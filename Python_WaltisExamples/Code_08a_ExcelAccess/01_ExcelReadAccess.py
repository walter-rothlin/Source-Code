
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('G:\_WaltisDaten\SourceCode\GitHosted\Python_WaltisExamples\Code_08a_ExcelAccess\TestPythonReadAccess.xlsx')

# Get sheet names
print(wb.sheetnames)

setUpSheet = wb['Set-Up']
aCell = setUpSheet['C6']
print('Content Set-Up.C6  :', aCell.value)
print('Row No.            :', aCell.row)
print('Column Letter      :', aCell.column)
print('Coordinates of cell:', aCell.coordinate)
print('Max row:',setUpSheet.max_row, '   Max col:', setUpSheet.max_column, '  Content:', wb['Set-Up'].cell(row=11, column=5).value)
print()
print('Content Set-Up.C6  :', wb['Set-Up']['C6'].value)
print('Content Set-Up.C6  :', wb['Set-Up'].cell(row=6, column=3).value)


wb['Set-Up']['A1'].value = 'Hello'
print('Content Set-Up.A1  :', wb['Set-Up']['A1'].value)