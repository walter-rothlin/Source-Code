#!/usr/bin/python3

# ------------------------------------------------------------------
# Name: FG_Abern_Einzug.py
# Source: https://raw.githubusercontent.com/walter-rothlin/Source-Code/master/Python_WaltisExamplesQR_Codes_QR_Rechnung/QR_Rechnung_Tobias/FG_Abern_Einzug.py
#
# Description: Produces QR-Rechnungen for Flurgenossenschaft Abern
#
# Autor: Walter Rothlin
#
# History:
# 17-Jun_2022   Walter Rothlin      Initial Version
# ------------------------------------------------------------------
from SwissQRInvoiceGenerator import *
from openpyxl import load_workbook
from waltisLibrary import *


if __name__ == '__main__':
    jahr = "2022"
    baseDestDir = 'N:\\06_FG_Abern\\' + jahr + '\\PerimeterRechnungen\\'
    baseSourceDir = 'N:\\06_FG_Abern\\'

    sourceDate = 'FG_Aberen_Kassier.xlsx'

    addr_FG_Aberen = {
        "name": "Flurgenossenschaft",
        "address": "'Aberen'",
        "zip_code": "8858",
        "city": "Innerthal",
        "country": "",
    }

    ibanNr = {
        "FG_Abern": "CH90 0077 7003 2922 1166 7"
    }

    # Load in the workbook (data_only=True returns values not formulas)
    wb = load_workbook(baseSourceDir + sourceDate,  data_only=True)
    print("Sheets in ", sourceDate, ":", wb.sheetnames, "\n\n")
    dataSheet = wb['Perimeter']

    # Strassenbenützung
    addr_strassenbenutzung = {
        "name": dataSheet['B17'].value + " " + dataSheet['C17'].value,
        "address": dataSheet['D17'].value,
        "zip_code": str(dataSheet['E17'].value),
        "city": dataSheet['F17'].value,
        "country": "",
    }

    pdfName = baseDestDir + "FG_Aberen_Strassenbenutzung 1"
    htmlName = baseDestDir + "FG_Aberen_Strassenbenutzung 1" + ".html"

    createQRInvoice(
        generateQRInvoiceData(
            ibanNr["FG_Abern"],
            addr_FG_Aberen,
            addr_strassenbenutzung,
            amount="50.00",
            additional_information="Strassenbeutzung 2022"),
        invoice_text_html="",
        pdfName=pdfName,
        htmlName=htmlName)

    print("Produced following files: ", pdfName)


    doLoop = True
    row = 2
    while doLoop:
        anschrift = dataSheet['A' + str(row)].value
        # print("==> ", anschrift, ";", type(anschrift), end="\n")

        vorname   = dataSheet['B' + str(row)].value
        if vorname is None:
            vorname = ""
        # print(";", vorname, ";", type(vorname), end="\n")

        name      = dataSheet['C' + str(row)].value
        # print(";", name, ";", type(name), end="\n")

        strasse   = dataSheet['D' + str(row)].value
        if strasse is None:
            strasse = ""
        # print(";", strasse, ";", type(strasse), end="\n")

        plz       = str(dataSheet['E' + str(row)].value)
        # print(";", plz, ";", type(plz), end="\n")

        ort       = dataSheet['F' + str(row)].value
        # print(";", ort, ";", type(ort), end="\n")

        betrag    = dataSheet['N' + str(row)].value
        # print(";", betrag, ";", type(betrag))


        if anschrift is not None:
            initialen = ""
            if len(vorname) > 0:
                initialen = vorname[0]

            if len(anschrift) <= 4:
                anschrift += " " + initialen + "." + name
            # else:
            #    anschrift += " c/o " + initialen + "." + name

            addr_kreditor = {
                "name": anschrift,
                "address": strasse,
                "zip_code": plz,
                "city": ort,
                "country": "",
            }
            pdfName = baseDestDir + "FG_Aberen_" + str(row)
            htmlName = baseDestDir + "FG_Aberen" + str(row) + ".html"
            print(pdfName + ":    " + anschrift)
            createQRInvoice(
                generateQRInvoiceData(
                    ibanNr["FG_Abern"],
                    addr_FG_Aberen,
                    addr_kreditor,
                    amount=betrag,
                    additional_information=jahr),
                invoice_text_html="",
                pdfName=pdfName,
                htmlName=htmlName)
            row += 1
        else:
            doLoop = False

        if (os.path.exists(htmlName)):
            os.remove(htmlName)
