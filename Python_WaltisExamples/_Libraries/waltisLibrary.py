#!/usr/bin/python3

# ------------------------------------------------------------------
# Name  : waltisLibrary.py
# Source: https://raw.githubusercontent.com/walter-rothlin/Source-Code/master/Python_WaltisExamples/_Libraries/waltisLibrary.py
#
# Description: Library - Module
#
#
# Autor: Walter Rothlin
#
# History:
# 06-Dec-2017   Walter Rothlin      Initial Version (Eigene Funktions von umrechnen.py ausgelagert)
# 28-May-2019   Walter Rothlin      Added Primzahlen functions
# 07-Jun-2019   Walter Rothlin      Merged with littlePythonLib.py
# 09-Apr-2020   Walter Rothlin      Refactoring isPrimzahl()
# 28-Sep-2020   Walter Rothlin      Added equalsWithinTolerance
# 19-Nov-2020   Walter Rothlin      Added calcCircle mit None parameter
# 10-Dec-2020   Walter Rothlin      Added fakultaet
# 21-Jan-2021   Walter Rothlin      Added ASCII Fct
# 17-Feb-2021   Walter Rothlin      File Operationen implementiert
# 11-Mar-2021   Walter Rothlin      More automated testing
# 22-Apr-2021   Walter Rothlin      Added URL functions
# 30-Apr-2021   Walter Rothlin      Added average(list)
# 08-Jun-2021   Walter Rothlin      Added readInt, readFloat
# 17-Jun-2021   Walter Rothlin      Added Mitternachts-Formel
# 10-Oct-2021   Walter Rothlin      Added placer, underline
# 13-Oct-2021   Walter Rothlin      Added File_getFileContent
# 16-Oct-2021   Walter Rothlin      Added printProgressBar
# 10-Nov-2021   Walter Rothlin      Added getRange
# 11-Nov-2021   Walter Rothlin      Added RegEx
# 13-Nov-2021   Walter Rothlin      Added crypto functions (shiftChr(), chipher(), encrypt(), decrypt())
# 22-Nov-2021   Walter Rothlin      Added readFile
# 22-Nov-2021   Walter Rothlin      Added Search and XML fct: dictify, getFieldFromTelSearchXML
#                                         getResultsFromAdressSearch,
#                                         getResults_search_ch,
#                                         getResults_geoAdmin
# 04-Dec-2021   Walter Rothlin      Added getMenuStrFromList
# 06-Dec-2021   Walter Rothlin      Added getFilenameWithoutExtension, getFilenameExtension, addTimestampToFileName, getPath
# 15-Jan-2022   Walter Rothlin      Added File_remove
# 19-May-2022   Walter Rothlin      Added howManyDigitsAreInString
# 05-Sep-2022   Stefan Rueeger      Typo fixed
# 26-Sep-2022   Walter Rothlin      halt() returns string
# 31-Dec-2022   Walter Rothlin      Added generateQRInvoiceData
# 06-Apr-2022   Walter Rothlin      Added readFloat_00() guard with regex
# 22-Jun-2023   Walter Rothlin      Added read_boolean
# 13-Jul-2023   Walter Rothlin      Added split_adress_street_nr()
# 03-Sep-2023   Walter Rothlin      Added Resuable DB und Excel Functions
# 12-Sep-2023   Walter Rothlin      Added is_US_Date, is_EU_Date
# 23-Sep-2023   Walter Rothlin      Added format_IBAN
#                                         split_plz_ort()
# 04-Oct-2023   Walter Rothlin      Added format_float(a_float, vorkommastellen=0, nachkommastellen=2, do_grouping=True)
#                                         get_record_details_from_db(db, table_name, key_id=None, db_attributes_names=[], as_json=True, take_action=True, verbal=False)
# 16-Oct-2023   Walter Rothlin      Added remove_all_enum_value_in_set()
# 23-Dec-2023   Walter Rothlin      Added convert_str_to_int_float_str
# 10-Feb-2024   Walter Rothlin      Added read_text_file_into_list_of_lines, remove_empty_line
# 15-Feb-2024   Walter Rothlin      Added get_sql_datums_update_value
# 14-Mar-2024   Walter Rothlin      Changes for BZU
# 04-Apr-2024   Walter Rothlin      Added SQL Functions:
#                                           get_table_records
#                                           create_sql_stmt_from_rs
#                                           create_insert_data_stmt
#                                           File_create
# 18-Apr-2024   Walter Rothlin      Changed create_insert_data_stmt()
#                                           getMenuStrFromList()
#                                           getValueFromDictList()
# 09-May-2024   Walter Rothlin      Added get_all_table_names_from_schema()
# 04-Jun-2024   Walter Rothlin      Added new version of create_insert_data_stmt()
# 06-Jun-2024   Walter Rothlin      Added verbal to File_create()
# 07-Jun-2024   Walter Rothlin      Added hashing in create_insert_data_stmt()
#                                   Added hash_string()
#                                   Added is_email()
#                                   Fixed problem with ' in create_insert_data_stmt()
# 07-Jul-2024   Walter Rothlin      Added DB unload/load-functions
#                                   Fixed issue in addTimestampToFileName() if it has more than one . in the name
#                                   Fixed issue get_exception_for_table_unload with empty exception lists
# 11-Jul-2024   Walter Rothlin      Fixed issues in unload_all_data_from_schema() and select_data_from_db_table()
#                                   Added format_sql_stmt(sql_statement, indent=4)
#                                   Removed double if verbal
# 20-Oct-2024   Walter Rothlin      Added PAIN001-Functions: get_pain001_template(), get_pain001_msg()
# 20-Oct-2024   Walter Rothlin      Fixed issue with datetime.datetime
#                                   Added get_actual_year()
#                                   Added read_string_with_default()
# 21-Oct-2024   Walter Rothlin      Added format_float_1()
# 26-Oct-2024   Walter Rothlin      Added get_possible_enum_values_from_db_attribute()
# 30-Oct-2024   Walter Rothlin      Changed datetime in convert_resultSet_to_insertSQL()
# 14-Nov-2024   Walter Rothlin      Changed get_pain001_msg() to use ISO-20022 Template
# 16-Nov-2024   Walter Rothlin      Fixed get_pain001_msg() to use proper ISO-20022 Template (Imotop)
# 19-Nov-2024   Walter Rothlin      Added get_link_to_map_for_adress()
#                                   Added get_coordinates_for_adresse()
# 15-Dec-2024   Walter Rothlin      Added get_sql_string_update_value(), get_sql_float_update_value()
# 18-Jan-2025   Walter Rothlin      Added search_address_geo_admin()
# 09-May-2025   Walter Rothlin      Added is_password_valid()
# ------------------------------------------------------------------

# toDo:
#  def File_readWithInludes returns noting

# ------------------------------------------------------------------

import inspect
import math
import os
import shutil
import time
from datetime import *
from pathlib import Path
import re
from rich import print
import openpyxl
import mysql.connector
import json
from lxml import etree
import xml.etree.ElementTree as ET
import urllib.parse
import locale
import hashlib
from jinja2 import Environment, FileSystemLoader, Template
import string

# Add dir to PYTHONPATH in a program
# ----------------------------------
# import sys
# sys.path.append('G:\_WaltisDaten\SourceCode\GitHosted\Python_WaltisExamples\_Libraries')
# https://www.datacamp.com/community/tutorials/modules-in-python?utm_source=adwords_ppc&utm_campaignid=898687156&utm_adgroupid=48947256715&utm_device=c&utm_keyword=&utm_matchtype=b&utm_network=g&utm_adpostion=&utm_creative=332602034343&utm_targetid=aud-299261629574:dsa-429603003980&utm_loc_interest_ms=&utm_loc_physical_ms=1030659&gclid=CjwKCAjw8MD7BRArEiwAGZsrBdWiIwMkjQ8l9_Gm52f1FEpiFVHlAgnMVf7UPzAtdNaeC7zq0691oRoC2D4QAvD_BwE#import

# Add dir to PYTHONPATH in PYCHARM
# --------------------------------
# Click an directory in project folders. Than Rigth Click in Project-Explorer  --> Mark Directory as --> Source Root
# https://intellij-support.jetbrains.com/hc/en-us/community/posts/115000782850-How-to-set-up-working-directory-in-PyCharm-and-package-import-

# Library fucntions
# =================
import requests
import json


def waltisPythonLib_Version():
    print("waltisLibrary.py: 2.0.1.2")


# Regular-Expressions
# ===================
regEx_email = r'([\w\.-]+)@([\w\.-]+)'
regEx_email_1 = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
regEx_Float = r'[+-]?\d*\.[0-9]+'
regEx_Int = r'[+-]?[0-9]'
regEx_Float_Or_Int = r'[+-]?\d*\.?\d+'
regEx_Date_US = r'[0-9]{4}.[0-9]{2}.[0-9]{2}'  # 1964-11-29
regEx_Date_EU = r'[0-9]{2}.[0-9]{2}.[0-9]{4}'  # 29.11.1964

# Konstanten
# ==========
pi = math.pi
halbBogen = 180


# Bildschirmsteuerung
# ===================
def VT52_cls():
    print("\033[2J", end="", flush=True)


def VT52_home():
    print("\033[H", end="", flush=True)


def VT52_cls_home():
    VT52_cls()
    VT52_home()


def halt(prompt="Weiter?"):
    return input(prompt)


# Print iterations progress
def printProgressBar(iteration, total, prefix='', suffix='', decimals=1, length=100, fill='█', printEnd="\r"):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end=printEnd)
    # Print New Line on Complete
    if iteration == total:
        print()


def TEST_printProgressBar(verbal=False):
    items = list(range(0, 57))
    l = len(items)

    # Initial call to print 0% progress
    printProgressBar(0, l, prefix='Progress:', suffix='Complete', length=50)
    for i, item in enumerate(items):
        # Do stuff...
        time.sleep(0.3)
        # Update Progress Bar
        printProgressBar(i + 1, l, prefix='Progress:', suffix='Complete', length=50)
    print("Completed")

# =================================
# Password functions and tests
# =================================
def is_password_valid(

        password,
        min_length=None,
        min_count_figures=None,
        min_count_uppercase=None,
        min_count_lowercase=None,
        min_count_special_char=None
):
    '''
# ----------------------------------------------
# Passwortregeln:
# ----------------------------------------------
# Ein korrektes Passwort muss je nach Vorgaben:
# - eine Mindestlänge haben
# - eine Mindestanzahl an Ziffern enthalten
# - eine Mindestanzahl an Großbuchstaben enthalten
# - eine Mindestanzahl an Kleinbuchstaben enthalten
# - eine Mindestanzahl an Sonderzeichen enthalten
# ----------------------------------------------
    '''
    if not isinstance(password, str):
        return False

    count_digits = sum(1 for c in password if c.isdigit())
    count_upper = sum(1 for c in password if c.isupper())
    count_lower = sum(1 for c in password if c.islower())
    count_special = sum(1 for c in password if c in string.punctuation)

    if min_length is not None and len(password) < min_length:
        return False
    if min_count_figures is not None and count_digits < min_count_figures:
        return False
    if min_count_uppercase is not None and count_upper < min_count_uppercase:
        return False
    if min_count_lowercase is not None and count_lower < min_count_lowercase:
        return False
    if min_count_special_char is not None and count_special < min_count_special_char:
        return False

    return True

# ---------------------------------------------------
# Testfunktion mit verschiedenen Test-Cases
# ---------------------------------------------------
def AUTO_TEST_is_password_valid():
    test_cases = [
        {
            "password": "Passw0rd",
            "rules": {
                "min_length": 8,
                "min_count_figures": 1,
                "min_count_uppercase": 1,
                "min_count_lowercase": 1,
                "min_count_special_char": 0,
            },
            "expected": True,
        },
        {
            "password": "short7!",
            "rules": {
                "min_length": 8,
                "min_count_figures": 1,
                "min_count_uppercase": 1,
                "min_count_lowercase": 1,
                "min_count_special_char": 1,
            },
            "expected": False,
        },
        {
            "password": "password1!",
            "rules": {
                "min_length": 8,
                "min_count_figures": 1,
                "min_count_uppercase": 1,
                "min_count_lowercase": 1,
                "min_count_special_char": 1,
            },
            "expected": False,
        },
        {
            "password": "ValidPassword123!",
            "rules": {
                "min_length": 10,
                "min_count_figures": 2,
                "min_count_uppercase": 1,
                "min_count_lowercase": 3,
                "min_count_special_char": 1,
            },
            "expected": True,
        },
    ]

    for i, test in enumerate(test_cases):
        result = is_password_valid(test["password"], **test["rules"])
        passed = result == test["expected"]
        print(f"Testfall {i + 1}: {'✅ OK' if passed else '❌ FEHLER'}")


# =================================
# verification functions for readln
# =================================
def is_email(email):
    return re.match(regEx_email_1, email) is not None


# ================
# Readln functions
# ================
def read_string_with_default(prompt="Input:", default_value=None):
    if default_value is None:
        in_str = input(prompt + ":")
    else:
        in_str = input(f'{prompt} ({default_value}):')
        if len(in_str) <= 0:
            in_str = default_value
    return in_str


def read_boolean(prompt="Boolean[Y/*N]:", true_val_liste=['Y', 'J', 'T'], default_value=False, verbal=False):
    ret_val = input(prompt).upper()
    if ret_val == '':
        ret_val = default_value

    elif ret_val[0] in true_val_liste:
        ret_val = True

    else:
        ret_val = default_value

    return ret_val


def TEST_readln(verbal=False):
    print("Test readInt:", end="")
    i1 = readInt("   readInt=")
    print(i1)

    print("\nTest readFloat:", end="")
    f1 = readFloat("    readFloat=")
    print(f1)

    print("\n1) Test read_Number:", end="")
    i1 = read_Number("int", prompt="   int:", postErrorStr=" Must be a {t:1s}!!!!")
    print(i1)

    print("\n2) Test read_Number:", end="")
    f1 = read_Number("float")
    print(f1)

    print("\n3) Test read_StrongType:", end="")
    i1 = read_Number("int", prompt="   int [0..100]:", postErrorStr=" Must be a {t:1s}!!!!", min=0, max=100)
    print(i1)

    print("\n4) Test read_Number:", end="")
    i1 = read_Number("int", prompt="   int >= -10:", postErrorStr=" Must be a {t:1s}!!!!", min=-10)
    print(i1)

    print("\nTest readFloat:", end="")
    f1 = readFloat(prompt="    float <= 150:", postErrorStr=" Must be a {t:1s}!!!!", max=150)
    print(f1)


def read_Number(type, prompt="Input [{t:1s}{lh:s}]:", preErrorStr="Wrong Format:", postErrorStr="   Must be a {t:1s}!",
                min=None, minErrorStr="Value must be greater or equal than {mi:1d}",
                max=None, maxErrorStr="Value must less or equal than {ma:1d}"):
    error = True
    userInputZahl = 0
    while error:
        try:
            if min is not None and max is not None:
                aString = input(prompt.format(t=type, lh="  " + str(min) + ".." + str(max)))
            elif min is not None:
                aString = input(prompt.format(t=type, lh="  " + str(min) + ".."))
            elif max is not None:
                aString = input(prompt.format(t=type, lh="  " + ".." + str(max)))
            else:
                aString = input(prompt.format(t=type, lh=""))
            if type == "int":
                userInputZahl = int(aString)
            elif type == "float":
                userInputZahl = float(aString)
            else:
                print("Unknown Type")
            error = False
            if (min is None) and (max is None):
                error = False
            else:
                if min is not None:
                    if userInputZahl < min:
                        print(minErrorStr.format(mi=min))
                        error = True
                if max is not None:
                    if userInputZahl > max:
                        print(maxErrorStr.format(ma=max))
                        error = True
        except ValueError:
            print(preErrorStr + aString + "    " + postErrorStr.format(t=type))
            error = True
    return userInputZahl


def readInt_0(prompt="Input [Int]:", preError="Wrong Format:", postError="   Must be a INT!"):
    error = True
    userInputStr = ""
    while error:
        try:
            # userInputInt = int(input(prompt))  # User Eingabe kann bei Fehlermeldung nicht reflektiert werden
            userInputStr = input(prompt)
            userInputInt = int(userInputStr)
            error = False
        except ValueError:
            print(preError + userInputStr + postError)
            error = True
    return userInputInt


def readInt_00(prompt="Input [Int]:", preError="Wrong Format:", postError="   Must be a INT!"):
    error = True
    userInputStr = ""
    while error:
        try:
            userInputInt = int(input(prompt))
            error = False
        except ValueError:
            print(preError + userInputStr + postError)
            error = True
    return userInputInt


def convert_str_to_int(int_str, default_value=None):
    try:
        int_str = int_str.replace(' ', '')
        int_str = int_str.replace("'", '')
        ret_value = math.floor(float(int_str))
    except Exception:
        ret_value = default_value
    return ret_value


def convert_str_to_int_float_str(a_string_to_cast, default=''):
    if a_string_to_cast == '':
        a_string_to_cast = default

    if isinstance(a_string_to_cast, str) and (a_string_to_cast.startswith('"') or a_string_to_cast.startswith("'")):
        a_string_to_cast = a_string_to_cast.replace('"', '')
        a_string_to_cast = a_string_to_cast.replace("'", '')
        ret_val = a_string_to_cast
    else:
        try:
            ret_val = int(a_string_to_cast)
        except Exception as e:
            try:
                ret_val = float(a_string_to_cast)
            except Exception as e:
                try:
                    ret_val = bool(a_string_to_cast)
                except Exception as e:
                    ret_val = a_string_to_cast
    return ret_val


def readFloat_0(prompt="float=", errPreMsg="Wrong input:", errPostMsg="   Must be a float!"):
    error = True
    while error:
        try:
            userInputStr = input(prompt)
            userInputStr = userInputStr.replace(',', '.')
            userInputStr = userInputStr.replace(' ', '')
            userInputStr = userInputStr.replace("'", '')
            userInputFloat = float(userInputStr)
            error = False
        except ValueError:
            print(errPreMsg + userInputStr + errPostMsg)
            error = True
    return userInputFloat


def readFloat_00(prompt="float=", errPreMsg="Wrong input:", errPostMsg="   Must be a float!"):
    error = True
    while error:
        userInputStr = input(prompt)
        userInputStr = userInputStr.replace(',', '.')
        userInputStr = userInputStr.replace(' ', '')
        userInputStr = userInputStr.replace("'", '')
        if re.fullmatch(regEx_Float_Or_Int, userInputStr):
            userInputFloat = float(userInputStr)
            error = False
        else:
            print(errPreMsg + userInputStr + errPostMsg)
            error = True
    return userInputFloat


def readFloat_1(prompt="Input [Float]:",
                preErrorStr="Wrong Format:",
                postErrorStr="   Must be an FLOAT!!!!",
                min=None, minErrorStr="Value must be greater or equal than {mi:1d}",
                max=None, maxErrorStr="Value must less or equal than {ma:1d}"):
    error = True
    userInputStr = ""
    while error:
        try:
            userInputStr = input(prompt)
            userInputStr = userInputStr.replace(',', '.')
            userInputStr = userInputStr.replace(' ', '')
            userInputStr = userInputStr.replace("'", '')
            userInputZahl = float(userInputStr)
            error = False
            if (min is None) and (max is None):
                error = False
            else:
                if min is not None:
                    if userInputZahl < min:
                        print(minErrorStr.format(mi=min))
                        error = True
                if max is not None:
                    if userInputZahl > max:
                        print(maxErrorStr.format(ma=max))
                        error = True
        except Exception:
            print(preErrorStr + userInputStr + postErrorStr)
            error = True
    return userInputZahl


def readInt_1(prompt="Input [Int]:",
              preErrorStr="Wrong Format:",
              postErrorStr="   Must be an Int!!!!",
              min=None, minErrorStr="Value must be greater or equal than {mi:1d}",
              max=None, maxErrorStr="Value must less or equal than {ma:1d}"):
    error = True
    userInputStr = ""
    while error:
        try:
            userInputStr = input(prompt)
            userInputStr = userInputStr.replace(',', '.')
            userInputStr = userInputStr.replace(' ', '')
            userInputStr = userInputStr.replace("'", '')
            userInputZahl = int(userInputStr)
            error = False
            if (min is None) and (max is None):
                error = False
            else:
                if min is not None:
                    if userInputZahl < min:
                        print(minErrorStr.format(mi=min))
                        error = True
                if max is not None:
                    if userInputZahl > max:
                        print(maxErrorStr.format(ma=max))
                        error = True
        except Exception:
            print(preErrorStr + userInputStr + postErrorStr)
            error = True
    return userInputZahl


def readFloat(prompt="Input [{t:1s}{lh:s}]:", preErrorStr="Wrong Format:", postErrorStr="   Must be a {t:1s}!",
              min=None, minErrorStr="Value must be greater or equal than {mi:1d}",
              max=None, maxErrorStr="Value must less or equal than {ma:1d}"):
    return read_Number("float", prompt=prompt, preErrorStr=preErrorStr, postErrorStr=postErrorStr,
                       min=min, minErrorStr=minErrorStr,
                       max=max, maxErrorStr=maxErrorStr)


def readInt(prompt="Input [{t:1s}{lh:s}]:", preErrorStr="Wrong Format:", postErrorStr="   Must be a {t:1s}!",
            min=None, minErrorStr="Value must be greater or equal than {mi:1d}",
            max=None, maxErrorStr="Value must less or equal than {ma:1d}"):
    return read_Number("int", prompt=prompt, preErrorStr=preErrorStr, postErrorStr=postErrorStr,
                       min=min, minErrorStr=minErrorStr,
                       max=max, maxErrorStr=maxErrorStr)


# ==========================
# Physikalische Umrechnungen
# ==========================
def grad2Rad(grad):
    return math.pi * grad / 180


def rad2Grad(rad):
    return 180 * rad / math.pi


def fahrenheit2Celsius(fahrenheit):
    return (fahrenheit - 32) / 1.8


def celsius2Fahrenheit(celsius):
    return (celsius * 1.8) + 32


def AUTO_TEST_pysikalische_umrechnungen_1(verbal=False):
    print(celsius2Fahrenheit(37.8))  # --> 100
    print(celsius2Fahrenheit(0))  # --> 32

    print(fahrenheit2Celsius(100))  # --> 37.8
    print(fahrenheit2Celsius(32))  # --> 0


def AUTO_TEST_pysikalische_umrechnungen_1a(verbal=False):
    print("celsius2Fahrenheit(37.8)) = ", celsius2Fahrenheit(37.8), "   Expected: 100")
    print("celsius2Fahrenheit(0))    = ", celsius2Fahrenheit(0), "   Expected: 32")

    print("fahrenheit2Celsius(100))  = ", fahrenheit2Celsius(100), "    Expected: 37.8")
    print("fahrenheit2Celsius(32))   = ", fahrenheit2Celsius(32), "    Expected: 32")


def AUTO_TEST_pysikalische_umrechnungen_2(verbal=False):
    inVal = 37.8
    expected = 100.04
    print(round(celsius2Fahrenheit(inVal), 2) == expected)

    inVal = 0
    expected = 32
    print(round(celsius2Fahrenheit(inVal), 2) == expected)

    inVal = 100
    expected = 37.8
    print(round(fahrenheit2Celsius(inVal), 2) == expected)

    inVal = 32
    expected = 0
    print(round(fahrenheit2Celsius(inVal), 2) == expected)


def AUTO_TEST_pysikalische_umrechnungen_3(verbal=False):
    """
    Alle AUTO_TEST_ Funktionen sind zum automatischen Testen von Funktionen. Dabei werden Testfälle ausgeführt und
    das Resultat direkt mit dem erwarteten Resultat verglichen. Falls das Resultat eines Funktionsaufrufes
    nicht mit dem erwarteten übereinstimmen, ist dieser Test-Case durchgefallen oder "failed".

    Bei jedem durchgefallenen Test-Case muss ich folgende Analyse machen:
    •	Erwartungswert richtig?
    o	Ja: Fehler in der Implementation der Funktion fixen und Test wiederholen
    o	Nein: Erwartungswert anpassen und Test wiederholen

    In AUTO_TEST_pysikalische_umrechnungen_3:
    •	celsius2Fahrenheit() mit 2 Test-Cases
    •	fahrenheit2Celsius() mit 2 Test-Cases
    """

    inVal = 37.8
    expected = 100.05
    if round(celsius2Fahrenheit(inVal), 2) != expected:
        print("ERROR (1):", end="")
        print("celsiusToFahrenheit(", inVal, ") = ", round(celsius2Fahrenheit(inVal), 2), "   Expected:", expected,
              sep="")

    inVal = 0
    expected = 32.0
    if round(celsius2Fahrenheit(inVal), 2) != expected:
        print("ERROR (2):", end="")
        print("celsiusToFahrenheit(", inVal, ") = ", round(celsius2Fahrenheit(inVal), 2), "   Expected:", expected,
              sep="")

    inVal = 100
    expected = 37.78
    if round(fahrenheit2Celsius(inVal), 2) != expected:
        print("ERROR (3):", end="")
        print("celsiusToFahrenheit(", inVal, ") = ", round(fahrenheit2Celsius(inVal), 2), "   Expected:", expected,
              sep="")

    inVal = 32
    expected = 0.1
    if round(fahrenheit2Celsius(inVal), 2) != expected:
        print("ERROR (4):", end="")
        print("celsiusToFahrenheit(", inVal, ") = ", round(fahrenheit2Celsius(inVal), 2), "   Expected:", expected,
              sep="")


def AUTO_TEST_pysikalische_umrechnungen_4(verbal=False):
    testSuite = "pysikalische_umrechnungen"
    testCases = """
    Nr    |Fct                   |Param1       |Param2       |Param3       |Expected
         1|celsius2Fahrenheit    |37.8         |0            |0            |100.04
         2|celsius2Fahrenheit    |0            |0            |0            |32.0
         3|fahrenheit2Celsius    |100          |0            |0            |37.78
         4|rad2Grad              |0            |0            |0            |0
    """

    testsPerformed = 0
    testsFailed = 0

    listOfTestCases = testCases.split("\n")
    for aTestCase in listOfTestCases[2:-1]:
        testsPerformed += 1
        listOfTestValues = aTestCase.split("|")
        case = listOfTestValues[0].strip()
        fct = listOfTestValues[1].strip()
        param_1 = float(listOfTestValues[2].strip())
        param_2 = float(listOfTestValues[3].strip())
        param_3 = float(listOfTestValues[4].strip())
        expectedResult = float(listOfTestValues[5].strip())
        # print(case, fct, param_1, param_2, param_3, expectedResult)
        if fct == "celsius2Fahrenheit":
            result = round(celsius2Fahrenheit(param_1), 2)
        elif fct == "fahrenheit2Celsius":
            result = round(fahrenheit2Celsius(param_1), 2)
        if result != expectedResult:
            testsFailed += 1
            print("Error in ", fct, "  ->", case, " (", testsPerformed, ")", sep="")
            print("   celsius2Fahrenheit(", param_1, ") = ", result, "    Expected:", expectedResult, sep="")
            print()
    print("\n")
    print("==> ", testSuite, ": Tests Performed:", testsPerformed, "   Tests Failed:", testsFailed, "    Passed:",
          round(100 - (100 * testsFailed / testsPerformed), 1), "%", sep="")


def AUTO_TEST_pysikalische_umrechnungen_5(verbal=False):
    testSuite = getMyFctName()[auto_test_fct_prefix_len:]
    testCases = """
    Nr    |Fct                   |Param1       |Param2       |Param3       |Expected
         1|celsius2Fahrenheit    |37.8         |0            |0            |100.04
         2|celsius2Fahrenheit    |0            |0            |0            |32.0
         3|fahrenheit2Celsius    |100          |0            |0            |37.78
         4|fahrenheit2Celsius    |32           |0            |0            |0
         5|rad2Grad              |3.141562     |0            |0            |180
         6|grad2Rad              |180          |0            |0            |3.14
    """

    testsPerformed = 0
    testsFailed = 0
    listOfTestCases = testCases.split("\n")
    for aTestCase in listOfTestCases[2:-1]:
        testsPerformed += 1
        listOfTestValues = aTestCase.split("|")
        case = listOfTestValues[0].strip()
        fct = listOfTestValues[1].strip()
        param_1 = float(listOfTestValues[2].strip())

        expectedResult = float(listOfTestValues[5].strip())
        # print(case, fct, param_1, param_2, param_3, expectedResult)
        possibles = globals().copy()
        possibles.update(locals())
        method = possibles.get(fct)
        result = None
        if method:
            result = round(method(param_1), 2)
        else:
            testsFailed += 1
            print("Error in testSuite:", testSuite, "   Function:", fct, "    case:(", case, ")",
                  "   Fct to test not found!!!", sep="")
        if result != expectedResult:
            testsFailed += 1
            print("Error in testSuite:", testSuite, "   Function:", fct, "    case:(", case, ")", sep="")
            print("   ", fct, "(", param_1, ") = ", result, "    Expected:", expectedResult, sep="")
            print()
    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


def AUTO_TEST_pysikalische_umrechnungen(verbal=False):
    # print(unterstreichen("AUTO_TEST_pysikalische_umrechnungen", "="))
    # print(unterstreichen("AUTO_TEST_pysikalische_umrechnungen_1", "-"))
    # AUTO_TEST_pysikalische_umrechnungen_1(verbal)

    # print()
    # print(unterstreichen("AUTO_TEST_pysikalische_umrechnungen_1a", "-"))
    # AUTO_TEST_pysikalische_umrechnungen_1a(verbal)

    # print()
    # print(unterstreichen("AUTO_TEST_pysikalische_umrechnungen_2", "-"))
    # AUTO_TEST_pysikalische_umrechnungen_2(verbal)

    # print()
    # print(unterstreichen("AUTO_TEST_pysikalische_umrechnungen_3", "-"))
    # AUTO_TEST_pysikalische_umrechnungen_3(verbal)

    # print()
    # print(unterstreichen("AUTO_TEST_pysikalische_umrechnungen_4", "-"))
    # AUTO_TEST_pysikalische_umrechnungen_4(verbal)

    # print()
    # print(unterstreichen("AUTO_TEST_pysikalische_umrechnungen_5", "-"))
    return AUTO_TEST_pysikalische_umrechnungen_5(verbal)


# Summen Reihen-Functionen
# ------------------------
def summeBis_MitFormel(bis):
    # Spezifikation: 1+2+3+4+5+6+...+bis = (bis * (bis+1) /2)   Der Returnparameter ist ein int
    # mit Gausschen Summenformel
    #
    # Test-Cases
    # ----------
    # summeBis_MitFormel(9) = 45
    # summeBis_MitFormel(4) = 10
    # summeBis_MitFormel(20) = 210

    return int(bis * (bis + 1) / 2)


def summeBis_MitLoop(bis):
    # Spezifikation: 1+2+3+4+5+6+...+bis Der Returnparameter ist ein int
    # Informatiker Lösung (langsamer in der Ausführung)
    #
    # Test-Cases
    # ----------
    # summeBis_MitLoop(9) = 45
    # summeBis_MitLoop(4) = 10
    # summeBis_MitLoop(20) = 210

    res = 0
    for i in range(bis + 1):
        res += i
    return res


def summe(bis, von=None):
    # Spezifikation: von+..+5+6+...+bis
    #
    # Test-Cases
    # ----------
    # summe(9) = 45
    # summe(4) = 10
    # summe(20) = 210
    # summe(100) = 5050
    # summe(9, 4) = 39
    # summe(4, 9) = 39
    # summe(-1, 1) = 0

    if von is None:
        return int(bis * (bis + 1) / 2)
    else:
        if von > bis:
            tmp = bis
            bis = von
            von = tmp
        return int(bis * (bis + 1) / 2) - int(von * (von + 1) / 2) + von


def AUTO_TEST_a_summe(verbal=False):
    testsPerformed = 0
    testsFailed = 0
    testSuite = "a_summe"

    # Test-Cases
    # ----------
    # summe(9) = 45
    # summe(4) = 10
    # summe(20) = 210
    # summe(100) = 5050
    # summe(9, 4) = 39
    # summe(4, 9) = 39
    # summe(-1, 1) = 0
    testsPerformed += 1
    if summe(9) != 45:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summe(4) != 10:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summe(20) != 210:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summe(100) != 5050:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summe(9, 4) != 39:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summe(4, 9) != 39:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summe(-1, 1) != 0:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


def AUTO_TEST_a_summeBis_MitFormel(verbal=False):
    testsPerformed = 0
    testsFailed = 0
    testSuite = "a_summeBis_MitFormel"

    # Test-Cases
    # ----------
    # summeBis_MitFormel(9) = 45
    # summeBis_MitFormel(4) = 10
    # summeBis_MitFormel(20) = 210
    testsPerformed += 1
    if summeBis_MitFormel(9) != 45:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summeBis_MitFormel(4) != 10:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summeBis_MitFormel(20) != 210:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


def AUTO_TEST_a_summeBis_MitLoop(verbal=False):
    testsPerformed = 0
    testsFailed = 0
    testSuite = "a_summeBis_MitLoop"

    # Test-Cases
    # ----------
    # summeBis_MitLoop(9) = 45
    # summeBis_MitLoop(4) = 10
    # summeBis_MitLoop(20) = 210
    testsPerformed += 1
    if summeBis_MitLoop(9) != 45:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summeBis_MitLoop(4) != 10:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summeBis_MitLoop(20) != 210:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


def AUTO_TEST_a_summe(verbal=False):
    testsPerformed = 0
    testsFailed = 0
    testSuite = "a_summe"

    # Test-Cases
    # ----------
    # summe(9) = 45
    # summe(4) = 10
    # summe(20) = 210
    # summe(100) = 5050
    # summe(9, 4) = 39
    # summe(4, 9) = 39
    # summe(-1, 1) = 0
    testsPerformed += 1
    testsPerformed += 1
    if summe(9) != 45:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summe(4) != 10:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summe(20) != 210:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summe(100) != 5050:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summe(9, 4) != 39:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summe(4, 9) != 39:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if summe(-1, 1) != 0:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


def AUTO_TEST_summe(verbal=False):
    testsPerformed = 0
    testsFailed = 0

    testSuite = getMyFctName()[auto_test_fct_prefix_len:]
    testCases = """
    Nr    |Fct            |bis           |von           |Expected
    Type  |               |int           |int           |int
         1|summe          |9             |None          |45
         2|summe          |4             |None          |10
         3|summe          |20            |None          |210
         4|summe          |100           |None          |5050
         5|summe          |9             |4             |39
         6|summe          |4             |9             |39
         7|summe          |-1            |1             |0
    """
    listOfTestCases = testCases.split("\n")
    headStr = listOfTestCases[1]
    typeStr = listOfTestCases[2]
    ##### TBI print(headStr)
    ##### TBI print(typeStr)

    for aTestCase in listOfTestCases[3:-1]:
        testsPerformed += 1
        listOfTestValues = aTestCase.split("|")
        case = listOfTestValues[0].strip()
        fct = listOfTestValues[1].strip()
        param_1 = int(listOfTestValues[2].strip())
        param_2_str = listOfTestValues[3].strip()
        if param_2_str != 'None':
            param_2 = int(param_2_str)
        else:
            param_2 = 1  # default value parameter 2
        expectedResult = int(listOfTestValues[4].strip())
        # print(case, fct, param_1, param_2, expectedResult)

        possibles = globals().copy()
        possibles.update(locals())
        method = possibles.get(fct)
        result = None
        if method:
            result = method(bis=param_1, von=param_2)
        else:
            testsFailed += 1
            print("Error in testSuite:", testSuite, "   Function:", fct, "    case:(", case, ")",
                  "   Fct to test not found!!!", sep="")
        if result != expectedResult:
            testsFailed += 1
            print("Error in testSuite:", testSuite, "   Function:", fct, "    case:(", case, ")", sep="")
            print("   ", fct, "(", param_1, ",", param_2, ") = ", result, "    Expected:", expectedResult, sep="")
            print()
    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


# Durchschnitt berechnen
# ======================
def average(nrList, untergrenze=0, obergrenze=None, verbal=False, debug=False):
    '''
    Specification:
    This functions calculates the average of all number elements in a given List:
    e.g.
        average([4, 6, 2, 5.5])                           --> 4.375
        average([4, 6, 2, 5.5, 3])                        --> 4.1
        average(["4", "6", "2", "5.5", "1.5*2"])          --> 4.1
        average(["4", "6", "2", "5.5", "1.5*2", "walti"]) --> 4.1
    '''

    if obergrenze is None:
        nrList = nrList[untergrenze:]
    else:
        nrList = nrList[untergrenze:obergrenze]
    if debug:
        print(nrList)
    aSumme = 0
    aCount = 0
    for aElement in nrList:
        try:
            if (type(aElement) == int) or (type(aElement) == float):
                aNum = aElement
            else:
                aNum = eval(aElement)
        except NameError as error:
            if verbal:
                print("WARNING: Element", aElement, "is not taken!!", error)
        else:
            aSumme += aNum
            aCount += 1
    return aSumme / aCount


def AUTO_TEST_average(verbal=False):
    testsPerformed = 0
    testsFailed = 0
    testSuite = "average"

    testsPerformed += 1
    param_1 = [4, 6, 2, 5.5]
    expectedResult = 4.375
    result = average(param_1)
    if result != expectedResult:
        testsFailed += 1
        print("Error in testSuite:   ", testSuite, ":  -> (", testsPerformed, ")", sep="")
        print("   average(", param_1, ") = ", result, "    Expected:", expectedResult, sep="")
        print()

    testsPerformed += 1
    param_1 = [4, 6, 2, 5.5, 3]
    expectedResult = 4.1
    result = average(param_1)
    if result != expectedResult:
        testsFailed += 1
        print("Error in testSuite:   ", testSuite, ":  -> (", testsPerformed, ")", sep="")
        print("   average(", param_1, ") = ", result, "    Expected:", expectedResult, sep="")
        print()

    testsPerformed += 1
    param_1 = ["4", "6", "2", "5.5", "1.5*2", "walti"]
    expectedResult = 4.1
    result = average(param_1)
    if result != expectedResult:
        testsFailed += 1
        print("Error in testSuite:   ", testSuite, ":  -> (", testsPerformed, ")", sep="")
        print("   average(", param_1, ") = ", result, "    Expected:", expectedResult, sep="")
        print()

    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


# print(average([4, 6, 2, 5.5, 3]) == 4.1)
# print(average() == 4.1)

# Fakultät berechnen
# ==================
def fakultaet(obergrenze, untergrenze=1):
    # Spezifikation: untergrenze*..*5*6*...*obergrenze
    # Test-Cases
    # ----------
    # fakultaet(10, 1) = 3628800
    # fakultaet(15, 13) = 2730
    # fakultaet(8, 3) = 20160
    # fakultaet(99, 98) = 9702
    # fakultaet(3) = 6
    # fakultaet(5) = 120

    fakultaet = -1
    if (obergrenze > 0):
        fakultaet = 1
        while obergrenze >= untergrenze:
            fakultaet = fakultaet * obergrenze
            obergrenze = obergrenze - 1
    return fakultaet


def AUTO_TEST_fakultaet(verbal=False):
    testsPerformed = 0
    testsFailed = 0

    testSuite = getMyFctName()[auto_test_fct_prefix_len:]
    testCases = """
    Nr    |Fct                |obergrenze    |untergrenze   |Expected
    Type  |                   |int           |int           |
         1|fakultaet          |10            |1             |3628800
         2|fakultaet          |15            |13            |2730
         3|fakultaet          |8             |3             |20160
         4|fakultaet          |99            |98            |9702
         5|fakultaet          |3             |              |6
         6|fakultaet          |5             |              |120
    """

    listOfTestCases = testCases.split("\n")
    headStr = listOfTestCases[1]
    typeStr = listOfTestCases[2]
    ##### TBI print(headStr)
    ##### TBI print(typeStr)

    for aTestCase in listOfTestCases[3:-1]:
        testsPerformed += 1
        listOfTestValues = aTestCase.split("|")
        case = listOfTestValues[0].strip()
        fct = listOfTestValues[1].strip()
        param_1 = int(listOfTestValues[2].strip())
        param_2_str = listOfTestValues[3].strip()
        if param_2_str != '':
            param_2 = int(param_2_str)
        else:
            param_2 = 1  # default value parameter 2
        expectedResult = int(listOfTestValues[4].strip())
        # print(case, fct, param_1, param_2, expectedResult)

        possibles = globals().copy()
        possibles.update(locals())
        method = possibles.get(fct)
        result = None
        if method:
            result = method(param_1, param_2)
        else:
            testsFailed += 1
            print("Error in testSuite:", testSuite, "   Function:", fct, "    case:(", case, ")",
                  "   Fct to test not found!!!", sep="")
        if result != expectedResult:
            testsFailed += 1
            print("Error in testSuite:", testSuite, "   Function:", fct, "    case:(", case, ")", sep="")
            print("   ", fct, "(", param_1, ",", param_2, ") = ", result, "    Expected:", expectedResult, sep="")
            print()
    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


# Primzahlen Functions
# ====================
def isPrimzahl(aZahl):
    '''
    Returns True/False if a given integer is a prime figure (only dividable by 1 or itself.

            Parameters:
                    aZahl (int): A positive Integer

            Returns:
                    True / False
    '''
    isPrim = False
    if aZahl == 1 or aZahl == 2:
        isPrim = True
    else:
        isPrim = True
        obergrenze = int(aZahl / 2) + 1
        # print("Obergrenze:",obergrenze)
        for i in range(2, obergrenze + 1):
            # print("    Test: aZahl % i = ",aZahl,"%",i,"=",aZahl % i)
            if (aZahl % i) == 0:
                isPrim = False
    return isPrim


def isPrimzahl_0(n):
    if n == 1:
        isPrimzahl = False
    else:
        t = 2
        while True:
            if t >= n:
                isPrimzahl = True
                break
            if n % t == 0:
                isPrimzahl = False
                break
            t += 1
    return isPrimzahl


def isPrimzahl_2(eineZahl):
    ist_eine_Primzahl = True
    if eineZahl == 1:
        ist_eine_Primzahl = False
    if eineZahl == 2:
        ist_eine_Primzahl = True

    obergrenze = int(eineZahl / 2) + 1
    for i in range(2, obergrenze):
        if eineZahl % i == 0:
            ist_eine_Primzahl = False
            break
    return ist_eine_Primzahl


def getNextPrimzahl(zahl):
    if ((zahl % 2) == 0):  # then Gerade
        aZahl = zahl + 1
    else:
        aZahl = zahl + 2
    # print("getNextPrimzahl::Startwert:",aZahl)
    while (isPrimzahl(aZahl) == False):
        aZahl = aZahl + 2
    return aZahl


def getPrevPrimzahl(zahl):
    if ((zahl % 2) == 0):  # then Gerade
        aZahl = zahl - 1
    else:
        aZahl = zahl - 2
    if (aZahl <= 1):
        return 1
    else:
        # print("getPrevPrimzahl::Startwert:",aZahl)
        while isPrimzahl(aZahl) == False:
            aZahl = aZahl - 2
    return aZahl


def getPrimezahlenListeAsListe(start, end):
    ret_list = []
    for i in range(start, end + 1):
        if (isPrimzahl(i)):
            ret_list.append(i)
    return ret_list


def getPrimezahlenListeAsString(start, end, sep=";"):
    string_list = [str(num) for num in getPrimezahlenListeAsListe(start, end)]
    return sep.join(string_list)


def getPrimezahlenListe(start, end, sep=";"):
    return getPrimezahlenListeAsString(start, end, sep=";")


def getPrimfactorsAsList(zahl):
    ret_list = []
    aZahl = abs(zahl)
    aDivisor = 2
    if zahl == 1 or zahl == 2:
        ret_list = [zahl]
    else:
        while isPrimzahl(aZahl) == False:
            if (aZahl % aDivisor) == 0:
                if aZahl > 1:
                    ret_list.append(aDivisor)
                aZahl = int(aZahl / aDivisor)  # Ganzzahlige division
            else:
                aDivisor = getNextPrimzahl(aDivisor)
        if aZahl > 1:
            ret_list.append(aZahl)
    return ret_list


def getPrimfactorsAsString(zahl, sep=";"):
    string_list = [str(num) for num in getPrimfactorsAsList(zahl)]
    return sep.join(string_list)


def getPrimfactors(zahl, sep=";"):
    return getPrimfactorsAsString(zahl, sep=";")


def getDivisorsAsList(zahl):
    ret_list = []
    aZahl = abs(zahl)
    aDivisor = 2
    while (aDivisor < aZahl):
        if ((aZahl % aDivisor) == 0):
            ret_list.append(aDivisor)
        aDivisor = aDivisor + 1
    return ret_list


def getDivisorsAsString(zahl, sep=";"):
    string_list = [str(num) for num in getDivisorsAsList(zahl)]
    return sep.join(string_list)


def getDivisors(zahl, sep=";"):
    return getDivisorsAsString(zahl, sep=";")


# Test der Functions primezahlen
# ------------------------------
def TEST_Primzahlen():
    for i in range(100):
        if (isPrimzahl(i)):
            print("{z:3d}: Ist eine Primzahl!!!".format(z=i))
        else:
            print("{z:3d}: Primzahlen:{s:30s}      Teiler    :{s1:30s}".format(z=i, s=getPrimfactors(i),
                                                                               s1=getDivisors(i)))


# csv- or set-functions
# =====================
def remove_set_value(set_values, value_to_remove, verbal=True):
    if verbal:
        print(f"Calling remove_set_value({set_values}, {value_to_remove})")

    set_values.discard(value_to_remove)
    set_values_str = ""
    for a_set_value in set_values:
        print("Value:", a_set_value)
        set_values_str += "," + a_set_value
    set_values_str = set_values_str[1:]
    print('set_values_str:', set_values_str)

    # set_values_str = 'Pächter,Hat_35a,Hat_16a,Bürger'
    return set_values_str


# String Functions
# ================
def remove_empty_line(lines, comment_str='#'):
    ret_list = []
    for a_line in lines:
        a_line = a_line.rstrip()
        a_line = a_line.lstrip()
        if len(a_line) > 0 and not a_line.startswith(comment_str):
            ret_list.append(a_line)
    return ret_list


def format_float(a_float, field_size=0, nachkommastellen=2, do_grouping=True):
    locale.setlocale(locale.LC_ALL, '')
    return locale.format_string("%" + str(field_size) + "." + str(nachkommastellen) + "f", a_float, grouping=do_grouping)


def format_float_1(a_float, field_size=0, nachkommastellen=2, do_grouping=True, euro_style=True):
    formatStr = '.' + str(nachkommastellen) + 'f'
    if do_grouping:
        formatStr = ',' + formatStr
    float_formatted = '{a_float_val:' + formatStr + '}'
    float_formatted = float_formatted.format(a_float_val=a_float)

    if euro_style:
        float_str = float_formatted.replace(',', "'")

    if field_size > 0:
        formatStr = '{float_str:>' + str(field_size) + 's}'
        float_str = formatStr.format(float_str=float_str)

    return float_str


def TEST__format_float_1(verbal=False):
    print(f':1234567890123456789012345678900:')
    a_float_value = 4678.206
    print(f':{format_float(a_float_value, field_size=15, nachkommastellen=3, do_grouping=True)}:')
    print(f':{format_float(a_float_value, field_size=15, nachkommastellen=2, do_grouping=True)}:')
    print(f':{format_float(a_float_value, field_size=15, nachkommastellen=2, do_grouping=False)}:')

    print(f':{format_float_1(a_float_value, field_size=15, nachkommastellen=3, do_grouping=True)}:')
    print(f':{format_float_1(a_float_value, field_size=15, nachkommastellen=2, do_grouping=True)}:')
    print(f':{format_float_1(a_float_value, field_size=15, nachkommastellen=2, do_grouping=False)}:')
    print(f':{format_float_1(a_float_value, nachkommastellen=2, do_grouping=False)}:')
    print(f':{format_float_1(a_float_value, nachkommastellen=2, do_grouping=True)}:')
    print(f':{format_float_1(a_float_value, nachkommastellen=4, do_grouping=True)}:')


def AUTO_TEST__format_float(verbal=False):
    testsPerformed = 0
    testsFailed = 0
    testSuite = "format_float"

    testsPerformed += 1
    case = 1
    if format_float(1234.3) != "1’234.30":
        print("Error in testSuite:   ", testSuite, ":   ", testSuite, "    case:", case, sep="")
        print("    Result: ", format_float(1234.3), "   Expected:", "1’234.30", end="\n\n")
        testsFailed += 1

    testsPerformed += 1
    case = 2
    if format_float(1234.3, nachkommastellen=3) != "1’234.300":
        print("Error in testSuite:   ", testSuite, ":   ", testSuite, "    case:", case, sep="")
        print("    Result: ", format_float(1234.3, nachkommastellen=3), "   Expected:", "1’234.300", end="\n\n")
        testsFailed += 1

    testsPerformed += 1
    case = 3
    if format_float(1234.3, nachkommastellen=3, do_grouping=False) != "1234.300":
        print("Error in testSuite:   ", testSuite, ":   ", testSuite, "    case:", case, sep="")
        print(f"    Result:{format_float(1234.3, nachkommastellen=3, do_grouping=False)}:   Expected:1234.300:", end="\n\n")
        testsFailed += 1

    testsPerformed += 1
    case = 4
    if format_float(1234.3, field_size=10, nachkommastellen=3, do_grouping=False) != "  1234.300":
        print("Error in testSuite:   ", testSuite, ":   ", testSuite, "    case:", case, sep="")
        print(f"    Result:{format_float(1234.3, field_size=10, nachkommastellen=3, do_grouping=False)}:   Expected:  1234.300:", end="\n\n")
        testsFailed += 1

    testsPerformed += 1
    case = 5
    if format_float(1234.3, field_size=10, nachkommastellen=2, do_grouping=True) != "  1’234.30":
        print("Error in testSuite:   ", testSuite, ":   ", testSuite, "    case:", case, sep="")
        print(f"    Result:{format_float(1234.3, field_size=10, nachkommastellen=2, do_grouping=True)}:   Expected:  1’234.30:", end="\n\n")
        testsFailed += 1

    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")

    return [testsPerformed, testsFailed]


def split_adress_street_nr(street_nr_string, verbal=False):
    parts = street_nr_string.split(' ')
    if verbal:
        print(parts)
        print('strasse:', parts[0:-1], '   nr:', parts[-1])

    if len(parts) == 1:
        return {'Street': parts[0], 'No': ''}
    else:
        return {'Street': ' '.join(parts[0:-1]), 'No': parts[-1]}


def AUTO_TEST__split_adress_street_nr():
    test_cases = ['Etzelstrasse 7',
                  'Rue de solei 788ab',
                  "hans' house 76",
                  'Bahnhostrasse']
    for a_test_case in test_cases:
        strasse_nr = split_adress_street_nr(a_test_case, verbal=False)
        print('Strasse:', strasse_nr['Street'], '  Nr:', strasse_nr['No'])


def split_plz_ort(plz_ort_string, verbal=False):
    parts = plz_ort_string.split(' ')
    if verbal:
        print(parts)
    return {'PLZ': parts[0], 'Ort': ' '.join(parts[1:])}


def split_familien_name(familien_name, sex='Herr', verbal=False):
    parts = familien_name.split('-')
    if verbal:
        print(parts, '   Sex:', sex)

    ret_val = {}
    if len(parts) == 1:
        parts = familien_name.split(' ')
        if len(parts) == 1:
            ret_val = {'Ledig_Name': parts[0], 'Partner_Name': '', 'Partner_Name_Angenommen': 0}
        else:
            if sex == 'Herr':
                ret_val = {'Ledig_Name': parts[1], 'Partner_Name': parts[0], 'Partner_Name_Angenommen': 1}
            else:
                ret_val = {'Ledig_Name': parts[0], 'Partner_Name': parts[1], 'Partner_Name_Angenommen': 0}
    else:
        if sex == 'Herr':
            ret_val = {'Ledig_Name': parts[0], 'Partner_Name': parts[1], 'Partner_Name_Angenommen': 0}
        else:
            ret_val = {'Ledig_Name': parts[1], 'Partner_Name': parts[0], 'Partner_Name_Angenommen': 1}

    return ret_val


def howManyDigitsAreInString_Classic(aString, trace=False):
    count = 0
    for c in aString:
        if c >= '0' and c <= '9':
            count += 1
    if trace:
        print("howManyDigitsAreInString_Classic:", aString, count)
    return count


def howManyDigitsAreInString_WithComprehension(aString, trace=False):
    count = len([d for d in aString if '0' <= d <= '9'])
    if trace:
        print("howManyDigitsAreInString_WithComprehension:", aString, "--> ", count)
    return count


def howManyDigitsAreInString_WithRegEx(aString, trace=False):
    import re

    copmiledRe = re.compile(r'\d')
    count = len(copmiledRe.findall(aString))
    if trace:
        print("howManyDigitsAreInString_WithRegEx:", aString, "--> ", count)
    return count


def howManyDigitsAreInString(aString, trace=False):
    # return howManyDigitsAreInString_Classic(aString, trace=trace)
    # return howManyDigitsAreInString_WithComprehension(aString, trace=trace)
    return howManyDigitsAreInString_WithRegEx(aString, trace=trace)


def addParity(binStr, oddParity=True):
    anzOne = 0
    for aBit in binStr:
        if aBit == "1":
            anzOne += 1
    if anzOne % 2 == 0:
        if oddParity:
            binStr = "1" + binStr
        else:
            binStr = "0" + binStr
    else:
        if oddParity:
            binStr = "0" + binStr
        else:
            binStr = "1" + binStr
    return binStr


def AUTO_TEST_addParity(verbal=False):
    testsPerformed = 0
    testsFailed = 0

    testSuite = getMyFctName()[auto_test_fct_prefix_len:]
    testCases = """
    Nr    |Fct                |Bitmuster    |Odd_Parity   |Param3       |Expected
         1|addParity          |1101110      |True         |0            |01101110
         2|addParity          |0101110      |True         |0            |10101110
         3|addParity          |110          |True         |0            |1110
         4|addParity          |1101110      |False        |0            |11101110
         5|addParity          |0101110      |False        |0            |00101110
         6|addParity          |110          |False        |0            |0110
         7|addParity          |110          |True         |0            |1110
         8|addParity          |111          |True         |0            |0111
    """

    listOfTestCases = testCases.split("\n")
    for aTestCase in listOfTestCases[2:-1]:
        testsPerformed += 1
        listOfTestValues = aTestCase.split("|")
        case = listOfTestValues[0].strip()
        fct = listOfTestValues[1].strip()
        param_1 = listOfTestValues[2].strip()
        param_2 = (listOfTestValues[3].strip() == "True")
        expectedResult = listOfTestValues[5].strip()
        # print(case, fct, param_1, param_2, expectedResult)

        possibles = globals().copy()
        possibles.update(locals())
        method = possibles.get(fct)
        result = None
        if method:
            result = method(param_1, param_2)
        else:
            testsFailed += 1
            print("Error in testSuite:   ", testSuite, ":   ", fct, "    case:()  ->", case, " (", testsPerformed,
                  ")   Fct to test not found!!!", sep="")
        if result != expectedResult:
            testsFailed += 1
            print("Error in testSuite:   ", testSuite, ":   ", fct, "    case:()  ->", case, " (", testsPerformed, ")",
                  sep="")
            print("   ", fct, "(", param_1, ",", param_2, ") = ", result, "    Expected:", expectedResult, sep="")
            print()
    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


def getValueFromDictList(menu_liste, key_name='nr', key_value='4', value_name='formel'):
    list_element = [an_element for an_element in menu_liste if an_element[key_name] == key_value]
    if list_element[0].get(value_name) is None:
        return ''
    else:
        return list_element[0].get(value_name)


def getMenuStrFromList(aList, titel="Menu-Items", indent="    ", itmeFormat="{i:2d}: {menuText:1s}", itemNrOffest=1, unterstreichen="-"):
    if aList is None:
        return 'None Menu'
    if len(aList) < 1:
        return 'Empty Menu'

    retStr = ''
    if isinstance(aList[0], str):
        retList = [(indent + itmeFormat).format(i=i + itemNrOffest, menuText=p) for i, p in zip(range(len(aList)), aList)]
        if unterstreichen is not None:
            retStr = indent + titel + "\n" + indent + unterstreichen * len(titel) + "\n" + "\n".join(retList)
        else:
            retStr = indent + titel + "\n" + "\n".join(retList)
    else:
        retStr = f'{indent}{titel}\n{indent}{unterstreichen * len(titel)}\n'
        for a_menu_item in aList:
            if a_menu_item['menu_str'].strip() == '':
                retStr += f"\n"
            else:
                retStr += f"{indent}{a_menu_item['nr']}: {a_menu_item['menu_str']}\n"
    return retStr


def TEST_getMenuFromList(verbal=False):
    menuList = ["Grad to Kelvin", "Kelvin to Grad", "Rad to Grad", "Grad to Rad"]
    print(getMenuStrFromList(menuList))


def showASCII_Table(firstVal=33, lastVal=126, sep=" --> ", end="\n"):
    print("{z:5s} {sep:1s} {ordDec:3s}     {ordHex:5s} {ordOct:6s} {ordBin:8s}   {odd:8s}   {even:8s}".format(z="ASCII",
                                                                                                              sep=sep,
                                                                                                              ordDec="Dez",
                                                                                                              ordHex="Hex",
                                                                                                              ordOct="Oct",
                                                                                                              ordBin="Bin",
                                                                                                              even="Even",
                                                                                                              odd="Odd"),
          sep="")
    for i in range(firstVal, lastVal + 1):
        print("{z:5s} {sep:1s} {ordDec:3d}     {ordHex:5s} {ordOct:6s} {ordBin:8s}".format(z=chr(i), sep=sep, ordDec=i,
                                                                                           ordHex=hex(i)[2:].upper(),
                                                                                           ordOct=oct(i)[
                                                                                                  2:].upper().rjust(3,
                                                                                                                    "0"),
                                                                                           ordBin=bin(i)[
                                                                                                  2:].upper().rjust(7,
                                                                                                                    "0")),
              "   ", addParity(bin(i)[2:].upper().rjust(7, "0")), "   ",
              addParity(bin(i)[2:].upper().rjust(7, "0"), False), sep="")


def hash_string(string, algorithm='md5', ret_length=None):
    # Create a hash object based on the specified algorithm
    if algorithm == 'md5':
        hash_object = hashlib.md5()
    elif algorithm == 'sha1':
        hash_object = hashlib.sha1()
    elif algorithm == 'sha256':
        hash_object = hashlib.sha256()
    else:
        raise ValueError("Unsupported algorithm. Use 'md5', 'sha1', or 'sha256'.")

    # Update the hash object with the bytes of the string
    hash_object.update(string.encode('utf-8'))

    # Get the hexadecimal representation of the hash
    hash_string = hash_object.hexdigest()
    if ret_length is not None:
        hash_string = hash_string[:ret_length]
    return hash_string


def left(s, amount):
    return s[:amount]


def right(s, amount):
    return s[-amount:]


def mid(s, offset, amount):
    return s[offset:offset + amount]


def toUpperCase(inString):
    retString = ""
    for aChar in inString:
        if (aChar >= 'a') and (aChar <= 'z'):
            retString = retString + chr(ord(aChar) - (ord('a') - ord('A')))
        else:
            retString = retString + aChar
    return retString


def toLowerCase(inString):
    retString = ""
    for aChar in inString:
        if (aChar >= 'A') and (aChar <= 'Z'):
            retString = retString + chr(ord(aChar) + (ord('a') - ord('A')))
        else:
            retString = retString + aChar
    return retString


def toFirstUpperCase(inString):
    firstPart = inString[:1]
    endPart = inString[1:]
    return toUpperCase(firstPart) + toLowerCase(endPart)


def generateStringRepeats(len, aStr=" "):
    # Spezifikation: Wiederholt den aStr so oft, dass der Return-String len lang ist
    # Test-Cases
    # ----------
    # generateStringRepeats(10, '-')     => '----------'
    # generateStringRepeats(9,  '+-=')   => '+-=+-=+-='
    # generateStringRepeats(3,  'A')     => 'AAA'
    # generateStringRepeats(5, '.-')     => '.-.-.'
    # generateStringRepeats(6)           => '      '
    return (aStr * len)[:len]


def getRange(startVal=0, endVal=8, inc=1):
    doLoop = True
    retList = []
    i = startVal
    while doLoop:
        if inc > 0:
            if i <= endVal:
                retList.append(i)
                i += inc
            else:
                doLoop = False
        elif inc < 0:
            if i >= endVal:
                retList.append(i)
                i += inc
            else:
                doLoop = False
        else:
            doLoop = False
    return retList


def AUTO_TEST_getRange(verbal=False):
    testsPerformed = 0
    testsFailed = 0
    testSuite = "getRange"

    testsPerformed += 1
    istList = getRange(3, 8)
    expList = [3, 4, 5, 6, 7, 8]
    if str(istList) != str(expList):
        testsFailed += 1
        print("Error in getRange")
        print("    Ist:", str(istList))
        print("    Exp:", str(expList))
        print()

    testsPerformed += 1
    istList = getRange(3, 8, 2)
    expList = [3, 5, 7]
    if str(istList) != str(expList):
        testsFailed += 1
        print("Error in getRange")
        print("    Ist:", str(istList))
        print("    Exp:", str(expList))
        print()

    testsPerformed += 1
    istList = getRange(0, -10, -3)
    expList = [0, -3, -6, -9]
    if str(istList) != str(expList):
        testsFailed += 1
        print("Error in getRange")
        print("    Ist:", str(istList))
        print("    Exp:", str(expList))
        print()

    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


def placer(strichArt="-", laenge=30):
    return generateStringRepeats(len=laenge, aStr=strichArt)


def AUTO_TEST_a_generateStringRepeats(verbal=False):
    testsPerformed = 0
    testsFailed = 0
    testSuite = "a_generateStringRepeats"

    # Test-Cases
    # ----------
    # generateStringRepeats(10, '-')     => '----------'
    # generateStringRepeats(9,  '+-=')   => '+-=+-=+-='
    # generateStringRepeats(3,  'A')     => 'AAA'
    # generateStringRepeats(5, '.-')     => '.-.-.'
    # generateStringRepeats(6)           => '      '
    testsPerformed += 1
    if generateStringRepeats(10, '-') != '----------':
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if generateStringRepeats(9, '+-=') != '+-=+-=+-=':
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if generateStringRepeats(3, 'A') != 'AAA':
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if generateStringRepeats(5, '.-') != '.-.-.':
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if generateStringRepeats(6) != '      ':
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


def underline(titleStr, strichArt="="):
    return unterstreichen(title=titleStr, aChar=strichArt, end="\n")


def unterstreichen(title, aChar="=", end="\n"):
    # Spezifikation: Unterstreich einen String auf dem Bildschirm
    #     Beispiel: print(unterstreichen('Die ist ein Test',"+")
    #               Die ist ein Test
    #               ++++++++++++++++
    # Test-Cases
    # ----------
    # unterstreichen('Hallo')                     => 'Hallo\n====='
    # unterstreichen('Die ist ein Test',"+")      => 'Die ist ein Test\n++++++++++++++++'
    return title + end + generateStringRepeats(len(title), aChar)


def AUTO_TEST_a_unterstreichen(verbal=False):
    testsPerformed = 0
    testsFailed = 0
    testSuite = "a_unterstreichen"

    # Test-Cases
    # ----------
    # unterstreichen('Hallo')                     => 'Hallo\n====='
    # unterstreichen('Die ist ein Test',"+")      => 'Die ist ein Test\n++++++++++++++++'
    testsPerformed += 1
    if unterstreichen('Hallo') != 'Hallo\n=====':
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if unterstreichen('Die ist ein Test', "+") != 'Die ist ein Test\n++++++++++++++++':
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1
    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


def TEST_stringFct():
    showASCII_Table()

    eingabeString = input("String:")
    print("toUpperCase(", eingabeString, ")           --> ", toUpperCase(eingabeString), sep="")
    print("toLowerCase(", eingabeString, ")           --> ", toLowerCase(eingabeString), sep="")
    print("toFirstUpperCase(", eingabeString, ")      --> ", toFirstUpperCase(eingabeString), sep="")


def hexStrToURLEncoded(inStr):
    # https://www.quora.com/How-do-I-convert-hex-into-a-string-using-Python
    retStr = ""
    instr = "".join(inStr.split())
    print("instr:", instr)
    listOfDoubleHexStr = []

    i = 0
    while i < len(instr):
        listOfDoubleHexStr.append(instr[i] + instr[i + 1])
        i += 2

    print("listOfDoubleHexStr:", listOfDoubleHexStr)
    for aDHex in listOfDoubleHexStr:
        print(aDHex, ":", end="")
        aChrOrd = ord(bytes.fromhex(aDHex))
        print(aChrOrd, ":", end="")
        partRetStr = ""
        if (aChrOrd > 32) and (aChrOrd < 127):
            partRetStr = bytes.fromhex(aDHex).decode('utf-8')
        else:
            partRetStr = "%" + aDHex + ""
        retStr += partRetStr
        print(partRetStr, ":")
    return retStr


def TEST_hexStrToURLEncoded():
    testStr_0 = "5350430d0a303230300d0a31"
    testStr_1 = """53 50 43 0d   0a 30 32 30 30 0d 0a 31
        0d 0a 43 48 39 33 33 30   30 30 30 30 30 31 39 30"""

    testStr_2 = """
        71 a4 00 e2 53 50 43 0d   0a 30 32 30 30 0d 0a 31
        0d 0a 43 48 39 33 33 30   30 30 30 30 30 31 39 30
        37 36 31 36 34 30 35 0d   0a 4b 0d 0a 4e 6f 76 61
        54 72 65 6e 64 20 53 65   72 76 69 63 65 73 20 47
        6d 62 48 0d 0a 42 61 68   6e 68 6f 66 73 74 72 61
        73 73 65 20 31 38 0d 0a   36 33 34 30 20 42 61 61
        72 0d 0a 0d 0a 0d 0a 43   48 0d 0a 0d 0a 0d 0a 0d
        0a 0d 0a 0d 0a 0d 0a 0d   0a 36 2e 30 35 0d 0a 43
        48 46 0d 0a 4b 0d 0a 57   61 6c 74 65 72 20 52 6f
        74 68 6c 69 6e 0d 0a 50   65 74 65 72 6c 69 77 69
        65 73 65 20 33 33 0d 0a   38 38 35 35 20 57 61 6e
        67 65 6e 0d 0a 0d 0a 0d   0a 43 48 0d 0a 51 52 52
        0d 0a 30 30 30 30 30 30   30 30 30 30 30 30 30 30
        30 34 30 34 34 33 30 33   38 32 39 39 35 0d 0a 0d
        0a 45 50 44 0d 0a 00 ec   11 ec 11 ec 11 ec 11 ec
        11 ec 11 ec 11 ec 11 ec   11 ec 11 ec 11 ec
    """

    testStr_3 = """
            00 01 02 03 04 05 06 07   08 09 0A 0B 0C 0d 0e 0f
            10 11 12 13 14 15 16 17   18 19 1A 1B 1C 1d 1e 1f
            20 21 22 23 24 25 26 27   28 29 2A 2B 2C 2d 2e 2f
            30 31 32 33 34 35 36 37   38 39 3A 3B 3C 3d 3e 3f
            40 41 42 43 44 45 46 47   48 49 4A 4B 4C 4d 4e 4f
            50 51 52 53 54 55 56 57   58 59 5A 5B 5C 5d 5e 5f
            60 61 62 63 64 65 66 67   68 69 6A 6B 6C 6d 6e 6f
            70 71 72 73 74 75 76 77   78 79 7A 7B 7C 7d 7e 7f
            80 81 82 83 84 85 86 87   88 89 8A 8B 8C 8d 8e 8f
            f0 f1 f2 f3 f4 f5 f6 f7   f8 f9 fA fB fC fd fe ff
    """
    print(hexStrToURLEncoded(testStr_3))


# UNICODE Functions
# =================
# https://www.geeksforgeeks.org/how-to-print-superscript-and-subscript-in-python/

def TEST_get_sup_super():
    EinsSub = get_sub('1')
    print("get_sub('1'):", EinsSub)
    print("get_sub('2'):", get_sub('2'))
    print("get_super('1'):", get_super('1'))


def get_sub(x):
    """
    function to convert to subscript
    Wandelt ein Zeichen in ein Index-Zeichen (tiefgestellt) um:

        get_sub('2') ==> ₂
    """
    normal = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+-=()"
    sub_s = "ₐ₈CDₑբGₕᵢⱼₖₗₘₙₒₚQᵣₛₜᵤᵥwₓᵧZₐ♭꜀ᑯₑբ₉ₕᵢⱼₖₗₘₙₒₚ૧ᵣₛₜᵤᵥwₓᵧ₂₀₁₂₃₄₅₆₇₈₉₊₋₌₍₎"
    res = x.maketrans(''.join(normal), ''.join(sub_s))
    return x.translate(res)


def get_super(x):
    """
    function to convert to superscript
    Wandelt ein Zeichen in ein Hoch-Zeichen (hochgestellt) um:

        get_super('2') ==> ²
    """
    normal = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+-=()"
    super_s = "ᴬᴮᶜᴰᴱᶠᴳᴴᴵᴶᴷᴸᴹᴺᴼᴾQᴿˢᵀᵁⱽᵂˣʸᶻᵃᵇᶜᵈᵉᶠᵍʰᶦʲᵏˡᵐⁿᵒᵖ۹ʳˢᵗᵘᵛʷˣʸᶻ⁰¹²³⁴⁵⁶⁷⁸⁹⁺⁻⁼⁽⁾"
    res = x.maketrans(''.join(normal), ''.join(super_s))
    return x.translate(res)


# Date and Timestamp
# ==================
def getTimestamp(preStr="", postStr="", formatString="nice"):
    """
        formatString == nice    ==> {ts:%Y-%m-%d %H:%M:%S}   2021-01-02 13:15:03
        formatString == Default ==> {ts:%Y%m%d%H%M%S}        20210102131503

        %a 	Weekday, short version 	Wed
        %A 	Weekday, full version 	Wednesday
        %w 	Weekday as a number 0-6, 0 is Sunday 	3
        %d 	Day of month 01-31 	31
        %b 	Month name, short version 	Dec
        %B 	Month name, full version 	December
        %m 	Month as a number 01-12 	12
        %y 	Year, short version, without century 	18
        %Y 	Year, full version 	2018
        %H 	Hour 00-23 	17
        %I 	Hour 00-12 	05
        %p 	AM/PM 	PM
        %M 	Minute 00-59 	41
        %S 	Second 00-59 	08
        %f 	Microsecond 000000-999999 	548513
        %z 	UTC offset 	+0100
        %Z 	Timezone 	CST
        %j 	Day number of year 001-366 	365
        %U 	Week number of year, Sunday as the first day of week, 00-53 	52
        %W 	Week number of year, Monday as the first day of week, 00-53 	52
        %c 	Local version of date and time 	Mon Dec 31 17:41:00 2018
        %x 	Local version of date 	12/31/18
        %X 	Local version of time 	17:41:00
        %% 	A % character 	%
        %G 	ISO 8601 year 	2018
        %u 	ISO 8601 weekday (1-7) 	1
        %V 	ISO 8601 weeknumber (01-53) 	01
    """
    formatStr = '{ts:%Y-%m-%d %H:%M:%S}'
    if (formatString == ""):
        formatStr = '{ts:%Y%m%d%H%M%S}'
    elif (formatString == "nice"):
        formatStr = '{ts:%Y-%m-%d %H:%M:%S}'
    else:
        formatStr = formatString
    retStr = formatStr.format(ts=datetime.now())
    # retStr = left(retStr,len(retStr)-2)
    return preStr + retStr + postStr


def get_actual_year():
    return int(getTimestamp(formatString="{ts:%Y}"))


def TEST_getTimestamp():
    print("TEST_getTimestamp...")
    print(datetime.now())


# True if (old-young > limit)
def checkTimeDifference(oldTimestamp, youngTimestamp, limit, gt=True):
    timeDiff = youngTimestamp - oldTimestamp
    secStr = str(timeDiff)[:7]
    if (gt):
        return secStr > limit
    else:
        return secStr < limit


# Math functions
# ==============
def equalsWithinTolerance(ist, soll, abweichungProzent=0.001):
    if (soll == 0) and (ist == 0):
        return True
    else:
        if abs(100 - (ist * 100 / soll)) > abweichungProzent:
            return False
        else:
            return True


def is_US_Date(date_str, reg_ex=regEx_Date_US, verbal=False):
    if verbal:
        print(f"is_US_Date('{date_str}', '{reg_ex}')")
    if re.fullmatch(reg_ex, date_str):
        return True
    else:
        return False


def is_EU_Date(date_str, reg_ex=regEx_Date_EU, verbal=False):
    if verbal:
        print(f"is_EU_Date('{date_str}', '{reg_ex}')")
    if re.fullmatch(reg_ex, date_str):
        return True
    else:
        return False


def isFloatEquals(ist, soll, roundDezimals=3):
    # Test-Cases
    # ----------
    # isFloatEquals(4.5321,4.5329,2)      => True
    # isFloatEquals(4.5321,4.5329,3)      => False
    return round(ist, roundDezimals) == round(soll, roundDezimals)


def AUTO_TEST_a_isFloatEquals(verbal=False):
    testsPerformed = 0
    testsFailed = 0
    testSuite = "a_isFloatEquals"

    # Test-Cases
    # ----------
    # isFloatEquals(4.5321,4.5329,2)      => True
    # isFloatEquals(4.5321,4.5329,3)      => False
    testsPerformed += 1
    if isFloatEquals(4.5321, 4.5329, 2) != True:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    testsPerformed += 1
    if isFloatEquals(4.5321, 4.5329, 3) != False:
        print("Failed in ", testSuite, "  Case:", testsPerformed)
        testsFailed += 1

    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


# Inspection functions
# ====================
def getMyFctName():
    # liste = inspect.stack()[1]
    # for aItem in liste:
    #     print("   --> ", aItem)
    # print(inspect.stack()[1][4])
    return inspect.stack()[1][3]


# REST / JSON functions
# =====================
# https://stackoverflow.com/questions/7320319/xpath-like-query-for-nested-python-dictionaries
def xPath_Get(mydict, path):
    elem = mydict
    try:
        for x in path.strip("/").split("/"):
            try:
                x = int(x)
                elem = elem[x]
            except ValueError:
                elem = elem.get(x)
    except:
        pass
    return elem


def AUTO_TEST_xPath_Get(verbal=False):
    if verbal:
        print("==> ", getMyFctName())

    foo = {
        'spam': 'eggs',
        'morefoo': [{
            'bar': 'soap',
            'morebar': {
                'bacon': {
                    'bla': '12345'
                }
            }
        },
            'Walt'
        ]
    }

    testCasesExecuted = 5
    testCasesFailed = 2
    testCases = """"
    Expected | Param_1                     
    1234     | /morefoo/0/morebar/bacon/bla
    Walti    | /morefoo/1
    """
    param1 = "/morefoo/0/morebar/bacon/bla"
    expextedVal1 = "1234"
    retVal = xPath_Get(foo, param1)
    if (retVal != expextedVal1):
        print("   ERROR in TEST: xPath_Get(foo, ", param1, ") = ", retVal, "    ==> expected: ", expextedVal1, sep="")

    param1 = "morefoo/1"
    expextedVal1 = "Walti"
    retVal = xPath_Get(foo, param1)
    if (retVal != expextedVal1):
        print("   ERROR in TEST: xPath_Get(foo, ", param1, ") = ", retVal, "    ==> expected: ", expextedVal1, sep="")

    if verbal:
        print("--> Test Cases Executed: {a:4d}".format(a=testCasesExecuted))
        print("--> Test Cases Failed  : {a:4d}".format(a=testCasesFailed))
    return {"TestName": getMyFctName(), "testCasesExecuted": testCasesExecuted, "testCasesFailed": testCasesFailed}


def format_IBAN(IBAN):
    IBAN_str = IBAN.replace(' ', '')
    IBAN_str = IBAN_str[:4] + ' ' + IBAN_str[4:8] + ' ' + IBAN_str[8:12] + ' ' + IBAN_str[12:16] + ' ' + IBAN_str[16:20] + ' ' + IBAN_str[20:]
    return IBAN_str


def generateQRInvoiceData(creditor_iban, creditor_addr, debitor_addr, amount, currency='CHF', reference=None, additional_information=""):
    invoice_data = {
        "creditor_iban": creditor_iban,
        "creditor_name": creditor_addr["name"],
        "creditor_address": creditor_addr["address"],
        "creditor_zip_code": creditor_addr["zip_code"],
        "creditor_city": creditor_addr["city"],
        "creditor_country": creditor_addr["country"],

        "debtor_name": debitor_addr["name"],
        "debtor_address": debitor_addr["address"],
        "debtor_zip_code": debitor_addr["zip_code"],
        "debtor_city": debitor_addr["city"],
        "debtor_country": debitor_addr["country"],

        "amount": amount,
        "currency": currency,

        "reference_type": "NON",
        "reference_number": "",

        "additional_information": additional_information,
    }

    if reference is not None:
        invoice_data["reference_type"] = reference["reference_type"]
        invoice_data["reference_number"] = reference["reference_number"]

    return invoice_data


# URL operations
# ==============
def loadAndSaveFileFromURL(url='http://google.com/favicon.ico'):
    filename = url.split('/')[-1]
    r = requests.get(url, allow_redirects=True)
    open(filename, 'wb').write(r.content)
    return filename


# File and Directory operations
# =============================
def getFilenameWithoutExtension(fileName):
    return fileName[:fileName.index(".")]


def getFilenameExtension(fileName):
    return fileName[fileName.index("."):]


def addTimestampToFileName(fileName, timestampFormat="{ts:%Y_%m_%d}"):
    indexBeforeFileType = fileName.rindex(".")
    fileName = fileName[:indexBeforeFileType] + "_" + getTimestamp(formatString=timestampFormat) + fileName[indexBeforeFileType:]
    return fileName


def getPath(full_filename=None):
    if full_filename is None or len(full_filename) == 0:
        return os.path.abspath(os.getcwd()) + "\\"
    else:
        directory, filename = os.path.split(full_filename)
        return directory


def getFilename(full_filename=None):
    if full_filename is None or len(full_filename) == 0:
        return ''
    else:
        directory, filename = os.path.split(full_filename)
        return filename


def createDirIfNotExists(dir_path="./TestData", access_rights=0o755, verbal=False):
    try:
        ## os.mkdir(dir_path, access_rights)
        os.makedirs(dir_path, access_rights)
    except OSError:
        if verbal:
            print("Creation of the directory %s failed" % dir_path)
    else:
        if verbal:
            print("Successfully created the directory %s " % dir_path)


def deleteDir(dir_path="./TestData", verbal=False):
    try:
        shutil.rmtree(dir_path)  # does it even if it contains files
        ## os.rmdir(dir_path)    # only if the director is empty
    except OSError:
        if verbal:
            print("Deletion of the directory %s failed" % dir_path)
    else:
        if verbal:
            print("Successfully deleted the directory %s " % dir_path)


#    TBC verallgemeinern start
# Cleanup - Rule N°1 : Verzeichnis wird durchsucht und alle Files mit .csv werden gelöscht
def file_cleanup1():
    logfiles = Path('C:\\Users\charl\PycharmProjects\Python_HWZ\programming_tools\projektarbeit_python')
    for file in logfiles.glob('*.csv'):
        if file.is_file():
            try:
                file.unlink()
                print("Files gelöscht")
            except OSError as error:
                print("File {} konnte nicht gelöscht werden: {}".format(file, error))


# Cleanup - Rule N°2 : Verzeichnis wird durchsucht und alle Files(.csv) älter als 7 Tage werden gelöscht
def File_cleanup2():
    logfiles = Path('C:\\Users\charl\PycharmProjects\Python_HWZ\programming_tools\projektarbeit_python')
    for file in logfiles.glob('*.csv'):
        create_time = os.path.getctime(file)
        if (time.time() - create_time) // (24 * 3600) >= 7:
            os.unlink(file)
            print('{} removed'.format(file))


def File_getAllLogFiles(path):
    files = os.listdir(path)
    files = list(filter(lambda file: file.endswith('.csv'), files))
    return [path + file for file in files]  # append path to the file to delete it later


def File_removeOldLogs(files, maxLogFiles=10):
    while len(files) > maxLogFiles:
        oldestFile = min(files, key=os.path.getctime)
        os.remove(oldestFile)
        files = File_getAllLogFiles(".")


def File_remove(fullFileName, verbal=True, withFullName=True):
    fname = os.path.basename(fullFileName)
    if os.path.exists(fullFileName):
        os.remove(fullFileName)
        if verbal:
            if withFullName:
                print("File {} deleted!".format(fullFileName))
            else:
                print("File {} deleted!".format(fname))

    else:
        if verbal:
            if withFullName:
                print("Die Datei {} existiert nicht".format(fullFileName))
            else:
                print("Die Datei {} existiert nicht".format(fname))


def File_cleanup(filename, directory, path_sign):
    """Cleanup Funktion löscht angegebene Dateien aus angegebenen Verzeichnisse"""
    path_sign = path_sign
    if directory == "":
        directory = os.curdir + path_sign
    if not filename == "":
        filename = str(directory) + path_sign + str(filename)
        if os.path.exists(filename):
            print("Datei {} löschen".format(filename))
            os.remove(filename)
        else:
            print("Die Datei {} existiert nicht".format(filename))
    else:
        for file in os.listdir(directory):
            if file.startswith("logger"):
                os.chdir(directory)
                print("Datei {} löschen".format(file))
                os.remove(file)


#    TBC verallgemeinern start

# File manipulation
# =================
def readFile(sourceFileFN, lineEnd="\n"):
    return File_getFileContent(sourceFileFN, returnType="String", lineEnd=lineEnd)


'''String or ListOfLines'''


def File_getFileContent(sourceFileFN, returnType="String", lineEnd="\n"):
    # with open(sourceFileFN, "r", encoding="utf-8") as f:
    with open(sourceFileFN, "r", encoding="utf-8") as f:
        lines = f.readlines()
        lines = [aLine.strip('\n\r') for aLine in lines]
    if returnType == "ListOfLines":
        return lines
    elif returnType == "String":
        return lineEnd.join(lines)
    else:
        return "ERROR: File_getFileContent unknown format!"


def read_text_file_into_list_of_lines(filename):
    try:
        with open(filename, 'r') as file:
            lines = file.readlines()
            return lines
    except FileNotFoundError:
        print(f"ERROR: File '{filename}' not found.")
        return []


def File_createTestFile(aFileFN, startLineNr=1, endLineNr=20, aHeader="", aFooter="", aContent=""):
    aTestFile = open(aFileFN, "w")
    if aHeader != "":
        aTestFile.write(aHeader + "\n")
    for lNr in range(startLineNr, endLineNr + 1):
        aTestFile.write(str(lNr) + aContent + "\n")

    if aFooter != "":
        aTestFile.write(aFooter + "\n")
    aTestFile.close()


def File_getCountOfLines(sourceFileFN):
    lines = []
    with open(sourceFileFN, "r", encoding="utf-8") as f:
        lines = f.readlines()
    return len(lines)


def File_deleteLines(sourceFileFN, destinationFileFN=None, deleteLineFrom=None, deleteLineTo=None, verbal=False):
    if destinationFileFN is None:
        destinationFileFN = sourceFileFN

    if (deleteLineFrom is None) and (deleteLineTo is None):
        deleteLineFrom = 1
        deleteLineTo = 0
    elif (deleteLineFrom is not None) and (deleteLineTo is None):
        deleteLineTo = 1000000
    elif (deleteLineFrom is None) and (deleteLineTo is not None):
        deleteLineFrom = 1
    else:
        pass  # NOP

    if verbal:
        print("    Delete from", deleteLineFrom, "to", deleteLineTo, end="")

    # File in eine Liste lesen
    with open(sourceFileFN, "r", encoding="utf-8") as f:
        lines = f.readlines()

    # In Liste Range löschen
    del lines[deleteLineFrom - 1:deleteLineTo]

    # Liste in ein File schreiben
    with open(destinationFileFN, "w", encoding="utf-8") as f:
        f.writelines(lines)


def File_addHeader(sourceFileFN, destinationFileFN=None, headerStr=""):
    if destinationFileFN is None:
        destinationFileFN = sourceFileFN

    # File in eine Liste lesen
    with open(sourceFileFN, "r", encoding="utf-8") as f:
        lines = f.readlines()

    # Liste in ein File schreiben
    aTestFile = open(destinationFileFN, "w")
    aTestFile.write(headerStr)
    aTestFile.write("\n")
    aTestFile.writelines(lines)
    aTestFile.close()


def File_create(filename, str_to_save='', mode='w', encoding="utf-8", verbal=False):
    if verbal:
        print(f'File written to {filename}')
    with open(filename, mode, encoding=encoding) as file:
        file.write(str_to_save)


def AUTO_TEST_a_File_addHeader(verbal=False):
    testsPerformed = 0
    testsFailed = 0
    testSuite = "a_File_addHeader"

    testPath = "./TestData/AUTO_TEST_a_File_addHeader"
    createDirIfNotExists(dir_path=testPath)

    testsPerformed += 1
    baseTestFile = testPath + "/Test_3.txt"
    testFileName = testPath + "/Test_3a.txt"

    File_createTestFile(baseTestFile)
    expectedResult = 22
    File_addHeader(baseTestFile, testFileName, headerStr="This is \na multiline String")
    if File_getCountOfLines(testFileName) != expectedResult:
        print("Error in testSuite:   ", testSuite, ":     case:", testsPerformed, sep="")
        print("    Test Failed: File_getCountOfLines(" + testFileName + ")", "Result: ",
              File_getCountOfLines(testFileName), "   Expected:", expectedResult, end="\n\n")
        testsFailed += 1

    testsPerformed += 1
    testFileName = testPath + "/Test_2b.txt"
    expectedResult = 21
    File_addHeader(baseTestFile, testFileName, headerStr="This is a singleline String")
    if File_getCountOfLines(testFileName) != expectedResult:
        print("Error in testSuite:   ", testSuite, ":     case:", testsPerformed, sep="")
        print("    Test Failed: File_getCountOfLines(" + testFileName + ")", "Result: ",
              File_getCountOfLines(testFileName), "   Expected:", expectedResult, end="\n\n")
        testsFailed += 1

    deleteDir(testPath)
    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


def File_addFooter(sourceFileFN, destinationFileFN=None, footerStr=""):
    if destinationFileFN is None:
        destinationFileFN = sourceFileFN

    # File in eine Liste lesen
    with open(sourceFileFN, "r", encoding="utf-8") as f:
        lines = f.readlines()

    # Liste in ein File schreiben
    aTestFile = open(destinationFileFN, "w")
    aTestFile.writelines(lines)
    aTestFile.write(footerStr)
    aTestFile.close()


def getRegExMatches(inString, regEx):
    matches = re.findall(regEx, inString)
    return matches


def getIncludeFileName(aTextLine, includePattern='# include:\S+'):
    aFilename = ""
    listOfMatches = getRegExMatches(aTextLine, includePattern)
    if len(listOfMatches) > 0:
        aFilename = listOfMatches[0][10:]  # hard codiert len('# include:')
    return aFilename


def TEST_getIncludeFileName():
    print(getIncludeFileName("3   # include:Test_1_With_Include_2.txt   kkkkkkkk"))
    print(getIncludeFileName("   4   # include:./hhhh/Test_1_With_Include_22.txt   "))


def File_readWithInludes(sourceFileFN, includePattern='# include:\S+', includeSearchPath="./", recLevel=0):
    '''
    returns a list of lines and resolves the includes
    Default include pattern is: <include:filename>
    Don't pass $recLevel. It is used internaly for recursion
    The function doesn't read any other character on the same line as the include command
    '''
    f = open(sourceFileFN, "r")
    fileContent = f.readlines()
    f.close()
    for aLine in fileContent:
        includeFileName = getIncludeFileName(aLine, includePattern)
        if includeFileName != "" and recLevel < 3:
            File_readWithInludes(includeFileName, includePattern, includeSearchPath, recLevel=recLevel + 1)
        print(fileContent)


def AUTO_TEST_FileFunctions(verbal=False):
    testsPerformed = 0
    testsFailed = 0

    testSuite = getMyFctName()[auto_test_fct_prefix_len:]

    testPath = "./TestData/AUTO_TEST_FileFunctions"
    createDirIfNotExists(dir_path=testPath)

    # -------------------------------------------------------------------------------
    fct = "File_getCountOfLines()"
    # -------------------------------------------------------------------------------
    case = 1
    testsPerformed += 1
    testFileName = testPath + "/Test_1.txt"
    File_createTestFile(testPath + "/Test_1.txt")
    countOfLine = File_getCountOfLines(testFileName)
    expectedResult = 20
    if (countOfLine != expectedResult):
        print("Error in testSuite:   ", testSuite, "(1):   ", fct, "    case:", case, sep="")
        print("    Test Failed: File_getCountOfLines(" + testFileName + ")", "Result: ",
              File_getCountOfLines("./TestData/Test_1.txt"), "   Expected:", expectedResult, end="\n\n")
        testsFailed += 1

    case = 2
    testsPerformed += 1
    testFileName = testPath + "/Test_2.txt"
    File_createTestFile(testFileName, aHeader="Nr |", aContent=" | Content", aFooter="File Ende", startLineNr=5,
                        endLineNr=45)
    countOfLine = File_getCountOfLines(testPath + "/Test_2.txt")
    expectedResult = 43
    if (countOfLine != expectedResult):
        print("Error in testSuite:   ", testSuite, "(2):   ", fct, "    case:", case, sep="")
        print("    Test Failed: File_getCountOfLines(" + testFileName + ")", "Result: ",
              File_getCountOfLines("./TestData/Test_2.txt"), "   Expected:", expectedResult, end="\n\n")
        testsFailed += 1
    # -------------------------------------------------------------------------------
    fct = "File_deleteLines()"
    # -------------------------------------------------------------------------------
    case = 1
    testsPerformed += 1
    testFileName = testPath + "/Test_1a.txt"
    expectedResult = 20
    File_deleteLines(testPath + "/Test_1.txt", testFileName, deleteLineFrom=None, deleteLineTo=None, verbal=False)
    if File_getCountOfLines(testFileName) != expectedResult:
        print("Error in testSuite:   ", testSuite, ":   ", fct, "    case:", case, sep="")
        print("    Test Failed: File_getCountOfLines(" + testFileName + ")", "Result: ",
              File_getCountOfLines(testFileName), "   Expected:", expectedResult, end="\n\n")
        testsFailed += 1

    case = 2
    testsPerformed += 1
    testFileName = testPath + "/Test_1b.txt"
    expectedResult = 14
    File_deleteLines(testPath + "/Test_1.txt", testFileName, deleteLineFrom=2, deleteLineTo=7)
    if File_getCountOfLines(testFileName) != expectedResult:
        print("Error in testSuite:   ", testSuite, ":   ", fct, "    case:", case, sep="")
        print("    Test Failed: File_getCountOfLines(" + testFileName + ")", "Result: ",
              File_getCountOfLines(testFileName), "   Expected:", expectedResult, end="\n\n")
        testsFailed += 1

    case = 3
    testsPerformed += 1
    testFileName = testPath + "/Test_1c.txt"
    expectedResult = 4
    File_deleteLines(testPath + "/Test_1.txt", testFileName, deleteLineFrom=5)
    if File_getCountOfLines(testFileName) != expectedResult:
        print("Error in testSuite:   ", testSuite, ":   ", fct, "    case:", case, sep="")
        print("    Test Failed: File_getCountOfLines(" + testFileName + ")", "Result: ",
              File_getCountOfLines(testFileName), "   Expected:", expectedResult, end="\n\n")
        testsFailed += 1

    case = 4
    testsPerformed += 1
    testFileName = testPath + "/Test_1d.txt"
    expectedResult = 13
    File_deleteLines(testPath + "/Test_1.txt", testFileName, deleteLineTo=7)
    if File_getCountOfLines(testFileName) != expectedResult:
        print("Error in testSuite:   ", testSuite, ":   ", fct, "    case:", case, sep="")
        print("    Test Failed: File_getCountOfLines(" + testFileName + ")", "Result: ",
              File_getCountOfLines(testFileName), "   Expected:", expectedResult, end="\n\n")
        testsFailed += 1

    # -------------------------------------------------------------------------------
    fct = "File_addHeader()"
    # -------------------------------------------------------------------------------
    case = 1
    testsPerformed += 1
    testFileName = testPath + "/Test_1a.txt"
    expectedResult = 22
    File_addHeader(testFileName, headerStr="Hallo this ist ein \nHeader")
    if File_getCountOfLines(testFileName) != expectedResult:
        print("Error in testSuite:   ", testSuite, ":   ", fct, "    case:", case, sep="")
        print("    Test Failed: File_getCountOfLines(" + testFileName + ")", "Result: ",
              File_getCountOfLines(testFileName), "   Expected:", expectedResult, end="\n\n")
        testsFailed += 1

    # -------------------------------------------------------------------------------
    fct = "File_addFooter()"
    # -------------------------------------------------------------------------------
    case = 1
    testsPerformed += 1
    testFileName = testPath + "/Test_1a.txt"
    expectedResult = 24
    File_addFooter(testFileName, footerStr="Hier ist ein \nFooter")
    if File_getCountOfLines(testFileName) != expectedResult:
        print("Error in testSuite:   ", testSuite, ":   ", fct, "    case:", case, sep="")
        print("    Test Failed: File_getCountOfLines(" + testFileName + ")", "Result: ",
              File_getCountOfLines(testFileName), "   Expected:", expectedResult, end="\n\n")
        testsFailed += 1

    #### TEST_getIncludeFileName()
    #### File_readWithInludes("./TestData/Test_1_With_Include_1.txt")
    deleteDir(testPath)
    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


# Geometrische Formen berechnen
# =============================

#  Kreis
#  -----
def calcCircle_Flaeche(radius=None, durchmesser=None, flaeche=None, umfang=None):
    if radius is not None:
        return radius ** 2 * math.pi
    elif durchmesser is not None:
        return (durchmesser / 2) ** 2 * math.pi
    elif flaeche is not None:
        return flaeche
    elif umfang is not None:
        return ((umfang / math.pi) / 2) ** 2 * math.pi
    else:
        return 0


def TEST_calcCircle_Flaeche():
    radius = 10
    print(calcCircle_Flaeche(radius=radius))
    print(calcCircle_Flaeche(durchmesser=2 * radius))
    print(calcCircle_Flaeche(umfang=2 * radius * math.pi))
    print(calcCircle_Flaeche(flaeche=radius ** 2 * math.pi))


def calcCircle_Umfang(radius=None, durchmesser=None, flaeche=None, umfang=None):
    if radius is not None:
        return radius * 2 * math.pi
    elif durchmesser is not None:
        return durchmesser * math.pi
    elif flaeche is not None:
        return (math.sqrt(flaeche / math.pi)) * 2 * math.pi
    elif umfang is not None:
        return umfang
    else:
        return 0


def TEST_calcCircle_Umfang():
    radius = 10
    print(calcCircle_Umfang(radius=radius))
    print(calcCircle_Umfang(durchmesser=2 * radius))
    print(calcCircle_Umfang(umfang=2 * radius * math.pi))
    print(calcCircle_Umfang(flaeche=radius ** 2 * math.pi))


def calcCircle_Radius(radius=None, durchmesser=None, flaeche=None, umfang=None):
    if radius is not None:
        return radius
    elif durchmesser is not None:
        return durchmesser / 2
    elif flaeche is not None:
        return math.sqrt(flaeche / math.pi)
    elif umfang is not None:
        return (umfang / math.pi) / 2
    else:
        return 0


def TEST_calcCircle_Radius():
    radius = 10
    print(calcCircle_Radius(radius=radius))
    print(calcCircle_Radius(durchmesser=2 * radius))
    print(calcCircle_Radius(umfang=2 * radius * math.pi))
    print(calcCircle_Radius(flaeche=radius ** 2 * math.pi))


def calcCircle_Durchmesser(radius=None, durchmesser=None, flaeche=None, umfang=None):
    if radius is not None:
        return radius * 2
    elif durchmesser is not None:
        return durchmesser
    elif flaeche is not None:
        return math.sqrt(flaeche / math.pi) * 2
    elif umfang is not None:
        return umfang / math.pi
    else:
        return 0


def TEST_calcCircle_Durchmesser():
    radius = 10
    print(calcCircle_Durchmesser(radius=radius))
    print(calcCircle_Durchmesser(durchmesser=2 * radius))
    print(calcCircle_Durchmesser(umfang=2 * radius * math.pi))
    print(calcCircle_Durchmesser(flaeche=radius ** 2 * math.pi))


def TEST_CircleFct():
    TEST_calcCircle_Flaeche()
    TEST_calcCircle_Umfang()
    TEST_calcCircle_Radius()
    TEST_calcCircle_Durchmesser()


# Quadratische Gleichungen
# ========================
def TEST_calcNulstellen():
    print('1)', calcNullstellen_checked('2', '1', '-4', True))
    try:
        print('2)', calcNullstellen('2', '1', '-4'))
    except TypeError:
        print('2)', "Exception geworfen!!!")

    print('3)', calcNullstellen_checked(2, 1, -4, True))
    print('4)', calcNullstellen_checked('2', '1', '-4', True))

    try:
        print('5)', calcNullstellen_checked('2', '1', '-4', False))
    except TypeError:
        print('5)', "Exception geworfen wegen falschen Datentypen der Parameter!!!")

    try:
        print('6)', calcNullstellen_checked(2, 1))
    except TypeError:
        print('6)', "Exception geworfen wegen falschen Datentypen der Parameter!!!")


def calcNullstellen(a, b, c):
    """
    Berechnung der Nullstellen einer quadratischen Funktion der Form:
       y = ax² + bx + c

       Mitternachtsformel: https://www.mathebibel.de/mitternachtsformel

       Positive Testfälle: a=2    b=1    c=-4      Diskriminante: 33    x₁=-1.69  x₂=1.19
       Positive Testfälle: a=1    b=0    c=0       Diskriminante:  0    x₁,₂=0
       Negative Testfälle: a=1    b=2    c=3       Diskriminante: -8    x₁=----   x₂=-----

        return {"Diskriminante": diskriminante, "Solutions": 2, "Solution Text": "Zwei Lösungen", "x1": x1, "x2": x2}
    """

    diskriminante = b ** 2 - 4 * a * c
    if diskriminante < 0:
        return {"Diskriminante": diskriminante, "Solutions": 0, "Solution Text": "Keine Lösung"}
    elif diskriminante == 0:
        x1 = (-b) / (2 * a)
        return {"Diskriminante": diskriminante, "Solutions": 1, "Solution Text": "Eine Lösung", "x1": x1, "x2": x1}
    else:
        x1 = (-b + math.sqrt(diskriminante)) / (2 * a)
        x2 = (-b - math.sqrt(diskriminante)) / (2 * a)
        return {"Diskriminante": diskriminante, "Solutions": 2, "Solution Text": "Zwei Lösungen", "x1": x1, "x2": x2}


def calcNullstellen_checked(a, b, c, ParameterCheck=True, ParamCheckErrorMsg="Falsche Parameter"):
    if ParameterCheck:
        try:
            loesungen = calcNullstellen(a, b, c)
            return loesungen
        except TypeError:
            return {"Solutions": -1, "Solution Text": ParamCheckErrorMsg}
    else:
        return calcNullstellen(a, b, c)


# Symetrische Cryptographie
# =========================
def shifter(sChr, sh):
    retVal = chr(ord(sChr) + sh)
    return retVal


# shifter using a ringbuffer (source
def shiftChr(aChar, shift):
    if (aChar >= " ") and (aChar <= "~"):
        return chr(((ord(aChar) - ord(' ') + shift) % (ord('~') - ord(' ') + 1)) + ord(' '))
    else:
        return aChar


# using comprehension
def chipher(text, cipherKey, encript=False):
    return "".join([element[1] for element in
                    [[string_in_list,
                      shiftChr(string_in_list, (1 if encript else -1) * ord(cipherKey[i % len(cipherKey)]))]
                     for i, string_in_list in zip(range(len(text)), text)]])


def encrypt(klartext, cipherKey):
    return chipher(klartext, cipherKey, encript=True)


def decrypt(geheimtext, cipherKey):
    return chipher(geheimtext, cipherKey, encript=False)


def encrypt_old(klartext, aKey):
    keyIndex = 0
    geheimtext = ""
    for aChar in klartext:
        aKeyChr = aKey[keyIndex]
        shifter = ord(aKeyChr)
        aSecretChr = shiftChr(aChar, shifter)
        # print(aChar, " (Rigth-Shift: ord(", aKeyChr, ") ", shifter, ") --> ", aSecretChr, sep="")
        keyIndex += 1
        if (keyIndex >= len(aKey)):
            keyIndex = 0
        geheimtext += aSecretChr
    return geheimtext


def decrypt_old(geheimtext, aKey):
    keyIndex = 0
    encryptedtext = ""
    for aChar in geheimtext:
        aKeyChr = aKey[keyIndex]
        shifter = ord(aKeyChr)
        decryptedChar = shiftChr(aChar, -shifter)
        # print(aChar, " (Left-Shift:        ", -shifter, ") --> ", decryptedChar, sep="", end="\n\n")
        keyIndex += 1
        if (keyIndex >= len(aKey)):
            keyIndex = 0
        encryptedtext += decryptedChar
    return encryptedtext


def AUTO_TEST_CryptDecrypt(verbal=False):
    testsPerformed = 0
    testsFailed = 0
    testSuite = "CryptDecrypt"

    klartext = "Hallo"
    keyStr = "12"

    testsPerformed += 1
    # -------------------------------------------------------------------------------
    fct = "encrypt()"
    # -------------------------------------------------------------------------------
    case = 1
    chiffrat = encrypt(klartext, keyStr)
    chiffratOld = encrypt_old(klartext, keyStr)
    if chiffrat != chiffratOld:
        print("Error in testSuite:   ", testSuite, ":   ", fct, "    case:", case, sep="")
        print("    Result: ", chiffrat, "   Expected:", chiffratOld, end="\n\n")
        testsFailed += 1

    testsPerformed += 1
    # -------------------------------------------------------------------------------
    fct = "decrypt()"
    # -------------------------------------------------------------------------------
    case = 1
    dechiffrat = decrypt(chiffrat, keyStr)
    dechiffratOld = decrypt_old(chiffratOld, keyStr)
    if dechiffrat != dechiffratOld:
        print("Error in testSuite:   ", testSuite, ":   ", fct, "    case:", case, sep="")
        print("    Result: ", dechiffrat, "   Expected:", dechiffratOld, end="\n\n")
        testsFailed += 1

    testsPerformed += 1
    # -------------------------------------------------------------------------------
    fct = "encrypt() and decrypt()"
    # -------------------------------------------------------------------------------
    if dechiffrat != klartext:
        print("Error in testSuite:   ", testSuite, ":   ", fct, "    case:", case, sep="")
        print("    Result: ", dechiffrat, "   Expected:", klartext, end="\n\n")
        testsFailed += 1

    if verbal:
        print("=>   ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v=testSuite), "Tests Performed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsPerformed), "      Tests Failed:",
              ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=testsFailed),
              "    Passed:{v:7.1f}".format(v=round(100 - (100 * testsFailed / testsPerformed), 1)), "%", sep="")
    return [testsPerformed, testsFailed]


# geo.admin search / tel.search / Open Street Map (OSM)
# =====================================================

def get_link_to_map_for_adress(ort='Wangen', strasse='Peterliwiese', hausnr='33', link_lable=None):
    if link_lable is None:
        link_lable = f'{strasse} {hausnr}, {ort}'
    return f'<a target="_new" href="https://search.ch/map/{ort},{strasse}-{hausnr}">{link_lable}</a>'

def get_coordinates_for_adresse(address='Peterliwiese 33, 8855 Wangen', country='ch', language='de'):
    BASE_URL = "https://nominatim.openstreetmap.org/search"

    params = {
        'q': f"{address}",  # Address query
        'format': 'json',  # Response format
        'addressdetails': 1,  # Include detailed address components
        'limit': 1,  # Limit to one result
        'countrycodes': f'{country}'  # Restrict to the US
    }

    headers = {
        'User-Agent': 'PythonApp/1.0 (your_email@example.com)'  # Replace with your email
    }

    response = requests.get(BASE_URL, params=params, headers=headers)

    if response.status_code == 200:
        result = response.json()
        # print("Result:", result)
        if result:
            return result[0]
            # lat = result[0]['lat']
            # lon = result[0]['lon']
            # display_name = result[0]['display_name']
            # print(f"Latitude: {lat}, Longitude: {lon}, Address: {display_name}")
        else:
            return "No results found"
    else:
        return f"Error: {response.status_code} - {response.reason}"

def search_address_geo_admin(address='Peterliwiese 33, 8855 Wangen', country='ch', language='de'):
    """
    Sucht eine Adresse in der GeoAdmin API und gibt die Ergebnisse zurück.

    :param address: Die zu suchende Adresse (z. B. "Bern, Bundesplatz 3").
    :param language: Sprache der Antwort, z. B. "de", "fr", "it", "en".
    :return: JSON-Antwort mit den Suchergebnissen.
    """
    base_url = "https://api3.geo.admin.ch/rest/services/api/SearchServer"
    params = {
        "searchText": address,  # Die Adresse, nach der gesucht wird
        "type": "locations",  # Typ der Suche (locations für Adressensuche)
        "lang": language,  # Sprache der Antwort
        "limit": 5,  # Max. Anzahl der Ergebnisse
    }

    try:
        response = requests.get(base_url, params=params)
        response.raise_for_status()  # Prüft auf HTTP-Fehler
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Fehler bei der Anfrage: {e}")
        return None


def getResults_geoAdmin(searchCriteriaEncoded, appId="", doTrace=False):
    serviceURL = "https://api3.geo.admin.ch/1912100956/rest/services/ech/SearchServer?sr=2056&searchText={search:2s}&lang=en&type=locations"
    requestStr = serviceURL.format(search=searchCriteriaEncoded)
    responseStr = requests.get(requestStr)
    jsonResponse = json.loads(responseStr.text)
    print("Request:\n", requestStr) if doTrace else False
    print("Response:\n", jsonResponse, "\n") if doTrace else False
    returnJSON = {'criteria': searchCriteriaEncoded,
                  'count': int(len(jsonResponse['results'])),
                  'results': []}

    recNr = 1
    # print("Parsed values (Records found:{recCount:2d}):".format(recCount=len(jsonResponse['results'])))
    for entry in jsonResponse['results']:
        details = entry['attrs']['detail']
        lon = entry['attrs']['lon']
        lat = entry['attrs']['lat']
        x = entry['attrs']['x']
        y = entry['attrs']['y']
        details = {'details': details,
                   'longitude': lon,
                   'latitude': lat,
                   'ch_x': x,
                   'ch_y': y}
        returnJSON['results'].append(details)
        print("\nRecord No: ", recNr) if doTrace else False
        print("  detail  :", details) if doTrace else False
        print("  lon     :", lon) if doTrace else False
        print("  lat     :", lat) if doTrace else False
        print("  x       :", x) if doTrace else False
        print("  y       :", y) if doTrace else False
        recNr += 1
    return returnJSON


def getFieldFromTelSearchXML(searchCH_Entry, namespaces, fieldname="type"):
    try:
        # print("++++++++++++++++++::::", fieldname, ":::")
        node = searchCH_Entry.find(fieldname, namespaces)
        # if fieldname == "extra":
        # print("----------------->")
        # print(node.text)
        # print(node.xPath("@type").text)
        # print("+++++++++++++++++>")
        retVal = node.text
    except AttributeError:
        retVal = ""
    return retVal


def getResults_search_ch(searchCriteriaEncoded, appId="8e8a84fd0f10d3b44920e49bc3b06a37", doTrace=False):
    serviceURL = "https://tel.search.ch/api/?q={search:2s}&key={appId:2s}"
    requestStr = serviceURL.format(search=searchCriteriaEncoded, appId=appId)
    responseStr = requests.get(requestStr).content
    print("Request:\n", requestStr) if doTrace else False
    ## print("Response:\n", responseStr, "\n\n\n")  if doTrace else False

    namespaces = {'tel': 'http://tel.search.ch/api/spec/result/1.0/',
                  'openSearch': 'http://a9.com/-/spec/opensearchrss/1.0/'}  # add more as needed
    dom = ET.fromstring(responseStr)
    countFound = int(dom.find('{http://a9.com/-/spec/opensearchrss/1.0/}totalResults').text)
    returnJSON = {'criteria': searchCriteriaEncoded,
                  'count': countFound,
                  'results': []}
    dom = etree.HTML(responseStr)
    value = dom.xpath('//entry')
    print("  Elements found  :", len(value)) if doTrace else False

    for aEntry in value:
        entryType = getFieldFromTelSearchXML(aEntry, namespaces, "type")
        name = getFieldFromTelSearchXML(aEntry, namespaces, "name")
        subname = getFieldFromTelSearchXML(aEntry, namespaces, "subname")
        firstname = getFieldFromTelSearchXML(aEntry, namespaces, "firstname")
        street = getFieldFromTelSearchXML(aEntry, namespaces, "street")
        streetno = getFieldFromTelSearchXML(aEntry, namespaces, "streetno")
        zip = getFieldFromTelSearchXML(aEntry, namespaces, "zip")
        city = getFieldFromTelSearchXML(aEntry, namespaces, "city")
        canton = getFieldFromTelSearchXML(aEntry, namespaces, "canton")
        country = getFieldFromTelSearchXML(aEntry, namespaces, "country")
        telNr = getFieldFromTelSearchXML(aEntry, namespaces, "phone")
        telNrExtra = getFieldFromTelSearchXML(aEntry, namespaces, "extra")
        content = getFieldFromTelSearchXML(aEntry, namespaces, "content")
        details = {'entryType': entryType,
                   'name': name,
                   'subname': subname,
                   'firstname': firstname,
                   'street': street,
                   'streetno': streetno,
                   'zip': zip,
                   'city': city,
                   'canton': canton,
                   'country': country,
                   'telNr': telNr,
                   'telNrExtra': telNrExtra,
                   'content': content
                   }
        returnJSON['results'].append(details)

        print("  aEntry  :", aEntry) if doTrace else False
        print("  Content :\n", content) if doTrace else False
        print("  telNr   :", telNr) if doTrace else False
        print("  Zip     :", zip) if doTrace else False
        print() if doTrace else False

    return returnJSON


def getResultsFromAdressSearch(searchCriteriaEncoded, doTrace=False):
    results = getResults_search_ch(searchCriteriaEncoded, doTrace=doTrace)
    resultsFoundInTelSearch = results['count']
    print("Records found with AdressSearch   :{recCount:2d}".format(recCount=resultsFoundInTelSearch)) if doTrace else False
    print(json.dumps(results, indent=4)) if doTrace else False
    for i in range(results['count']):
        # print("i:", i, results['results'][i]['street']) if doTrace else False
        # print("i:", i, results['results'][i]['streetno']) if doTrace else False
        # print("i:", i, results['results'][i]['zip']) if doTrace else False
        # print("i:", i, results['results'][i]['city']) if doTrace else False
        # print() if doTrace else False
        searchCriteria = str(results['results'][i]['street']) + ' ' + results['results'][i]['streetno'] + ', ' + results['results'][i]['zip'] + ' ' + results['results'][i]['city']
        print("searchCriteria:", searchCriteria) if doTrace else False
        searchCriteriaEncoded = urllib.parse.quote_plus(searchCriteria)
        resultsGoeAdmin = getResults_geoAdmin(searchCriteriaEncoded, doTrace=doTrace)
        if resultsGoeAdmin['count'] > 0:
            # print("resultsGoeAdmin:", resultsGoeAdmin['results'][0]) if doTrace else False
            results['results'][i]['geoDetails'] = resultsGoeAdmin['results'][0]['details']
            results['results'][i]['longitude'] = resultsGoeAdmin['results'][0]['longitude']
            results['results'][i]['latitude'] = resultsGoeAdmin['results'][0]['latitude']
            results['results'][i]['ch_x'] = resultsGoeAdmin['results'][0]['ch_x']
            results['results'][i]['ch_y'] = resultsGoeAdmin['results'][0]['ch_y']
    return results


def dictify(context, names):
    node = context.context_node
    rv = []
    rv.append('__dictify_start_marker__')
    names = names.split('|')
    for n in names:
        if n.startswith('@'):
            val = node.attrib.get(n[1:])
            if val != None:
                rv.append(n)
                rv.append(val)
        else:
            children = node.findall(n)
            for child_node in children:
                rv.append(n)
                rv.append(child_node.text)
    rv.append('__dictify_end_marker__')
    return rv


# =====================
# Reusable DB-Functions
# =====================
def get_sql_string_update_value(attr_name, new_value):
    if new_value is None or str(new_value) == 'None' or new_value == '':
        ret_val = f"{attr_name} = NULL"
    else:
        ret_val = f"{attr_name} = '{new_value}'"
    return ret_val

def get_sql_float_update_value(attr_name, new_value):
    if new_value is None or str(new_value) == 'None' or new_value == '':
        ret_val = f"{attr_name} = NULL"
    else:
        ret_val = f"{attr_name} = {new_value}"
    return ret_val

def get_sql_datums_update_value(attr_name, new_value, date_str_format='%d.%m.%Y'):
    if new_value is None or str(new_value) == 'None' or new_value == '':
        ret_val = f"{attr_name} = NULL"
    else:
        ret_val = f"{attr_name} = STR_TO_DATE('{new_value}', '{date_str_format}')"
    return ret_val


def remove_all_enum_value_in_set(db, table, attribute_name, enum_val_to_remove, key_attr_name='ID', take_action=False, verbal=False):
    myCursor = db.cursor(dictionary=True)
    records_affected = 0
    if verbal:
        print(f'''
           --> Calling remove_all_enum_value_in_set(db,
                                    table                   = {table}, 
                                    attribute_name          = {attribute_name},
                                    enum_val_to_remove      = {enum_val_to_remove},
                                    key_attr_name           = {key_attr_name},
                                    take_action             = {take_action},                                    
                                    verbal                  = {verbal})''')

    update_person = f"""
        UPDATE {table} AS p1
        JOIN (
            SELECT {key_attr_name}, removeSetValue({attribute_name}, '{enum_val_to_remove}') AS replaceSetValue
            FROM {table}
            WHERE FIND_IN_SET('{enum_val_to_remove}', {attribute_name}) > 0
        ) AS subquery
        ON p1.{key_attr_name} = subquery.{key_attr_name}
        SET p1.{attribute_name} = subquery.replaceSetValue;
     """

    if verbal:
        print(update_person)

    if take_action:
        myCursor.execute(update_person)
        db.commit()
        records_affected = myCursor.rowcount

    return records_affected


def get_record_details_from_db(db, table_name, key_id=None, db_attributes_names=[], as_json=True, take_action=True, verbal=False):
    if verbal:
        print(f'''
           --> Calling get_record_details_from_db(db,
                                    table_name       = {table_name},
                                    key_id           = {key_id},
                                    attributes       = {db_attributes_names},
                                    take_action      = {take_action},
                                    verbal           = {verbal})''')

    myCursor = db.cursor(dictionary=as_json)
    if len(db_attributes_names) == 0:
        fields = '*'
    else:
        fields = ', '.join(db_attributes_names)

    if key_id is not None:
        sql_select = f"SELECT {fields} FROM {table_name} WHERE ID={key_id}"
    else:
        sql_select = f"SELECT {fields} FROM {table_name}"

    if verbal:
        print(sql_select)
    myCursor.execute(sql_select)
    myresult = myCursor.fetchall()
    return myresult


def mysql_db_connect(db_host='localhost', port=3306, db_schema='stammdaten', db_user_name=None, password=None, trace=False):
    if trace:
        print(f"Connecting to '{db_schema:s}@{db_host:s}' with user '{db_user_name:s}'....", end="", flush=True)
    db_connection = mysql.connector.connect(
        host=db_host,
        port=port,
        user=db_user_name,
        password=password,
        database=db_schema,
        auth_plugin='mysql_native_password'
    )
    db_connection.set_charset_collation('utf8mb4', 'utf8mb4_bin')
    if trace:
        print("completed!")
    return db_connection


def do_db_connect(password,
                  user,
                  host='localhost',
                  port=3306,
                  schema='sakila',
                  db_type='MySql', verbal=False):
    if db_type == 'MySql':
        if verbal:
            print(f'Connecting to {user}@{host} --> {schema}...', end='', )
        db_connection = mysql.connector.connect(
            host=host,
            user=user,
            password=password,
            database=schema,
            port=port,
            auth_plugin='mysql_native_password'
        )
        if verbal:
            print('....connected!')

    if db_type == 'Access':
        pass

    return db_connection


def get_record_count(db=None, db_tbl_name=None, retValueWithTblName=True):
    sql_select = f'SELECT count(*) FROM {db_tbl_name} '
    mycursor = db.cursor()
    mycursor.execute(sql_select)
    myresult = mycursor.fetchall()
    if retValueWithTblName:
        return {'rows_in_db   (' + db_tbl_name + '):': myresult[0][0]}
    else:
        return myresult[0][0]


# uses the view Table_Meta_Data
def get_db_attr_type(db, table, attribute, take_action=False, verbal=False):
    if verbal:
        print(f'''
           --> Calling get_db_attr_type(db,
                                    table                  = {table}, 
                                    attribute              = {attribute},
                                    take_action            = {take_action},
                                    verbal                 = {verbal})''')

    select_sql = f"SELECT Attr_Type,Enum_Set_Values FROM Table_Meta_data WHERE `Table` = '{table}' AND `Attribute` = '{attribute}'"
    if verbal:
        print(select_sql)
    mycursor = db.cursor()
    mycursor.execute(select_sql)
    ret_val = {}
    for aRec in mycursor.fetchall():
        ret_val = {
            'type': aRec[0].decode('ascii'),
            'enums': ''  # aRec[1]    # .decode('ascii')
        }
    return ret_val


def get_all_table_names_from_schema(db, schema=None, object_types=None, verbal=False):
    '''
    object_types: BASE TABLE, VIEW
    '''

    if verbal:
        print(f'''
           --> Calling get_all_table_names_from_schema(db,
                        schema        = {schema}, 
                        object_types  = {object_types},
                        verbal        = {verbal})''')

    where_list = []
    if schema is not None:
        where_list.append(f"`TABLE_SCHEMA` = '{schema}'")
    if object_types is not None:
        where_list.append(f"`TABLE_TYPE` = '{object_types}'")
    where_clause = ' AND '.join(where_list)
    if len(where_clause) > 0:
        where_clause = 'WHERE' + where_clause
    select_sql = f'''
        SELECT 
            TABLE_SCHEMA, 
            TABLE_NAME, 
            TABLE_TYPE
        FROM
            INFORMATION_SCHEMA.TABLES
        {where_clause};
    '''
    if verbal:
        print(select_sql)

    mycursor = db.cursor()
    mycursor.execute(select_sql)
    return mycursor.fetchall()


def get_db_attr_type_new(db, schema_name=None, table_name=None, attribute_name=None, take_action=False, verbal=False):
    if verbal:
        print(f'''
           --> Calling get_db_attr_type_new(db,
                                    schema_name            = {schema_name},
                                    table_name             = {table_name}, 
                                    attribute_name         = {attribute_name},
                                    take_action            = {take_action},
                                    verbal                 = {verbal})''')

    where_clause_list = []
    where_clause_str = ''
    if schema_name is not None:
        where_clause_list.append(f"TABLE_SCHEMA = '{schema_name}'")
    if table_name is not None:
        where_clause_list.append(f"TABLE_NAME = '{table_name}'")
    where_clause_str = ' AND '.join(where_clause_list)
    if where_clause_str != '':
        where_clause_str = 'WHERE ' + where_clause_str

    select_sql = f"""
SELECT
	TABLE_SCHEMA   AS TABLE_SCHEMA,
	TABLE_NAME     AS TABLE_NAME,
	COLUMN_NAME    AS Attribute,
	DATA_TYPE      AS Attr_Type,
	COLUMN_KEY     AS Is_key,
	COLUMN_TYPE    AS Attr_Type_Values,
	'TBD' AS Set_Values
FROM INFORMATION_SCHEMA.COLUMNS
{where_clause_str}
    """

    if verbal:
        print(select_sql)
    mycursor = db.cursor()
    mycursor.execute(select_sql)
    ret_val = {}
    for aRec in mycursor.fetchall():
        print(aRec)
    return ret_val


def update_db_attributes(db=None,
                         db_tbl_name=None,
                         id_attr_name='ID', id=None,
                         name_values_to_update=None,
                         take_action=False, verbal=False):
    if verbal:
        print(f'''
            --> Calling update_db_attributes(db,
                           db_tbl_name                 = {db_tbl_name}, 

                           id_attr_name     = {id_attr_name}, 
                           id               = {id},

                           name_values_to_update= {name_values_to_update},

                           take_action      = {take_action},
                           verbal           = {verbal})''')

    return {'dataset_changed': 1, 'attributes_changed': 7}


def update_db_attribute(db=None,
                        db_tbl_name=None, db_attr_name=None, db_attr_type='varchar', db_attr_set_enum_values='',
                        id_attr_name='ID', id=None,
                        new_value=None, new_value_format=None,
                        take_action=False, verbal=False):
    if verbal:
        print(f'''
           --> Calling update_db_attribute(db,
                          db_tbl_name                 = {db_tbl_name}, 
                          db_attr_name                = {db_attr_name},
                          db_attr_type                = {db_attr_type}, 
                          db_attr_set_enum_values     = {db_attr_set_enum_values},

                          id_attr_name     = {id_attr_name}, 
                          id               = {id},

                          new_value        = {new_value},
                          new_value_format = {new_value_format}, 

                          take_action      = {take_action},
                          verbal           = {verbal})''')

    update_count = 0
    sql_select = f'SELECT {db_attr_name} FROM {db_tbl_name} WHERE {id_attr_name} = {id}'
    if verbal:
        print(sql_select)
    mycursor = db.cursor()
    mycursor.execute(sql_select)
    myresult = mycursor.fetchall()
    if len(myresult) == 1:
        old_value = myresult[0][0]
        if verbal:
            print(old_value, '--->', new_value)

        sql_update = None
        if new_value == 'NULL' or new_value == 'None':
            sql_update = f"UPDATE {db_tbl_name} SET {db_attr_name} = NULL WHERE {id_attr_name} = {id}"
        elif old_value != new_value:
            if db_attr_type == 'varchar':
                if new_value[0] == '+' or new_value[0] == ';' or new_value[0] == '|':
                    if len(old_value) > 0:
                        sql_update = f"UPDATE {db_tbl_name} SET {db_attr_name} = CONCAT({db_attr_name}, '{new_value}') WHERE {id_attr_name} = {id}"
                    else:
                        sql_update = f"UPDATE {db_tbl_name} SET {db_attr_name} = '{new_value[1:]}' WHERE {id_attr_name} = {id}"
                else:
                    sql_update = f"UPDATE {db_tbl_name} SET {db_attr_name} = '{new_value}' WHERE {id_attr_name} = {id}"

            elif db_attr_type == 'enum':
                sql_update = f"UPDATE {db_tbl_name} SET {db_attr_name} = '{new_value}' WHERE {id_attr_name} = {id}"

            elif db_attr_type == 'set':
                if new_value[0] == '+':
                    new_value = new_value[1:]
                    sql_update = f"UPDATE {db_tbl_name} SET {db_attr_name} = addSetValue({db_attr_name}, '{new_value}') WHERE {id_attr_name} = {id}"
                elif new_value[0] == '-':
                    new_value = new_value[1:]
                    new_value = remove_set_value(old_value, new_value)
                    if new_value == '':
                        sql_update = f"UPDATE {db_tbl_name} SET {db_attr_name} = NULL WHERE {id_attr_name} = {id}"
                    else:
                        # sql_update = f"UPDATE {db_tbl_name} SET {db_attr_name} = removeSetValue({db_attr_name}, '{new_value}') WHERE {id_attr_name} = {id}"
                        sql_update = f"UPDATE {db_tbl_name} SET {db_attr_name} = '{new_value}' WHERE {id_attr_name} = {id}"
                else:
                    sql_update = f"UPDATE {db_tbl_name} SET {db_attr_name} = '{new_value}' WHERE {id_attr_name} = {id}"

            elif db_attr_type == 'date':
                new_value = str(new_value)[:10]
                date_format_sql = ''
                if is_US_Date(new_value):
                    date_format_sql = '%Y-%m-%d'
                    new_value = re.sub(r'\D+', '-', new_value)
                elif is_EU_Date(new_value):
                    date_format_sql = '%d.%m.%Y'
                    new_value = re.sub(r'\D+', '.', new_value)

                sql_update = f"UPDATE {db_tbl_name} SET {db_attr_name} = STR_TO_DATE('{new_value}', '{date_format_sql}') WHERE {id_attr_name} = {id}"

            else:
                sql_update = f"UPDATE {db_tbl_name} SET {db_attr_name} = {new_value} WHERE {id_attr_name} = {id}"

        if verbal:
            print('sql_update:', sql_update)

        if take_action and sql_update is not None:
            mycursor.execute(sql_update)
            update_count += 1
            db.commit()
    else:
        if verbal:
            print(myresult)

    return update_count


def get_table_records(db, table_name, where_clause=None, as_dictionary=True, take_action=False, verbal=False):
    if verbal:
        print(f'''
           --> Calling get_table_records(db,
                                    table_name             = {table_name}, 
                                    take_action            = {take_action},
                                    verbal                 = {verbal})''')

    select_stmt = f"SELECT * FROM {table_name}"
    if where_clause is not None:
        select_stmt += ' WHERE ' + where_clause
    if verbal:
        print(select_stmt)

    mycursor = db.cursor(dictionary=as_dictionary)
    mycursor.execute(select_stmt)
    my_results = mycursor.fetchall()
    if verbal:
        for a_data_set in my_results:
            print(a_data_set)
        print('Count:', len(my_results))
    return my_results


def create_sql_stmt_from_rs(result_set, table_name='Language', as_csv=False, take_action=False, verbal=False):
    if verbal:
        print(result_set)

    ret_str = ''
    attr_list = []
    for attr_name in result_set[0]:
        attr_list.append('`' + attr_name + '`')

    ret_str += f"""INSERT INTO `{table_name}` ({', '.join(attr_list)}) VALUES \n"""
    for a_tuple in result_set:
        value_list = []
        for attr_name, attr_value in a_tuple.items():
            if verbal:
                print('-->', attr_name, attr_value, type(attr_value))
            if isinstance(attr_value, datetime.datetime):
                attr_value = f"STR_TO_DATE('{attr_value}', '%Y-%m-%d %H:%i:%s')"
            elif isinstance(attr_value, datetime.date):
                attr_value = f"STR_TO_DATE('{attr_value}', '%Y-%m-%d')"
            elif isinstance(attr_value, set):
                attr_value = str(attr_value)
                if attr_value == 'set()':
                    # print(attr_value)
                    attr_value = 'NULL'
                else:
                    attr_value = attr_value.replace("', '", ",")
                    attr_value = attr_value[:-2]  # .replace(r"'{", "")
                    attr_value = attr_value[2:]  # .replace(r"}'", "")
                    attr_value = f"'{attr_value}'"
            elif isinstance(attr_value, str):
                attr_value = attr_value.replace("'", "\\'")
                attr_value = attr_value.replace("\n", "<br/>")
                attr_value = f"'{attr_value}'"
            elif attr_value is None:
                attr_value = 'NULL'
            else:
                attr_value = f"{attr_value}"
            value_list.append(attr_value)
            if verbal:
                print('==>', attr_name, attr_value, type(attr_value), "\n")
        if verbal:
            print('\n\n', value_list)
        ret_str += f"""  ({', '.join(value_list)}),\n"""
    ret_str = ret_str[0:-2] + ";"
    return ret_str


def create_insert_data_stmt_old(db_schema, table_name, where_clause=None):  # replace on 4.6.2024
    data_sets = get_table_records(db_schema, table_name=table_name, where_clause=where_clause, as_dictionary=True, take_action=True, verbal=False)
    if len(data_sets) > 0:
        insert_str = create_sql_stmt_from_rs(
            data_sets,
            table_name=table_name, as_csv=False, take_action=False, verbal=False)
        if where_clause is None:
            where_clause = ''
        else:
            where_clause = f' WHERE {where_clause}'
    else:
        insert_str = ' -- Table is empty'
    insert_str = f"""  -- Extracted at: {getTimestamp()}
    -- SELECT * FROM `{table_name}` {where_clause};    Tuples found: {len(data_sets)}
    {insert_str}"""
    return insert_str


def create_insert_data_stmt(db_schema, table_name, where_clause=None, fields_to_select=None, fields_to_hash=None, verbal=False):
    my_cursor = db_schema.cursor(dictionary=True)

    if where_clause is None:
        where_clause = ''
    else:
        where_clause = f' WHERE {where_clause}'

    if fields_to_select is None:
        fields_to_select = '*'
    else:
        fields_to_select = ','.join(fields_to_select)

    if fields_to_hash is None:
        fields_to_hash = []

    select_statement = f'SELECT {fields_to_select} FROM `{table_name}`{where_clause};'
    if verbal:
        print(f'    --> {select_statement}')
    my_cursor.execute(select_statement)
    result_set = my_cursor.fetchall()
    count_of_records = len(result_set)

    if count_of_records > 0:
        attribute_list = list(result_set[0].keys())
        attributes = str(attribute_list).replace('[', '').replace(']', '').replace("'", '`')
        ret_str = f'''
    INSERT INTO `{table_name}` ({attributes}) VALUES'''

        insert_tuple_str = ''
        for a_tuple in result_set:
            a_tuple_str = '        ('
            for a_attr in attribute_list:
                value_of_attr = a_tuple[a_attr]
                type_of_attr = type(value_of_attr)

                if a_attr in fields_to_hash:
                    if value_of_attr is not None and len(value_of_attr) > 0:
                        if is_email(value_of_attr):
                            email_part_0 = value_of_attr.split('@')[0]
                            email_part_1 = value_of_attr.split('@')[1]
                            value_of_attr_hash = hash_string(email_part_0, ret_length=len(email_part_0))
                            value_of_attr_hash_1 = value_of_attr[0] + value_of_attr_hash[1:] + '@' + email_part_1
                            if verbal:
                                print(f'         {a_attr}={value_of_attr} EMAIL: {value_of_attr_hash} : {value_of_attr_hash_1}')
                            value_of_attr = value_of_attr_hash_1
                        else:
                            value_of_attr_hash = hash_string(value_of_attr, ret_length=len(value_of_attr))
                            value_of_attr_hash_1 = value_of_attr[0] + value_of_attr_hash[1:]
                            if verbal:
                                print(f'         {a_attr}={value_of_attr} HASH: {value_of_attr_hash} : {value_of_attr_hash_1}')
                            value_of_attr = value_of_attr_hash_1
                # print(f'{a_attr}={value_of_attr}  [{type_of_attr}]')
                if isinstance(value_of_attr, datetime.datetime):
                    a_tuple_str += f"STR_TO_DATE('{value_of_attr}', '%Y-%m-%d %H:%i:%s')" + ', '

                elif isinstance(value_of_attr, datetime.date):
                    a_tuple_str += f"STR_TO_DATE('{value_of_attr}', '%Y-%m-%d')" + ', '

                elif isinstance(value_of_attr, set):
                    value_of_attr = str(value_of_attr)
                    if value_of_attr == 'set()':
                        # print(value_of_attr)
                        a_tuple_str += 'NULL' + ', '
                    else:
                        value_of_attr = value_of_attr.replace("', '", ",")
                        value_of_attr = value_of_attr[:-2]  # .replace(r"'{", "")
                        value_of_attr = value_of_attr[2:]  # .replace(r"}'", "")
                        value_of_attr = value_of_attr.replace("'", "'")
                        a_tuple_str += f"'{value_of_attr}'" + ', '

                elif isinstance(value_of_attr, str):
                    value_of_attr = value_of_attr.replace("'", r"\'")
                    a_tuple_str += f"'{value_of_attr}'" + ', '

                elif value_of_attr is None:
                    a_tuple_str += 'NULL' + ', '

                else:
                    a_tuple_str += f"{value_of_attr}" + ', '

            a_tuple_str = a_tuple_str[:-2]
            a_tuple_str += '),'
            insert_tuple_str += '\n' + a_tuple_str
        insert_tuple_str = insert_tuple_str[:-1]
        # print(insert_tuple_str)
        ret_str += insert_tuple_str + ';'
    else:
        ret_str = f' -- WARNING: Table `{table_name}` is empty!'
    if verbal:
        print('\n')
    return f'''
    -- -------------------------------
    -- {select_statement}   --> {count_of_records} records found
    -- Extracted at: {getTimestamp()}{ret_str}

  '''


# ========================
# DB unload/load-functions
# ========================
def format_sql_stmt(sql_statement, indent=4):
    sql_statement = sql_statement.replace('\n', ' ')
    sql_statement = sql_statement.replace('  ', ' ')
    sql_statement = sql_statement.replace('  ', ' ')
    sql_statement = sql_statement.replace('  ', ' ')
    sql_statement = sql_statement.replace('  ', ' ')
    sql_statement = sql_statement.replace(' ;', ';')
    if isinstance(indent, int):
        indent = f'{" ":{indent}s}'
    return indent + sql_statement


def get_all_table_names_from_schema(db, schema=None, object_types=None, verbal=False):
    '''
    object_types: BASE TABLE, VIEW
    '''

    if verbal:
        print(f'''
           --> Calling get_all_table_names_from_schema(db,
                        schema        = {schema}, 
                        object_types  = {object_types},
                        verbal        = {verbal})''')

    where_list = []
    if schema is not None:
        where_list.append(f"`TABLE_SCHEMA` = '{schema}'")
    if object_types is not None:
        where_list.append(f"`TABLE_TYPE` = '{object_types}'")
    where_clause = ' AND '.join(where_list)
    if len(where_clause) > 0:
        where_clause = 'WHERE' + where_clause
    select_sql = f'''
        SELECT 
            TABLE_SCHEMA, 
            TABLE_NAME, 
            TABLE_TYPE
        FROM
            INFORMATION_SCHEMA.TABLES
        {where_clause};
    '''

    if verbal:
        print(select_sql)

    mycursor = db.cursor()
    mycursor.execute(select_sql)
    return mycursor.fetchall()


def get_possible_enum_values_from_db_attribute(db, schema=None, table_name=None, column_name=None, verbal=False):
    """
    Fetches all possible values for a SET column in a MySQL table.

    Parameters:
    - table_name (str): Name of the table containing the SET column.
    - column_name (str): Name of the SET column.

    Returns:
    - list of str: Possible values in the SET column.
    """

    if verbal:
        print(f'''
           --> Calling get_possible_set_values_from_db(db,
                        schema      = {schema}, 
                        table_name  = {table_name},
                        column_name = {column_name},
                        verbal      = {verbal})''')
    select_sql = f'''
        SELECT 
            COLUMN_TYPE
        FROM 
            INFORMATION_SCHEMA.COLUMNS
        WHERE 
            TABLE_SCHEMA = '{schema}'     AND
            TABLE_NAME   = '{table_name}' AND
            COLUMN_NAME  = '{column_name}';
    '''

    my_cursor = db.cursor(dictionary=True)
    if verbal:
        print(select_sql)
    my_cursor.execute(select_sql)
    result = my_cursor.fetchone()['COLUMN_TYPE']

    if verbal:
        print('result --> ', result)

    if result:
        # Use regex to extract values from the SET definition
        column_type = result.decode() if isinstance(result, bytes) else result
        if verbal:
            column_type
        values = re.findall(r"'(.*?)'", column_type)
        values.sort()
        if verbal:
            print('--->', values)
        return values
    else:
        return []  # No result found


def convert_resultSet_to_insertSQL(table_name, result_set=None, fields_to_hash=None, verbal=False):
    if fields_to_hash is None:
        fields_to_hash = []

    if result_set is not None and len(result_set) > 0:
        insert_str = f"INSERT INTO `{table_name}` ("
        attr_list = list(result_set[0].keys())
        for a_attr_name in attr_list:
            insert_str += f"`{a_attr_name}`, "

        insert_str = insert_str[:-2] + ") VALUES\n"

        for a_tuple in result_set:
            tuple_str = '('
            for an_attr_name in attr_list:
                an_attr_value = a_tuple[an_attr_name]

                if an_attr_name in fields_to_hash:
                    if an_attr_value is not None and len(an_attr_value) > 0:
                        if is_email(an_attr_value):
                            email_part_0 = an_attr_value.split('@')[0]
                            email_part_1 = an_attr_value.split('@')[1]
                            value_of_attr_hash = hash_string(email_part_0, ret_length=len(email_part_0))
                            value_of_attr_hash_1 = an_attr_value[0] + value_of_attr_hash[1:] + '@' + email_part_1
                            an_attr_value = value_of_attr_hash_1
                        else:
                            value_of_attr_hash = hash_string(an_attr_value, ret_length=len(an_attr_value))
                            value_of_attr_hash_1 = an_attr_value[0] + value_of_attr_hash[1:]
                            if verbal:
                                print(f'         {an_attr_name}={an_attr_value} HASH: {value_of_attr_hash} : {value_of_attr_hash_1}')
                            an_attr_value = value_of_attr_hash_1

                if isinstance(an_attr_value, str):
                    an_attr_value = an_attr_value.replace("'", r"\'")
                    tuple_str += f"'{an_attr_value}'" + ', '

                elif an_attr_value is None:
                    tuple_str += 'NULL' + ', '

                elif isinstance(an_attr_value, datetime):
                    tuple_str += f"STR_TO_DATE('{an_attr_value}', '%Y-%m-%d %H:%i:%s')" + ', '

                elif isinstance(an_attr_value, date):
                    tuple_str += f"STR_TO_DATE('{an_attr_value}', '%Y-%m-%d')" + ', '

                elif isinstance(an_attr_value, set):
                    an_attr_value = str(an_attr_value)
                    if an_attr_value == 'set()':
                        # print(attr_value)
                        tuple_str += 'NULL' + ', '
                    else:
                        an_attr_value = an_attr_value.replace("', '", ",")
                        an_attr_value = an_attr_value[:-2]  # .replace(r"'{", "")
                        an_attr_value = an_attr_value[2:]  # .replace(r"}'", "")
                        # attr_value = attr_value.replace("'", "'")
                        an_attr_value = f"'{an_attr_value}'" + ', '
                        tuple_str += an_attr_value
                else:
                    tuple_str += f"{str(an_attr_value)}, "

            tuple_str = "   " + tuple_str[:-2] + "),\n"
            insert_str += tuple_str
        insert_str = insert_str[:-2] + ";"
    else:
        insert_str = ""
    return insert_str


def do_prepare_db_attributes(attributes_to_select=None, indent=4):
    if isinstance(indent, int):
        indent = f'{" ":{indent}s}'

    ret_str = ''
    if attributes_to_select is None:
        ret_str = "*"
    elif isinstance(attributes_to_select, str):
        ret_str = attributes_to_select
    elif isinstance(attributes_to_select, list):
        ret_str = f',\n{indent}'.join(attributes_to_select)

    ret_str = f'\n{indent}{ret_str}'
    return ret_str


def do_prepare_order_by(order_by=None, indent=4):
    if isinstance(indent, int):
        indent = f'{" ":{indent}s}'

    ret_str = ''
    if order_by is None:
        ret_str = ""
    elif isinstance(order_by, str):
        ret_str = order_by
    elif isinstance(order_by, list):
        ret_str = f',\n{indent}'.join(order_by)

    if ret_str != '':
        ret_str = '\nORDER BY\n' + indent + ret_str
    return ret_str


def select_data_from_db_table(db_connection,
                              table_name,

                              attribute_list=None,
                              where_clause=None,
                              order_by_list=None,

                              db_type='MySql',
                              indent=4,
                              dictionary=True,
                              verbal=False):
    if isinstance(indent, int):
        indent = f'{" ":{indent}s}'
        # print(':', indent, ':', sep='')

    if where_clause is None:
        where_clause = ''
    else:
        where_clause = f"\nWHERE\n{indent}{where_clause}"

    sql_statement = f"""SELECT {do_prepare_db_attributes(attribute_list, indent=len(indent))}\nFROM {table_name} {where_clause} {do_prepare_order_by(order_by_list, indent=len(indent))};"""

    sql_statement = format_sql_stmt(sql_statement)

    if verbal:
        print(sql_statement)

    my_cursor = db_connection.cursor(dictionary=dictionary)
    my_cursor.execute(sql_statement)
    my_resultset = my_cursor.fetchall()
    if verbal:
        print(f"{indent}{len(my_resultset)} record(s) found")
    return {
        'timestamp': f'{datetime.now():%Y-%m-%d %H:%M:%S}',
        'count': len(my_resultset),
        'select': sql_statement,
        'rs': my_resultset
    }


def unload_data_from_db_table(db_connection,
                              table_name,
                              attribute_list=None,
                              where_clause=None,
                              order_by_list=None,
                              fields_to_hash=None,
                              verbal=False
                              ):
    rs = select_data_from_db_table(db_connection,
                                   table_name=table_name,
                                   attribute_list=attribute_list,
                                   where_clause=where_clause,
                                   order_by_list=order_by_list,
                                   verbal=verbal
                                   )
    insert_string = convert_resultSet_to_insertSQL(
        table_name,
        rs['rs'],
        fields_to_hash=fields_to_hash,
        verbal=verbal
    )

    insert_string = f"""
/*
-- Extracted at: {rs['timestamp']}
-- Count       : {rs['count']}
-- {rs['select'].strip()}
*/
{insert_string}

    """
    return insert_string


def get_exception_for_table_unload(table_name, exception_list, verbal=False):
    if verbal:
        print(f'''
        get_exception_for_table_unload('{table_name}', {exception_list}, {verbal} has been called)
        ''')

    ret_exception = None
    if exception_list is not None and len(exception_list) > 0:
        for an_exception in exception_list:
            if an_exception['table'].lower() == table_name.lower():
                ret_exception = an_exception
                break

    return ret_exception


def unload_all_data_from_schema(db_connection,
                                schema_name=None,
                                tables_to_read_with_exceptions=None,
                                verbal=False):
    obj_type = 'BASE TABLE'
    all_tables = get_all_table_names_from_schema(db_connection, schema=schema_name, object_types=obj_type, verbal=False)
    if verbal:
        print(f'{len(all_tables)} "{obj_type}" found in "{schema_name}"!\n')

    max_len = 0
    for a_table_name in all_tables:
        length = len(a_table_name[1])
        if length > max_len:
            max_len = length
    max_table_name_length = f'{max_len}s'

    insert_string = ''
    for a_table in all_tables:
        if verbal:
            print(f'--> {a_table[0]}::{a_table[1]}')
        exception = get_exception_for_table_unload(a_table[1], tables_to_read_with_exceptions, verbal=False)

        if verbal and exception is not None:
            print(f'    Exceptions for "{a_table[1]}" found\n          {exception}\n')

        if exception is None:
            fields_to_select = None
            order_by_list = None
            where_clause = None
            do_verbal = None
            fields_to_hash = None
        else:
            if exception.get('action') == 'Exclude':
                print(f'WARNING: {a_table[1]:{max_table_name_length}} Data not extracted!\n')
                continue
            fields_to_select = exception.get('fields')
            order_by_list = exception.get('order_by_list')
            where_clause = exception.get('where_clause')
            do_verbal = exception.get('verbal')
            fields_to_hash = exception.get('fields_to_hash')
            print(f'WARNING: {a_table[1]:{max_table_name_length}} Data with exceptions extracted\n')

        insert_string += "\n\n" + unload_data_from_db_table(
            db_connection,
            table_name=a_table[1],
            attribute_list=fields_to_select,
            where_clause=where_clause,
            order_by_list=order_by_list,
            fields_to_hash=fields_to_hash,
            verbal=verbal
        )
    return insert_string


def get_enum_values_from_db_attribute(db_connection, schema_name=None, attribute_name=None, verbal=False):
    print(f'get_enum_values_from_db_attribute(schema_name={schema_name}, attribute_name={attribute_name})')
    return ['Bürger', 'Hat_16a']


# ------------------------
# Reusable Excel-Functions
# ------------------------
def is_not_empty(value):
    return value is not None and value != ''


def are_all_values_empty(ws, title_row=1, row=1, exclude_attr_names=[]):
    column_header_list = get_existing_column_titles(ws, title_row=title_row, exclude_attr_names=exclude_attr_names, key_only=False)
    ret_value = get_cell_values_by_column_titles(ws, title_row=title_row, row=row, column_names=column_header_list, do_reset_cell=False, reset_cell_value=None, take_action=False, verbal=False)

    non_empty_values = {key: value for key, value in ret_value.items() if is_not_empty(value)}
    # print(non_empty_values)
    if len(non_empty_values) == 0:
        return True
    else:
        return False


def get_existing_column_titles(ws, title_row=1, exclude_attr_names=[], key_only=False):
    ret_value = {cell.value: cell.column for cell in ws[title_row] if cell.value not in exclude_attr_names}
    if key_only:
        ret_value = list(ret_value.keys())
    return ret_value


def get_cell_values_by_column_titles(ws, title_row=1, row=1, column_names=[], do_reset_cell=False, reset_cell_value=None, take_action=False, verbal=False):
    if verbal:
        print(f'''
           --> Calling get_cell_values_by_column_titles(ws,
                                    title_row        = {title_row}, 
                                    row              = {row},
                                    column_name      = {column_names},
                                    do_reset_cell    = {do_reset_cell},
                                    reset_cell_value = {reset_cell_value},
                                    take_action      = {take_action},
                                    verbal           = {verbal})''')

    ret_values = {}
    for column_name in column_names:
        value = get_cell_value_by_column_title(ws, title_row=title_row, row=row, column_name=column_name, do_reset_cell=do_reset_cell, reset_cell_value=reset_cell_value, take_action=take_action, verbal=verbal)[column_name]
        if value is None:
            ret_values[column_name] = ''
        else:
            ret_values[column_name] = value
    return ret_values


def get_cell_value_by_column_title(ws, title_row=1, row=1, column_name='ID', do_reset_cell=False, reset_cell_value=None, take_action=False, verbal=False):
    if verbal:
        print(f'''
           --> Calling get_cell_value_by_column_title(ws,
                                    title_row        = {title_row}, 
                                    row              = {row},
                                    column_name      = {column_name},
                                    do_reset_cell    = {do_reset_cell},
                                    reset_cell_value = {reset_cell_value},
                                    take_action      = {take_action},
                                    verbal           = {verbal})''')

    column_title_to_index = {cell.value: cell.column for cell in ws[title_row]}
    if False:
        print('column_title_to_index:', column_title_to_index)
    column_letter = openpyxl.utils.get_column_letter(column_title_to_index[column_name])
    cell_value = ws[column_letter + str(row)].value
    if verbal:
        print(f"{column_name}  --> ws['{column_letter}{str(row)}'].value = {cell_value}")

    if do_reset_cell:
        ws[column_letter + str(row)].value = reset_cell_value
        return {column_name: cell_value,
                'cell_column': column_letter,
                'cell_row': row,
                f'ws["{column_letter}{str(row)}"].value': cell_value,
                'cell_value_reseted_to': reset_cell_value}
    else:
        return {column_name: cell_value,
                'cell_column': column_letter,
                'cell_row': row,
                f'ws["{column_letter}{str(row)}"].value': cell_value}


def set_cell_value_by_column_title(new_cell_value, ws, title_row=1, row=1, column_name='ID', take_action=False, verbal=False):
    if verbal:
        print(f'''
           --> Calling set_cell_value_by_column_title({new_cell_value},
                                    ws,
                                    title_row={title_row}, 
                                    row={row},
                                    column_name={column_name},
                                    verbal={verbal})''')

    column_title_to_index = {cell.value: cell.column for cell in ws[title_row]}
    column_letter = openpyxl.utils.get_column_letter(column_title_to_index[column_name])
    old_cell_value = ws[column_letter + str(row)].value
    if verbal:
        print(f"""
             Old value of {column_name} --> ws['{column_letter}{str(row)}'].value = {old_cell_value}
             New value of {column_name} --> ws['{column_letter}{str(row)}'].value = {new_cell_value}
             """)
    if old_cell_value is None and new_cell_value is not None:
        action = 'create'
    elif old_cell_value != new_cell_value:
        action = 'update'
    elif new_cell_value is None and old_cell_value is not None:
        action = 'deleted'
    else:
        action = 'no change'

    if take_action and action != 'no change':
        ws[column_letter + str(row)].value = new_cell_value
    return {'Old value:' + column_name: old_cell_value,
            'New value:' + column_name: new_cell_value,
            'action': action,
            'cell_col': column_letter,
            'cell_row': row,
            f'ws["{column_letter}{str(row)}"].value': new_cell_value}


# ========================
# PAIN001 Functions (Zahlungsauftrags-Fileformat)
# ========================
def get_pain001_template(use_iso_norm=True):
    pain001_template = '''
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Document xmlns="http://www.six-interbank-clearing.com/de/pain.001.001.03.ch.02.xsd">
    <CstmrCdtTrfInitn>
        <GrpHdr>
            <MsgId>{{ Debitor_Info.Header_ID }}</MsgId>
            <CreDtTm>{{ Debitor_Info.Creation_Time }}</CreDtTm>
            <NbOfTxs>{{ Zahlungs_liste|length }}</NbOfTxs>
            <CtrlSum>{{ summery_of_payments.total_amount }}</CtrlSum>
            <InitgPty>
                <Nm>{{ Debitor_Info.Name.strip() }}</Nm>
            </InitgPty>
        </GrpHdr>
        <PmtInf>
            <PmtInfId>{{ summery_of_payments.pain_id }}</PmtInfId>
            <PmtMtd>TRF</PmtMtd>
            <NbOfTxs>{{ summery_of_payments.count_of_payments }}</NbOfTxs>
            <PmtTpInf>
                <SvcLvl>
                    <Cd>SEPA</Cd>
                </SvcLvl>
            </PmtTpInf>
            <ReqdExctnDt>{{ Debitor_Info.Valuta }}</ReqdExctnDt>
            <Dbtr>
                <Nm>{{ Debitor_Info.Name.strip() }}</Nm>
                <PstlAdr>
                    <Ctry>{{ Debitor_Info.Country_Code.strip().upper() }}</Ctry>
                </PstlAdr>
            </Dbtr>
            <DbtrAcct>
                <Id>
                    <IBAN>{{ Debitor_Info.IBAN.replace(' ','').upper()  }}</IBAN>
                </Id>
            </DbtrAcct>
            <DbtrAgt>
                <FinInstnId>
                    <BIC>{{ Debitor_Info.BIC }}</BIC>
                </FinInstnId>
            </DbtrAgt>

            {% for a_payment in Zahlungs_liste %}
            <CdtTrfTxInf>
                <PmtId>
                    <InstrId>{{ loop.index}}</InstrId>
                    <EndToEndId>{{ loop.index + summery_of_payments.pain_start_id }}</EndToEndId>
                </PmtId>
                <Amt>
                    <InstdAmt Ccy="{{ a_payment.Ccy.replace(' ','').upper()  }}">{{ a_payment.Amount }}</InstdAmt>
                </Amt>
                <Cdtr>
                    <Nm>{{ a_payment.Receiver_Name.strip()  }}</Nm>
                </Cdtr>
                <CdtrAcct>
                    <Id>
                        <IBAN>{{ a_payment.IBAN.replace(' ','').upper() }}</IBAN>
                    </Id>
                </CdtrAcct>
                <RmtInf>
                    <Ustrd>{{ a_payment.Reason }}</Ustrd>
                </RmtInf>
            </CdtTrfTxInf>
            {% endfor %}
        </PmtInf>
    </CstmrCdtTrfInitn>
</Document>     
    '''

    iso20022_template_old = '''
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Document xmlns="urn:iso:std:iso:20022:tech:xsd:pain.001.001.03">
    <CstmrCdtTrfInitn>
        <GrpHdr>
            <MsgId>{{ Debitor_Info.Header_ID }}</MsgId>
            <CreDtTm>{{ Debitor_Info.Creation_Time }}</CreDtTm>
            <NbOfTxs>{{ Zahlungs_liste|length }}</NbOfTxs>
            <CtrlSum>{{ summery_of_payments.total_amount }}</CtrlSum>
            <InitgPty>
                <Nm>{{ Debitor_Info.Name.strip() }}</Nm>
            </InitgPty>
        </GrpHdr>
        <PmtInf>
            <PmtInfId>{{ summery_of_payments.payment_info_id }}</PmtInfId>
            <PmtMtd>TRF</PmtMtd>
            <BtchBookg>true</BtchBookg>
            <NbOfTxs>{{ summery_of_payments.count_of_payments }}</NbOfTxs>
            <CtrlSum>{{ summery_of_payments.total_amount }}</CtrlSum>
            <PmtTpInf>
                <InstrPrty>NORM</InstrPrty>
                <SvcLvl>
                    <Cd>SEPA</Cd>
                </SvcLvl>
            </PmtTpInf>
            <ReqdExctnDt>{{ Debitor_Info.Execution_Date }}</ReqdExctnDt>
            <Dbtr>
                <Nm>{{ Debitor_Info.Name.strip() }}</Nm>
                <PstlAdr>
                    <Ctry>{{ Debitor_Info.Country_Code.strip().upper() }}</Ctry>
                </PstlAdr>
            </Dbtr>
            <DbtrAcct>
                <Id>
                    <IBAN>{{ Debitor_Info.IBAN.replace(' ', '').upper() }}</IBAN>
                </Id>
            </DbtrAcct>
            <DbtrAgt>
                <FinInstnId>
                    <BIC>{{ Debitor_Info.BIC }}</BIC>
                </FinInstnId>
            </DbtrAgt>
            <ChrgBr>SLEV</ChrgBr>

            {% for a_payment in Zahlungs_liste %}
            <CdtTrfTxInf>
                <PmtId>
                    <InstrId>{{ loop.index }}</InstrId>
                    <EndToEndId>{{ loop.index + summery_of_payments.pain_start_id }}</EndToEndId>
                </PmtId>
                <Amt>
                    <InstdAmt Ccy="{{ a_payment.Ccy.replace(' ', '').upper() }}">{{ a_payment.Amount }}</InstdAmt>
                </Amt>
                <CdtrAgt>
                    <FinInstnId>
                        <BIC>{{ a_payment.Creditor_BIC }}</BIC>
                    </FinInstnId>
                </CdtrAgt>
                <Cdtr>
                    <Nm>{{ a_payment.Receiver_Name.strip() }}</Nm>
                    <PstlAdr>
                        <Ctry>{{ a_payment.Receiver_Country_Code.strip().upper() if a_payment.Receiver_Country_Code else "CH" }}</Ctry>
                    </PstlAdr>
                </Cdtr>
                <CdtrAcct>
                    <Id>
                        <IBAN>{{ a_payment.IBAN.replace(' ', '').upper() }}</IBAN>
                    </Id>
                </CdtrAcct>
                <RmtInf>
                    <Ustrd>{{ a_payment.Remittance_Info }}</Ustrd>
                </RmtInf>
            </CdtTrfTxInf>
            {% endfor %}
        </PmtInf>
    </CstmrCdtTrfInitn>
</Document>
    '''

    iso20022_template = '''

<?xml version="1.0" encoding="UTF-8"?>
<Document xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns="http://www.six-interbank-clearing.com/de/pain.001.001.03.ch.02.xsd">
  <CstmrCdtTrfInitn>
    <GrpHdr>
      <MsgId>{{ Debitor_Info.Header_ID }}</MsgId>
      <CreDtTm>{{ Debitor_Info.Creation_Time }}</CreDtTm>
      <NbOfTxs>{{ Zahlungs_liste|length }}</NbOfTxs>
      <CtrlSum>{{ summery_of_payments.total_amount }}</CtrlSum>
      <InitgPty>
        <Nm>{{ Debitor_Info.Name.strip() }}</Nm>
        <CtctDtls>
          <Nm>Walter Rothlin, Peterliwiese 33, 8855 Wangen SZ</Nm>
          <Othr>Version 3.1.0.4</Othr>
        </CtctDtls>
      </InitgPty>
    </GrpHdr>
    <PmtInf>
      <PmtInfId>{{ summery_of_payments.pain_id }}</PmtInfId>
      <PmtMtd>TRF</PmtMtd>
      <BtchBookg>true</BtchBookg>
      <NbOfTxs>{{ summery_of_payments.count_of_payments }}</NbOfTxs>
      <CtrlSum>{{ summery_of_payments.total_amount }}</CtrlSum>
      <ReqdExctnDt>{{ Debitor_Info.Valuta }}</ReqdExctnDt>
      <Dbtr>
        <Nm>{{ Debitor_Info.Name.strip() }}</Nm>
      </Dbtr>
      <DbtrAcct>
        <Id>
          <IBAN>{{ Debitor_Info.IBAN.replace(' ','').upper()  }}</IBAN>
        </Id>
      </DbtrAcct>
      <DbtrAgt>
        <FinInstnId>
          <ClrSysMmbId>
            <ClrSysId>
              <Cd>CHBCC</Cd>
            </ClrSysId>
            <MmbId>777</MmbId>
          </ClrSysMmbId>
        </FinInstnId>
      </DbtrAgt>

        {% for a_payment in Zahlungs_liste %}
        <CdtTrfTxInf>
            <PmtId>
                <InstrId>{{ loop.index }}</InstrId>
                <EndToEndId>{{ loop.index + summery_of_payments.pain_start_id }}</EndToEndId>
            </PmtId>
            <Amt>
                <InstdAmt Ccy="{{ a_payment.Ccy.replace(' ','').upper() }}">{{ a_payment.Amount }}</InstdAmt>
            </Amt>
            <Cdtr>
                <Nm>{{ a_payment.Receiver_Name.strip() }}</Nm>
                <PstlAdr>
                    <StrtNm>{{ a_payment.Receiver_Strasse.strip() }}</StrtNm>
                    <PstCd>{{ a_payment.Receiver_PLZ.strip() }}</PstCd>
                    <TwnNm>{{ a_payment.Receiver_Ort.strip() }}</TwnNm>
                    <Ctry>CH</Ctry>
                </PstlAdr>
            </Cdtr>
            <CdtrAcct>
                <Id>
                    <IBAN>{{ a_payment.IBAN.replace(' ','').upper() }}</IBAN>
                </Id>
            </CdtrAcct>
            <RmtInf>
                <Ustrd>{{ a_payment.Reason }}</Ustrd>
            </RmtInf>
        </CdtTrfTxInf>
        {% endfor %}
    </PmtInf>
  </CstmrCdtTrfInitn>
</Document>    

    '''

    if use_iso_norm:
        return_tmpl = iso20022_template
    else:
        return_tmpl = pain001_template
    return return_tmpl


def get_pain001_msg(debitor_info,
                    payment_list,
                    template_path='./Templates',
                    template_filename=None,
                    pain_id=f'{datetime.utcnow().strftime("%y%m%d%H%M%S")}',
                    pain_start_id=int(f'{datetime.utcnow().strftime("%d%m%y%H%M%S")}')):
    total_amount = 0
    for a_payment in payment_list:
        total_amount += float(a_payment['Amount'])

    summery_of_payments = {
        'count_of_payments': len(payment_list),
        'total_amount': f'{str(round(total_amount, 2))}',
        'pain_id': pain_id,
        'pain_start_id': pain_start_id,
    }

    if template_filename is not None:
        env = Environment(loader=FileSystemLoader(template_path))
        template = env.get_template(template_filename)
    else:
        template = Template(get_pain001_template())

    return template.render(summery_of_payments=summery_of_payments, Zahlungs_liste=payment_list, Debitor_Info=debitor_info)


def TEST_Pain001():
    Debitor_Info_Geno = {
        'Header_ID': '547362488991',
        'Creation_Time': f'{datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%f%z")}',  # 2023-11-17T14:48:04.485+01:00
        'Name': 'Genossame Wangen SZ',
        'Country_Code': 'CH',
        'IBAN': 'CH5100777001561771945',
        'BIC': 'KBSZCH22XXX',
        'Valuta': '2024-11-02'
    }

    Debitor_Info_WR = {
        'Header_ID': '547362488991',
        'Creation_Time': f'{datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%f%z")}',  # 2023-11-17T14:48:04.485+01:00
        'Name': 'Walter Rothlin',
        'Country_Code': 'CH',
        'IBAN': 'CH4900777004546031980',  # Börsenkonto SZKB
        'BIC': 'KBSZCH22XXX',
        'Valuta': '2024-10-21'
    }

    Zahlungs_liste = [
        {'Ccy': 'CHF',
         'Amount': '2.95',
         'Receiver_Name': 'Tobias Rothlin',
         'IBAN': 'CH40 0077 7003 6561 2009 5',  # Haushaltsgeld SZKB
         'Reason': 'Test von PAIN001 (2.95)'},
        {'Ccy': 'chf',
         'Amount': '1.05',
         'Receiver_Name': '  Tobias Rothlin  ',
         'IBAN': 'ch40 0077 7003 6561 2009 5',  # Haushaltsgeld SZKB
         'Reason': 'Test von PAIN001 (1.05)'},
    ]

    pain001_msg = get_pain001_msg(debitor_info=Debitor_Info_WR, payment_list=Zahlungs_liste)
    print(pain001_msg)
    with open('generated_pain001.xml', 'w') as file:
        # Write the string to the file
        file.write(pain001_msg.strip())


# ===========================================================
# MAIN
# ===========================================================

if __name__ == '__main__':

    autoTest = True

    if not autoTest:
        pass  # NOP in Python
        # AUTO_TEST_is_password_valid()
        # AUTO_TEST_xPath_Get(verbal=True)
        # AUTO_TEST__split_adress_street_nr()
        # TEST_stringFct()
        # TEST_hexStrToURLEncoded()
        # TEST_getTimestamp()
        # TEST_readln(verbal=True)
        # TEST_get_sup_super()
        # TEST_calcNulstellen()
        # TEST_printProgressBar()
        # TEST_getMenuFromList()
        # TEST_Pain001()
        # TEST__format_float_1()

    # Automated Tests
    # ===============
    if autoTest:
        auto_test_suiteNameLength = 40
        auto_test_testStatistics_anzStellen = 4
        auto_test_testStatistics_totalLength = 107
        auto_test_fct_prefix = "AUTO_TEST_"
        auto_test_fct_prefix_len = len(auto_test_fct_prefix)

        doVerbal = True
        totalTests = [0, 0]
        testStat = AUTO_TEST_pysikalische_umrechnungen(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_a_summeBis_MitFormel(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_a_summeBis_MitLoop(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_a_summe(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_a_generateStringRepeats(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_a_unterstreichen(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_summe(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_fakultaet(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_average(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_FileFunctions(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_a_File_addHeader(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_addParity(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_a_isFloatEquals(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_getRange(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST_CryptDecrypt(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        testStat = AUTO_TEST__format_float(verbal=doVerbal)
        totalTests[0] += testStat[0]
        totalTests[1] += testStat[1]

        deleteDir("./TestData")
        if doVerbal:
            print(generateStringRepeats(auto_test_testStatistics_totalLength, '-'))
            print("===> ", ("{v:" + str(auto_test_suiteNameLength) + "s}").format(v="Total:"), "Tests Performed:",
                  ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=totalTests[0]),
                  "      Tests Failed:",
                  ("{v:" + str(auto_test_testStatistics_anzStellen) + "d}").format(v=totalTests[1]),
                  "    Passed:{v:7.1f}".format(v=round(100 - (100 * totalTests[1] / totalTests[0]), 1)), "%", sep="")
            print(generateStringRepeats(auto_test_testStatistics_totalLength, '='))

        print("\n\n")
        ex_TestFaelle = totalTests[0]
        print("Ihre provisorisches Prüfungsbewertung:")
        print("     Total zu implementierende Testfälle: {p:3d}".format(p=ex_TestFaelle))
        print("     Testfälle implementiert            : {p:3d}".format(p=totalTests[0]))
        print("     Testfälle failed                   : {p:3d}".format(p=totalTests[1]))
        p_testabdeckung = 100 * totalTests[0] / ex_TestFaelle
        p_implement = 100 * (totalTests[0] - totalTests[1]) / totalTests[0]
        p_TotalPunkte = p_testabdeckung + p_implement
        print("\n")
        print("     Punkte für Testabdeckung        : {p:6.2f}".format(p=p_testabdeckung))
        print("     Punkte für Implementierung      : {p:6.2f}".format(p=p_implement))
        print("\n")
        print("     Total Punkte         : {p:6.2f}".format(p=p_TotalPunkte))
        print("     Provisorische Note   : {p:5.1f}".format(p=(5 / 200) * p_TotalPunkte + 1))
        print("                            =======")
