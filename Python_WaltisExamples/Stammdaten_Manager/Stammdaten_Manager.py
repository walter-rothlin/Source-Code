#!/usr/bin/python3

# ------------------------------------------------------------------
# Name  : Stammdaten_Manager.py
# Source: https://raw.githubusercontent.com/walter-rothlin/Source-Code/master/Python_WaltisExamples/Stammdaten_Manager/Stammdaten_Manager.py
#
# Description: Manager für Stammdaten
#
# Autor: Walter Rothlin
#
# History:
# 01-Jan-2023   Walter Rothlin      Initial Version
# ------------------------------------------------------------------
import mysql.connector
import sqlparse
import csv
import json
from openpyxl import load_workbook
from waltisLibrary import *

# Lambda function to check if a x is from WAHR / FALSCH.
ifTrue     = lambda x: True if (x == 'WAHR' or x == 'TRUE') else False
ifIntEmpty = lambda x: True if (x == '' or x == 'TRUE') else False

def db_connect(host='localhost', schema='stammdaten', user=None, password=None, trace=False):
    if trace:
        print("Connecting to " + schema + "@" + host + "....", end="", flush=True)
    db_connection = mysql.connector.connect(
          host        = host,
          user        = user,
          password    = password,
          database    = schema,
          auth_plugin = 'mysql_native_password'
    )
    if trace:
        print("completed!")
    return db_connection


#  DATE_FORMAT(last_update, '%Y%m%d_%H%i%s')   AS Last_Update
def get_fields_from_view_or_table_as_list(db_connection, tbl_or_view_name='Personen_Daten', return_as_list=True, sep=",\n", end="\n"):
    sql_stm_get_desc = 'Desc ' + tbl_or_view_name
    mycursor = db_connection.cursor(dictionary=True)
    mycursor.execute(sql_stm_get_desc)
    myresult = mycursor.fetchall()

    if return_as_list:
        ret_value = []
        for aDataSet in myresult:
            ret_value.append(aDataSet['Field'])
    else:
        ret_value = ''
        for aDataSet in myresult:
            ret_value += aDataSet['Field'] + sep
        ret_value = ret_value[:-len(sep)]
        ret_value += end
    return ret_value

def get_data_from_view_or_table(db_connection,
                                tbl_or_view_name='Personen_Daten',
                                fields_to_select=None,
                                where_clause=None,
                                sort_criteria=None,
                                result_as='CSV',  # JSON CSV
                                sep=";",
                                end="\n",
                                trace=False):
    if fields_to_select is None:
        fields_to_select = get_fields_from_view_or_table_as_list(db_connection, tbl_or_view_name=tbl_or_view_name)

    field_as_string = ','.join(fields_to_select)
    header_string = sep.join(fields_to_select)
    sql_stm_select_from_tbl_or_view = "SELECT " + field_as_string + " FROM " + tbl_or_view_name
    if where_clause is not None:
        sql_stm_select_from_tbl_or_view += " WHERE " + where_clause
    if sort_criteria is not None:
        sql_stm_select_from_tbl_or_view += " ORDER BY " + sort_criteria

    sql_formatted = sqlparse.format(sql_stm_select_from_tbl_or_view, reindent=True, keyword_case='upper')

    if trace:
        print(sql_formatted)
    mycursor = stammdaten_schema.cursor(dictionary=True)
    mycursor.execute(sql_formatted)
    myresult = mycursor.fetchall()

    if result_as == 'CSV':
        ret_str = header_string + end
        for aDataSet in myresult:
            value_list = []
            for anAttr in fields_to_select:
                value_list.append(str(aDataSet[anAttr]))
            ret_str += sep.join(value_list)
            ret_str += end
        return ret_str
    else:
        return myresult

def csv_to_json(csvFilePath, delimiter=',', jsonFilePath=None):
    jsonArray = []

    # read csv file
    with open(csvFilePath, encoding='utf-8') as csvf:
        # load csv file data using csv library's dictionary reader
        # csvReader = csv.DictReader(csvf, delimiter=delimiter, quoting=csv.QUOTE_NONE)
        csvReader = csv.DictReader(csvf)


        # convert each csv row into python dict
        for row in csvReader:
            # add this python dict to json array
            jsonArray.append(row)

    # convert python jsonArray to JSON String and write to file
    if jsonFilePath is not None:
        with open(jsonFilePath, 'w', encoding='utf-8') as jsonf:
            jsonString = json.dumps(jsonArray, indent=4)
            jsonf.write(jsonString)

    return jsonArray

def read_from_file_and_call_stored_procedure(
        db_connection,
        csv_file_name=r'N:\02_EDV\Orte_Import.csv',
        csv_delimiter=';',   # Not functioning yet (file has to be separated by ,
        proc_name='getOrtId',
        db_file_mapping={'PLZ': 'PLZ', 'Name': 'Ortsname', 'Kanton': 'Kanton', 'TelVorwahl': 'Tel_Vorwahl'}):

    mycursor = stammdaten_schema.cursor()
    input_csv_json = csv_to_json(csv_file_name, delimiter=',')
    # print(input_csv_json)

    # print('db_file_mapping:', db_file_mapping)
    requested_attributs = db_file_mapping.keys()
    arg_counts=len(requested_attributs)
    # print('requested_attributs:', requested_attributs)

    print(proc_name, ":", len(input_csv_json))
    for a_data_tuple in input_csv_json:
        # print('-->', a_data_tuple)
        values = []
        for a_attribute in requested_attributs:
            values.append(a_data_tuple[db_file_mapping[a_attribute]])
            # print('==>', a_attribute, a_data_tuple[db_file_mapping[a_attribute]])
        print(','.join(requested_attributs), ' --> ', ','.join(values))

        if arg_counts == 0:
            args = ('x')
        elif arg_counts == 1:
            args = (values[0], 'x')
        elif arg_counts == 2:
            args = (values[0], values[1], 'x')
        elif arg_counts == 3:
            args = (values[0], values[1], values[2], 'x')
        elif arg_counts == 4:
            args = (values[0], values[1], values[2], values[3], 'x')
        elif arg_counts == 5:
            args = (values[0], values[1], values[2], values[3], values[4], 'x')
        elif arg_counts == 6:
            args = (values[0], values[1], values[2], values[3], values[4], values[5], 'x')
        elif arg_counts == 7:
            args = (values[0], values[1], values[2], values[3], values[4], values[5], values[6], 'x')
        elif arg_counts == 8:
            args = (values[0], values[1], values[2], values[3], values[4], values[5], values[6], values[7], 'x')
        elif arg_counts == 9:
            args = (values[0], values[1], values[2], values[3], values[4], values[5], values[6], values[7], values[8], 'x')
        elif arg_counts == 10:
            if proc_name == 'getPersonenId':
                if values[5] == 'Ja':
                    Name_Angenommen = True
                else:
                    Name_Angenommen = False

                args = ('BuergerDB', values[0], values[1], values[2], values[3], values[4], Name_Angenommen, values[6], values[7], values[8], values[9], 'x')
        elif arg_counts == 11:
            args = (values[0], values[1], values[2], values[3], values[4], values[5], values[6], values[7], values[8], values[9], values[10], 'x')
        else:
            print("ERROR: arg_counts ", arg_counts, " NOT defined")
        print(proc_name, args, sep="", end="  -->  ")
        result_args = mycursor.callproc(proc_name, args)
        print(proc_name, result_args, sep="")
        if proc_name == 'getPersonenId':
            personen_id = result_args[11]


            # Update Adressen
            postfach = a_data_tuple['strPostfach']
            # print(str(personen_id) + ":Postfach:" + postfach)
            sql = "UPDATE ADRESSEN SET Postfachnummer = '" + postfach + "' WHERE id=(SELECT Privat_Adressen_id from Personen where id=" + str(personen_id) + ")"
            # print(sql)
            mycursor.execute(sql)
            db_connection.commit()

            # Update Telefonnummer
            prio = 0
            telnr_ids = []
            if a_data_tuple['strTelefon_Privat1'] != "":
                args = ('0041', '', a_data_tuple['strTelefon_Privat1'], 'Privat', 'Festnetz', prio, 'x')
                result_args = mycursor.callproc('getTelefonnummerId', args)
                telnr_ids.append(result_args[6])
                prio += 1

            if a_data_tuple['strTelefon_Privat2'] != "":
                args = ('0041', '', a_data_tuple['strTelefon_Privat2'], 'Privat', 'Festnetz', prio, 'x')
                result_args = mycursor.callproc('getTelefonnummerId', args)
                telnr_ids.append(result_args[6])
                prio += 1

            if a_data_tuple['strTelefon_Geschäft1'] != "":
                args = ('0041', '', a_data_tuple['strTelefon_Geschäft1'], 'Geschaeft', 'Festnetz', prio, 'x')
                result_args = mycursor.callproc('getTelefonnummerId', args)
                telnr_ids.append(result_args[6])
                prio += 1

            if a_data_tuple['strTelefon_Geschäft2'] != "":
                args = ('0041', '', a_data_tuple['strTelefon_Geschäft2'], 'Geschaeft', 'Festnetz', prio, 'x')
                result_args = mycursor.callproc('getTelefonnummerId', args)
                telnr_ids.append(result_args[6])
                prio += 1

            if a_data_tuple['strMobil_Telefon'] != "":
                args = ('0041', '', a_data_tuple['strMobil_Telefon'], 'Privat', 'Mobile', prio, 'x')
                result_args = mycursor.callproc('getTelefonnummerId', args)
                telnr_ids.append(result_args[6])
                prio += 1

            if a_data_tuple['strFaxnummer_Privat'] != "":
                args = ('0041', '', a_data_tuple['strFaxnummer_Privat'], 'Privat', 'FAX', prio, 'x')
                result_args = mycursor.callproc('getTelefonnummerId', args)
                telnr_ids.append(result_args[6])
                prio += 1

            if a_data_tuple['strFaxnummer_Geschäft'] != "":
                args = ('0041', '', a_data_tuple['strFaxnummer_Geschäft'], 'Geschaeft', 'FAX', prio, 'x')
                result_args = mycursor.callproc('getTelefonnummerId', args)
                telnr_ids.append(result_args[6])
                prio += 1

            print("personen_id:", personen_id, "TelIds:", telnr_ids)
            for aTelnNerId in telnr_ids:
                sql = "INSERT INTO Personen_has_telefonnummern (Personen_id, Telefonnummern_id) VALUES (" + str(personen_id) + "," + str(aTelnNerId) + ")"
                print(sql)
                mycursor.execute(sql)
                db_connection.commit()

            # Update eMail
            prio = 0
            email_ids = []
            if a_data_tuple['strEmail'] != "":
                args = (a_data_tuple['strEmail'], 'Sonstige', prio, 'x')
                # print("     args:", args)
                result_args = mycursor.callproc('getEmailAdrId', args)
                # print("     resu:", args)
                email_ids.append(result_args[3])
                prio += 1


            print("personen_id:", personen_id, "eMailIds:", email_ids, "len(eMailIds):", len(email_ids))
            for aEmailId in email_ids:
                sql = "INSERT INTO Personen_has_email_adressen (Personen_id, EMail_Adressen_id) VALUES (" + str(personen_id) + "," + str(aEmailId) + ")"
                # print(sql)
                mycursor.execute(sql)
                db_connection.commit()

            # IBAN
            if a_data_tuple['strKonto_Nr'] != "":
                iban_str = a_data_tuple['strKonto_Nr']
                print("personen_id:", personen_id, "iban_str:", iban_str)
                sql = """
                         INSERT INTO iban 
                             (Personen_id, 
                             Nummer,
                             Prio) 
                         VALUES (""" + str(personen_id) + """,
                                 '""" + iban_str + """',
                                 0)
                    """
                # print(sql)
                mycursor.execute(sql)
                db_connection.commit()

            # Legacy-Fields
            lintNutzenKey = a_data_tuple['lintNutzenKey']
            if lintNutzenKey == '':
                lintNutzenKey = '-1'

            match a_data_tuple['lintZivilstandkey']:
                case '-1':
                    zivilstand = 'Unbestimmt'
                case '1':
                    zivilstand = 'Ledig'
                case '2':
                    zivilstand = 'Verheiratet'
                case '3':
                    zivilstand = 'Geschieden'
                case '4':
                    zivilstand = 'Getrennt'
                case '5':
                    zivilstand = 'Verwitwet'
                case other:
                    zivilstand = 'Leer'
            Zivilstand = "Zivilstand = '" + str(zivilstand) + "',"

            # Boolean Values
            bolNutzenberechtigung = ifTrue(a_data_tuple['bolNutzenberechtigung'])
            bolBaulandgesuch = ifTrue(a_data_tuple['bolBaulandgesuch'])
            bolMutationen_Aktuell = ifTrue(a_data_tuple['bolMutationen_Aktuell'])
            bolMitarbeiter_Genossenrat = ifTrue(a_data_tuple['bolMitarbeiter_Genossenrat'])
            bolWeggezogen = ifTrue(a_data_tuple['bolWeggezogen'])
            bolAktiv = ifTrue(a_data_tuple['bolAktiv'])
            bolBürgerauflage = ifTrue(a_data_tuple['bolBürgerauflage'])

            kategorien = []
            bolLandwirt = ifTrue(a_data_tuple['bolLandwirt'])
            if bolLandwirt:
                kategorien.append('Landwirt')

            bolGenossenbürger = ifTrue(a_data_tuple['bolGenossenbürger'])
            if bolGenossenbürger:
                kategorien.append('Bürger')

            if len(kategorien) > 0:
                kategorien = ','.join(kategorien)
                kategorien = "Kategorien = '" + kategorien + "',"
            else:
                kategorien = ""

            # Changed Mapping
            bolChronik = ifTrue(a_data_tuple['bolChronik'])
            if bolChronik:
                Chronik_Bezogen_Am = "Chronik_Bezogen_Am = str_to_date('5.8.1960','%d.%m.%Y'),"
            else:
                Chronik_Bezogen_Am = ""

            memBaulandgesuch_Info = a_data_tuple['memBaulandgesuch_Info']
            if memBaulandgesuch_Info == '':
                Baulandgesuch_Details = ""
            else:
                Baulandgesuch_Details = "Baulandgesuch_Details = '" + str(memBaulandgesuch_Info) + "',"

            memBemerkungen = a_data_tuple['memBemerkungen']
            if memBemerkungen == '':
                Bemerkungen = ""
            else:
                Bemerkungen = "Bemerkungen = '" + str(memBemerkungen) + "',"

            # Dates
            if a_data_tuple['datGeburtstag'] != '':
                Geburtstag = "Geburtstag = str_to_date('" + a_data_tuple['datGeburtstag'] + "','%d.%m.%Y'),"
            else:
                Geburtstag = ""

            if a_data_tuple['datGestorben'] != '':
                Todestag = "Todestag = str_to_date('" + a_data_tuple['datGestorben'] + "','%d.%m.%Y'),"
            else:
                Todestag = ""

            if a_data_tuple['datBaulandgesuch_Termin'] != '':
                Baulandgesuch_Eingereicht_Am = "Baulandgesuch_Eingereicht_Am = str_to_date('" + a_data_tuple['datBaulandgesuch_Termin'] + "','%d.%m.%Y'),"
            else:
                Baulandgesuch_Eingereicht_Am = ""

            if a_data_tuple['datBauland_Kauf'] != '':
                Bauland_Gekauft_Am = "Bauland_Gekauft_Am = str_to_date('" + a_data_tuple['datBauland_Kauf'] + "','%d.%m.%Y'),"
            else:
                Bauland_Gekauft_Am = ""

            if a_data_tuple['datAngemeldet'] != '':
                Angemeldet_Am = "Angemeldet_Am = str_to_date('" + a_data_tuple['datAngemeldet'] + "','%d.%m.%Y'),"
            else:
                Angemeldet_Am = ""




            # print("personen_id:", personen_id, "bolNutzenberechtigung:", bolNutzenberechtigung)
            sql = """
                     UPDATE Personen SET
                                -- Date Values
                                """ + Geburtstag + """
                                """ + Todestag + """
                                """ + Baulandgesuch_Eingereicht_Am + """
                                """ + Bauland_Gekauft_Am + """
                                """ + Angemeldet_Am + """
                                
                                -- Changed Mapping
                                """ + Chronik_Bezogen_Am + """
                                """ + Bemerkungen + """
                                """ + Baulandgesuch_Details + """
                                """ + kategorien + """
                                """ + Zivilstand + """
                                
                     
                                -- Legacy Fields
                                lintNutzenKey            = """ + str(lintNutzenKey) + """,
                                
                                -- Boolean Values
                                bolNutzenberechtigung    = """ + str(bolNutzenberechtigung) + """,
                                bolBaulandgesuch         = """ + str(bolBaulandgesuch) + """,
                                bolMutationen_Aktuell    = """ + str(bolMutationen_Aktuell) + """,
                                bolMitarbeiter_Genossame = """ + str(bolMitarbeiter_Genossenrat) + """,
                                bolWeggezogen            = """ + str(bolWeggezogen) + """,
                                bolAktiv                 = """ + str(bolAktiv) + """,
                                bolLandwirt              = """ + str(bolLandwirt) + """,
                                bolGenossenbürger        = """ + str(bolGenossenbürger) + """,
                                bolBürgerauflage         = """ + str(bolBürgerauflage) + """
                                
                     WHERE id = """ + str(personen_id) + """
                """
            print(sql)
            mycursor.execute(sql)
            db_connection.commit()
    '''
    mycursor = db_connection.cursor()
    
    print(sql)
    mycursor.execute(sql)

    # "
    # val = ("3000", "Bern")
    # mycursor.execute(sql, val)
    
    '''

def load_data_from_BuergerDB():

    if True:
        read_from_file_and_call_stored_procedure(
            db_connection=stammdaten_schema,
            csv_file_name=r'N:\02_EDV\Land_Import.csv',
            csv_delimiter=',',
            proc_name='getLandId',
            db_file_mapping={'Name': 'Landname', 'Code': 'Code', 'Landesvorwahl': 'Landesvorwahl'})

    if True:
        read_from_file_and_call_stored_procedure(
            db_connection=stammdaten_schema,
            csv_file_name=r'N:\02_EDV\Orte_Import.csv',
            csv_delimiter=',',
            proc_name='getOrtId',
            db_file_mapping={'PLZ': 'PLZ', 'Name': 'Ortsname', 'Kanton': 'Kanton', 'Tel_Vorwahl': 'TelVorwahl'})


    if True:
        read_from_file_and_call_stored_procedure(
            db_connection=stammdaten_schema,
            csv_file_name=r'N:\02_EDV\Personen_Import.csv',
            csv_delimiter=',',
            proc_name='getPersonenId',
            db_file_mapping={'Sex': 'Sex', 'Firma': 'Firma', 'Vorname': 'Vorname', 'Ledig_Name': 'Ledig_Name', 'Partner_Name': 'Partner_Name', 'Partner_Name_Angenommen': 'Partner_Name_Angenommen', 'Strasse': 'Strasse', 'Hausnummer': 'Hausnummer', 'PLZ': 'PLZ', 'Ort': 'Ort'})


def searchPerson(vorname, ledigname, partnername=None, gebdatum_YYYYMMDD=None):
    mycursor = stammdaten_schema.cursor()
    sql = """
             SELECT
                ID
             FROM Personen_Daten
             WHERE Vorname    = '""" + str(vorname)   + """' AND
                   Ledig_Name = '""" + str(ledigname) + """'
        """
    # print(sql)
    mycursor.execute(sql)
    myresult = mycursor.fetchall()
    if len(myresult) == 0:
        # print("Not Found:", len(myresult), "\n", sql)
        return None
    elif len(myresult) > 1:
        # print("Multiple found:", len(myresult), "\n", sql)
        return -1
    elif len(myresult) == 1:
        return myresult[0][0]

def load_bewirtschafter_details_from_Landteil_EXCEL ():
    wb = load_workbook(r'V:\Landwirtschaft\Pachtland\Landwirtschaftliche_Gütertabelle Stand_2023_02_17.xlsx')
    print(wb.sheetnames)

    Bewirtschafter_Sicht = wb['Bewirtschafter_Sicht']
    column = 0
    columnHeaders = {}
    while Bewirtschafter_Sicht.cell(row=2, column=column+1).value is not None:
        headerStr = str(Bewirtschafter_Sicht.cell(row=2, column=column+1).value).replace('\n', '')
        columnHeaders.update({headerStr: column+1})
        column += 1
    print(columnHeaders)

    row = 3
    last_value = ''
    while Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['BewirtschafterNummer']).value is not None:
        aDataSet = []

        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['BewirtschafterVorname']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['Name']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['Name4']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['Adresse']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['PLZ']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['Ort']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['Geb.-Datum']).value)[0:10].replace('-', ''))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['Tel.-Nr.']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['Mail']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['BewirtschafterNummer']).value))

        if aDataSet[0] is not None and aDataSet[9].count("Ergebnis") != 1:
            new_value = aDataSet[0] + ';' + aDataSet[1] + ';' + aDataSet[6]
            if new_value != last_value:
                last_value = new_value
                # print(new_value)
                aDataSet.insert(0, str(searchPerson(aDataSet[0], aDataSet[1], aDataSet[2], aDataSet[6])))
                print(';'.join(aDataSet))
        row += 1


def reco_personen():
    wb = load_workbook(r'V:\Landwirtschaft\Pachtland\Infotabellen_Landwirte_2023_02_25.xlsx')
    # print(wb.sheetnames)

    name_liste = wb['Bewirtschafter']
    column = 0
    columnHeaders = {}
    while name_liste.cell(row=1, column=column+1).value is not None:
        headerStr = str(name_liste.cell(row=1, column=column+1).value).replace('\n', '')
        columnHeaders.update({headerStr: column+1})
        column += 1
    print(columnHeaders)

    row = 2
    last_value = ''
    while name_liste.cell(row=row, column=columnHeaders['Paechter_Id']).value is not None:
        select_criterias = [str(name_liste.cell(row=row, column=columnHeaders['Ledig_Name']).value),
                            str(name_liste.cell(row=row, column=columnHeaders['Vorname']).value),
                            str(name_liste.cell(row=row, column=columnHeaders['Adresse']).value),
                            str(name_liste.cell(row=row, column=columnHeaders['Partner_Name']).value)]
        person_id_from_file = name_liste.cell(row=row, column=columnHeaders['Paechter_Id']).value
        rs = find_person(select_criterias)

        # print(rs)
        if len(rs) == 1:
            if rs[0] == 0:
                print(f'{person_id_from_file:4d} not found in DB!', str(select_criterias))

                args = ('Loader_1',  'x')
                proc_name = 'getPersonenId'
                result_args = mycursor.callproc(proc_name, args)
                print(proc_name, result_args, sep="")
                if proc_name == 'getPersonenId':
                    personen_id = result_args[11]
            else:
                if person_id_from_file == rs[0]:
                    db_attr_list = ['`Betriebs_Nr`', '`Geburtstag`', '`Tel_Nr`', '`eMail`']
                    db_record = get_person_details_from_DB_by_ID(rs[0], db_attr_list)[0]
                    print(f'{rs[0]:4d}    o.k.', str(db_record))
                else:
                    db_attr_list = ['Vorname', 'Ledig_Name', 'Partner_Name', 'Private_Strassen_Adresse', 'Private_PLZ', 'Private_Ort', 'Betriebs_Nr', 'Geburtstag']
                    db_record = get_person_details_from_DB_by_ID(rs[0], db_attr_list)[0]
                    print(person_id_from_file, rs[0], 'Wrong Person found!!!!!!', str(select_criterias), str(db_record))
        else:
            # print(person_id_from_file, "  ;".join(rs), ' multiple found in DB!')
            print(person_id_from_file, str(rs), ' multiple found in DB!')
        row += 1



def load_landteile_from_Landteil_EXCEL():
    wb = load_workbook(r'V:\Landwirtschaft\Pachtland\Landwirtschaftliche_Gütertabelle Stand_2023_02_17.xlsx')
    print(wb.sheetnames)

    Bewirtschafter_Sicht = wb['Bewirtschafter_Sicht']
    column = 0
    columnHeaders = {}
    while Bewirtschafter_Sicht.cell(row=2, column=column+1).value is not None:
        headerStr = str(Bewirtschafter_Sicht.cell(row=2, column=column+1).value).replace('\n', '')
        columnHeaders.update({headerStr: column+1})
        column += 1
    print(columnHeaders)

    row = 3
    last_value = ''
    while Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['BewirtschafterNummer']).value is not None:
        aDataSet = []

        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['Gemeinde']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['ParzellenNummer']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['ParzellenFlurname']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['BG/GGja / nein']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['FlächeAren']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['Pacht vonGenossame/Are']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['Pachtzins/AreFr.']).value)[0:10].replace('-', ''))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['Tel.-Nr.']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['Mail']).value))
        aDataSet.append(str(Bewirtschafter_Sicht.cell(row=row, column=columnHeaders['PachtzinsFr.']).value))

        if aDataSet[0] is not None and aDataSet[9].count("Ergebnis") != 1:
            new_value = aDataSet[0] + ';' + aDataSet[1] + ';' + aDataSet[6]
            if new_value != last_value:
                last_value = new_value
                # print(new_value)
                print(';'.join(aDataSet))
        row += 1

def get_WHERE_for_person_search(criteria_list, count_of_criterias_use=None):
    where_clauses_AND = []
    if count_of_criterias_use is None:
        count_of_criterias_use = len(criteria_list)

    used_items = []
    # print('==> count_of_criterias_use:', count_of_criterias_use, 'criteria_list:', len(criteria_list))
    if count_of_criterias_use > len(criteria_list):
        count_of_criterias_use = len(criteria_list)
    for ci in range(count_of_criterias_use):
        used_items.append("Such_Begriff LIKE BINARY '%" + criteria_list[ci] + "%'")

    # print(used_items)
    where_clause = ' AND \n              '.join(used_items)
    return where_clause

def get_SELECT_for_person_search(criteria_list, count_of_criterias_use=None):
    where_clause = get_WHERE_for_person_search(criteria_list, count_of_criterias_use)

    sql = """
        SELECT
            ID AS ID
        FROM Personen_Daten 
        WHERE """ + where_clause + """
        ORDER BY Familien_Name, Vorname;
    """
    # print(sql)
    return sql

def find_person_fix(criteria_list, count_of_criterias_use=None, doDebug=False):
    if count_of_criterias_use is None:
        count_of_criterias_use = len(criteria_list)
    mycursor = stammdaten_schema.cursor()
    mycursor.execute(get_SELECT_for_person_search(criteria_list, count_of_criterias_use))
    return mycursor.fetchall()

def find_person(criteria_list, count_of_criterias_use=1, doDebug=False):
        # print('==> ', criteria_list)
        if count_of_criterias_use > len(criteria_list):
            rs = find_person_fix(criteria_list)
            return [x[0] for x in rs]
        else:
            myresult = find_person_fix(criteria_list, count_of_criterias_use)
            # print('myresult:', myresult)
            if len(myresult) == 0:
                # print("Not Found:", len(myresult), "\n", sql)
                return [0]
            elif len(myresult) > 1:
                # print("Multiple found:", len(myresult), "\n", sql)
                return find_person(criteria_list, count_of_criterias_use + 1)
            elif len(myresult) == 1:
                return [myresult[0][0]]

def get_person_details_from_DB_by_ID(id, attr_list=['*']):
    fieldStr = (',\n            ').join(attr_list)

    sql = """
        SELECT
            """ + fieldStr + """
        FROM Personen_Daten 
        WHERE ID = """ + str(id) + """;
    """
    ### print(sql)
    mycursor = stammdaten_schema.cursor()
    mycursor.execute(sql)
    return mycursor.fetchall()

def TEST_findPerson():
    search_criteria = ["Bruhin", "Eugen", "Pesenti"]
    print(find_person(search_criteria))

    search_criteria = ["Rothlin", "Walter"]
    print(find_person(search_criteria))

if __name__ == '__main__':
    stammdaten_schema = db_connect(host='localhost',
                                   schema='stammdaten',
                                   user="App_User_Stammdaten",
                                   password="1234ABCD12abcd",
                                   trace=True)

    load_data_from_BuergerDB()

    # TEST_findPerson()
    # reco_personen()

    # load_bewirtschafter_details_from_Landteil_EXCEL()
    # load_landteile_from_Landteil_EXCEL()


    '''
    print(get_data_from_view_or_table(stammdaten_schema))
    print('\n\n')
    print(get_data_from_view_or_table(stammdaten_schema,
                                      tbl_or_view_name='Personen_Daten',
                                      fields_to_select=['Firma', 'Vorname', 'Ledig_Name']))
    print('\n\n')
    print(get_data_from_view_or_table(stammdaten_schema,
                                      tbl_or_view_name='Personen_Daten',
                                      fields_to_select=['Firma', 'Vorname', 'Ledig_Name'],
                                      where_clause="Ledig_Name='Rothlin'",
                                      sort_criteria="Vorname",
                                      result_as='JSON',
                                      trace=True))

    put_data_in_table(stammdaten_schema)
    '''





