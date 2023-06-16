#!/usr/bin/python3

# ------------------------------------------------------------------
# Name  : Genossame_Manager.py
# Source: https://raw.githubusercontent.com/walter-rothlin/Source-Code/master/Python_WaltisExamples/Genossame_Manager/Stammdaten_Manager.py
#
# Description: Manager für Genossame-Daten
#
# Autor: Walter Rothlin
#
# History:
# 01-Jan-2023   Walter Rothlin      Initial Version
# ------------------------------------------------------------------
import mysql.connector
import sqlparse
import csv
import json
import pandas as pd
import openpyxl
from waltisLibrary import *

# Lambda function to check if a x is from WAHR / FALSCH.
ifTrue     = lambda x: True if (x == 'WAHR' or x == 'TRUE') else False
ifIntEmpty = lambda x: True if (x == '' or x == 'TRUE') else False

def db_connect(host='localhost', port=3306, schema='stammdaten', user=None, password=None, trace=False):
    if trace:
        print("Connecting to " + schema + "@" + host + "....", end="", flush=True)
    db_connection = mysql.connector.connect(
          host        = host,
          port        = port,
          user        = user,
          password    = password,
          database    = schema,
          auth_plugin = 'mysql_native_password'
    )
    if trace:
        print("completed!")
    return db_connection


def get_record_count(db=None, db_tbl_name=None, retValueWithTblName=True):
    sql_insert = f'SELECT count(*) FROM {db_tbl_name} '
    mycursor = db.cursor()
    mycursor.execute(sql_insert)
    myresult = mycursor.fetchall()
    if retValueWithTblName:
        return {'rows_in_db   (' + db_tbl_name + '):': myresult[0][0]}
    else:
        return myresult[0][0]


def import_data_from_EXCEL(filename, sheet_name, csv_db_col_mapping={}, db=None, db_tbl_name=None, verbal=False, ingnore_sql_error=False):
    sheet_data = pd.read_excel(filename, sheet_name=sheet_name)
    csv_column_names = csv_db_col_mapping.keys()
    # print('columns:', csv_column_names)
    df = pd.DataFrame(sheet_data, columns=csv_column_names)

    row_counter_file_loaded = 0
    row_counter_file_not_loaded = 0
    mycursor = db.cursor()
    for index, row in df.iterrows():
        row_counter_file_loaded += 1
        csv_values = []
        db_attributes_names = []

        for a_csv_column_name in csv_column_names:
            default_value = str(csv_db_col_mapping[a_csv_column_name].get('default_value'))
            attribute_type = str(csv_db_col_mapping[a_csv_column_name].get('db_attribute_type'))
            a_csv_column_value = str(row[a_csv_column_name]).replace("'", "\\'")

            # print('1)', a_csv_column_name, ':: ', attribute_type, ': ', a_csv_column_value, '  (', default_value, ')', sep='', end='')
            if (a_csv_column_value == 'nan' or a_csv_column_value == 'NaT' or a_csv_column_value == ''):
                if default_value != '':
                    a_csv_column_value = default_value
                else:
                    a_csv_column_value = ""
            elif a_csv_column_value == 'False':
                a_csv_column_value = "0"
            elif a_csv_column_value == 'True':
                a_csv_column_value = "1"
            else:
                pass

            # print('-->', attribute_type)
            if attribute_type == 'int':
                # print('RegEx:', a_csv_column_value)
                a_csv_column_value = convert_str_to_int(a_csv_column_value)
            # print('   Final:', a_csv_column_value)
            if a_csv_column_value is not None and a_csv_column_value != 'None':
                if attribute_type == 'int':
                    csv_values.append(str(a_csv_column_value))
                    # csv_values.append('{zahl:11d}'.format(zahl=a_csv_column_value))
                else:
                    csv_values.append("'" + a_csv_column_value + "'")
                db_attributes_names.append(str(csv_db_col_mapping[a_csv_column_name]['db_table_attribute_name']))

        csv_values_as_string = ', '.join(csv_values)
        db_attributes_names_as_string = ', '.join(db_attributes_names)


        # print('db_attributes_types:', db_attributes_types)

        if ingnore_sql_error:
            sql_insert = f'''
             INSERT IGNORE INTO {db_tbl_name} ({db_attributes_names_as_string}) VALUES 
                    ({csv_values_as_string})
            '''
        else:
            sql_insert = f'''
             INSERT INTO {db_tbl_name} ({db_attributes_names_as_string}) VALUES 
                    ({csv_values_as_string})
            '''

        try:
            mycursor.execute(sql_insert)
        except Exception:
            if verbal:
                print('ERROR:', sql_insert)
            row_counter_file_not_loaded += 1
    db.commit()

    prompt_file = 'rows_in_file (' + sheet_name + '):'
    rec_count_file = {prompt_file: row_counter_file_loaded}

    prompt_db = 'rows_in_db   (' + db_tbl_name + '):'
    rec_count_db = get_record_count(db, db_tbl_name)

    prompt_not_loaded = 'rows_not_loaded:'
    rec_not_loaded = {prompt_not_loaded: row_counter_file_not_loaded}

    if rec_count_file[prompt_file] != rec_count_db[prompt_db]:
        print('\nWARNING:')
        if verbal:
            print('1) rec_count_db  :', str(rec_count_db))
            print('2) rec_count_file:', str(rec_count_file))
            print('3) rec_not_loaded:', str(rec_not_loaded))


    return [rec_count_file, rec_count_db, rec_not_loaded]

def initial_load(inport_excel_fn, tables_to_load, db_connection):
    for a_table in tables_to_load:
        if a_table == 'Länder':
            resultat = import_data_from_EXCEL(inport_excel_fn,
                                   sheet_name='Länder',
                                   csv_db_col_mapping={'ID': {'db_table_attribute_name': 'ID'},
                                                       'Land': {'db_table_attribute_name': 'Name'},
                                                       'Code': {'db_table_attribute_name': 'Code'},
                                                       'Landesvorwahl': {'db_table_attribute_name': 'Landesvorwahl'},
                                                       'last_update': {'db_table_attribute_name': 'last_update'}
                                                       },
                                   db=db_connection,
                                   db_tbl_name='Land')
            print(resultat)

        if a_table == 'Orte':
            resultat = import_data_from_EXCEL(inport_excel_fn,
                                   sheet_name='Orte',
                                   csv_db_col_mapping={'ID': {'db_table_attribute_name': 'ID'},
                                                       'PLZ': {'db_table_attribute_name': 'PLZ'},
                                                       'Name': {'db_table_attribute_name': 'Name'},
                                                       'Kanton': {'db_table_attribute_name': 'Kanton'},
                                                       'Land_ID': {'db_table_attribute_name': 'Land_ID'},
                                                       'last_update': {'db_table_attribute_name': 'last_update'}
                                                       },
                                   db=db_connection,
                                   db_tbl_name='Orte')
            print(resultat)

        if a_table == 'Adressen':
            resultat = import_data_from_EXCEL(inport_excel_fn,
                                   sheet_name='Adressen',
                                   csv_db_col_mapping={'ID': {'db_table_attribute_name': 'ID'},
                                                       'Strasse': {'db_table_attribute_name': 'Strasse'},
                                                       'Hausnummer': {'db_table_attribute_name': 'Hausnummer'},
                                                       'Postfachnummer': {'db_table_attribute_name': 'Postfachnummer'},
                                                       'Adresszusatz': {'db_table_attribute_name': 'Adresszusatz'},
                                                       'Wohnung': {'db_table_attribute_name': 'Wohnung'},
                                                       'Kataster_Nr': {'db_table_attribute_name': 'Kataster_Nr'},
                                                       'x_CH1903': {'db_table_attribute_name': 'x_CH1903'},
                                                       'Y_CH1903': {'db_table_attribute_name': 'y_CH1903'},
                                                       'Politisch_Wangen': {'db_table_attribute_name': 'Politisch_Wangen'},
                                                       'Ort_ID': {'db_table_attribute_name': 'Orte_ID'},
                                                       'last_update': {'db_table_attribute_name': 'last_update'}
                                                       },
                                   db=db_connection,
                                   db_tbl_name='Adressen')
            print(resultat)

        if a_table == 'Personen':
            '''
            'Partner_ID': {'db_table_attribute_name': 'Partner_ID',
                           'default_value': 'None',
                           'db_attribute_type': 'int'},
            'Vater_ID': {'db_table_attribute_name': 'Vater_ID',
                           'default_value': 'None',
                           'db_attribute_type': 'int'},
            'Mutter_ID': {'db_table_attribute_name': 'Mutter_ID',
                           'default_value': 'None',
                           'db_attribute_type': 'int'},
            '''
            resultat = import_data_from_EXCEL(inport_excel_fn,
                                   sheet_name='Personen',
                                   csv_db_col_mapping={'ID': {'db_table_attribute_name': 'ID'},
                                                       'Source': {'db_table_attribute_name': 'Source'},
                                                       'History': {'db_table_attribute_name': 'History'},
                                                       'Bemerkungen': {'db_table_attribute_name': 'Bemerkungen'},
                                                       'Sex': {'db_table_attribute_name': 'Sex'},
                                                       'Firma': {'db_table_attribute_name': 'Firma'},
                                                       'Vorname': {'db_table_attribute_name': 'Vorname'},
                                                       'Vorname_2': {'db_table_attribute_name': 'Vorname_2'},
                                                       'Ledig_Name': {'db_table_attribute_name': 'Ledig_Name'},
                                                       'Partner_Name': {'db_table_attribute_name': 'Partner_Name'},
                                                       'Partner_Name_Angenommen': {'db_table_attribute_name': 'Partner_Name_Angenommen'},
                                                       'AHV_Nr': {'db_table_attribute_name': 'AHV_Nr'},
                                                       'Betriebs_Nr': {'db_table_attribute_name': 'Betriebs_Nr'},
                                                       'Zivilstand': {'db_table_attribute_name': 'Zivilstand'},
                                                       'Kategorien': {'db_table_attribute_name': 'Kategorien'},
                                                       'Geburtstag': {'db_table_attribute_name': 'Geburtstag'},
                                                       'Todestag': {'db_table_attribute_name': 'Todestag'},
                                                       'Nach_Wangen_Gezogen': {'db_table_attribute_name': 'Nach_Wangen_Gezogen'},
                                                       'Von_Wangen_Weggezogen': {'db_table_attribute_name': 'Von_Wangen_Weggezogen'},
                                                       'Baulandgesuch_Eingereicht_Am': {'db_table_attribute_name': 'Baulandgesuch_Eingereicht_Am'},
                                                       'Bauland_Gekauft_Am': {'db_table_attribute_name': 'Bauland_Gekauft_Am'},
                                                       'Baulandgesuch_Details': {'db_table_attribute_name': 'Baulandgesuch_Details'},
                                                       'Angemeldet_Am': {'db_table_attribute_name': 'Angemeldet_Am'},
                                                       'Bezahlt_Aufnahme_Gebühr': {'db_table_attribute_name': 'Bezahlt_Aufnahme_Gebühr'},
                                                       'Aufgenommen_Am': {'db_table_attribute_name': 'Aufgenommen_Am'},
                                                       'Sich_Für_Bürgertag_Angemeldet_Am': {'db_table_attribute_name': 'Sich_Für_Bürgertag_Angemeldet_Am'},
                                                       'Neubürgertag_gemacht_Am': {'db_table_attribute_name': 'Neubürgertag_gemacht_Am'},
                                                       'Ausbezahlter_Bürgertaglohn': {'db_table_attribute_name': 'Ausbezahlter_Bürgertaglohn'},
                                                       'Funktion_Uebernommen_Am': {'db_table_attribute_name': 'Funktion_Uebernommen_Am'},
                                                       'Funktion_Abgegeben_Am': {'db_table_attribute_name': 'Funktion_Abgegeben_Am'},
                                                       'Chronik_Bezogen_Am': {'db_table_attribute_name': 'Chronik_Bezogen_Am'},
                                                       'Newsletter_Abonniert_Am': {'db_table_attribute_name': 'Newsletter_Abonniert_Am'},
                                                       'Privat_Adressen_ID': {'db_table_attribute_name': 'Privat_Adressen_ID'},
                                                       'Geschaefts_Adressen_ID': {'db_table_attribute_name': 'Geschaefts_Adressen_ID'},
                                                       'last_update': {'db_table_attribute_name': 'last_update'}
                                                       },
                                   db=db_connection,
                                   db_tbl_name='Personen')
            print(resultat)

        if a_table == 'IBAN':
            resultat = import_data_from_EXCEL(inport_excel_fn,
                                   sheet_name='IBAN_Liste',
                                   csv_db_col_mapping={'IBAN_ID': {'db_table_attribute_name': 'ID'},
                                                       'IBAN_Nummer': {'db_table_attribute_name': 'Nummer'},
                                                       'Bezeichnung': {'db_table_attribute_name': 'Bezeichnung'},
                                                       'Bankname': {'db_table_attribute_name': 'Bankname'},
                                                       'Bankort': {'db_table_attribute_name': 'Bankort'},
                                                       'Pers_ID': {'db_table_attribute_name': 'Personen_ID'},
                                                       'Prio': {'db_table_attribute_name': 'Prio'},
                                                       'last_update': {'db_table_attribute_name': 'last_update'}
                                                       },
                                   db=db_connection,
                                   db_tbl_name='IBAN')
            print(resultat)

        if a_table == 'EMail':
            resultat = import_data_from_EXCEL(inport_excel_fn,
                                   verbal=False,
                                   sheet_name='EMail_Liste',
                                   csv_db_col_mapping={'Email_ID': {'db_table_attribute_name': 'ID'},
                                                       'eMail_adresse': {'db_table_attribute_name': 'eMail'},
                                                       'Type': {'db_table_attribute_name': 'Type'},
                                                       'Prio': {'db_table_attribute_name': 'Prio'},
                                                       'last_update': {'db_table_attribute_name': 'last_update'}
                                                       },
                                   db=db_connection,
                                   db_tbl_name='email_adressen')
            print(resultat)

        if a_table == 'Person_Has_EMail':
            resultat = import_data_from_EXCEL(inport_excel_fn,
                                   sheet_name='EMail_Liste',
                                   csv_db_col_mapping={'Pers_ID': {'db_table_attribute_name': 'Personen_ID'},
                                                       'Email_ID': {'db_table_attribute_name': 'EMail_adressen_ID'},
                                                       'last_update': {'db_table_attribute_name': 'last_update'}
                                                       },
                                   db=db_connection,
                                   db_tbl_name='personen_has_email_adressen')
            print(resultat)

        if a_table == 'Telefon':
            resultat = import_data_from_EXCEL(inport_excel_fn,
                                   verbal=False,
                                   sheet_name='Telefon_Liste',
                                   csv_db_col_mapping={'Tel_ID': {'db_table_attribute_name': 'ID'},
                                                       'Laendercode': {'db_table_attribute_name': 'Laendercode'},
                                                       'Vorwahl': {'db_table_attribute_name': 'Vorwahl'},
                                                       'Nummer': {'db_table_attribute_name': 'Nummer'},
                                                       'Type': {'db_table_attribute_name': 'Type'},
                                                       'Endgeraet': {'db_table_attribute_name': 'Endgeraet'},
                                                       'Prio': {'db_table_attribute_name': 'Prio'},
                                                       'last_update': {'db_table_attribute_name': 'last_update'}
                                                       },
                                   db=db_connection,
                                   db_tbl_name='Telefonnummern')
            print(resultat)

        if a_table == 'Person_Has_Telefonnummer':
            resultat = import_data_from_EXCEL(inport_excel_fn,
                                   sheet_name='Telefon_Liste',
                                   csv_db_col_mapping={'Pers_ID': {'db_table_attribute_name': 'Personen_ID'},
                                                       'Tel_ID': {'db_table_attribute_name': 'Telefonnummern_ID'},
                                                       'last_update': {'db_table_attribute_name': 'last_update'}
                                                       },
                                   db=db_connection,
                                   db_tbl_name='personen_has_telefonnummern')
            print(resultat)

def inital_load_fromExcel(input_fn, db_connection, doit=False):
    if doit:
        initial_load(input_fn, ['Länder', 'Orte', 'Adressen', 'Personen'], db_connection)
        initial_load(input_fn, ['IBAN'], db_connection)
        initial_load(input_fn, ['EMail', 'Person_Has_EMail'], db_connection)
        initial_load(input_fn, ['Telefon', 'Person_Has_Telefonnummer'], db_connection)
def inserts_from_excel(filename, sheet_name, attribut, db_connection, verbal=False):

    myCursor = db_connection.cursor()
    insert_count = 0
    sheet_data = pd.read_excel(filename, sheet_name=sheet_name)
    if attribut == 'EMAIL':
        print('Adding eMail adresses ..   ', end='')
        if verbal:
            print()
        if sheet_name == 'Unbereinigt_email_telnr_iban':
            df = pd.DataFrame(sheet_data, columns=['ID', 'eMail_ID', 'eMail'])
            for index, row in df.iterrows():
                # print()
                # print(index, row)
                pers_ID = str(row[0]).replace('.0', '')
                email_ID = str(row[1]).replace('.0', '')
                email = str(row[2])
                if email_ID == 'nan' and email != 'nan':
                    insert_count += 1
                    args = (pers_ID, email, 'Sonstige', 0, 'x')  # 'x' in the Tuple will be replaced by OUT-Value
                    if verbal:
                        print('    addEmailAdr', str(args), sep='')
                    result_args = myCursor.callproc('addEmailAdr', args)

    if attribut == 'TELNR':
        insert_count = 0
        print('Adding TelNrs ..   ', end='')
        if verbal:
            print()
        if sheet_name == 'Unbereinigt_email_telnr_iban':
            df = pd.DataFrame(sheet_data, columns=['ID', 'Tel_Nr_ID', 'Tel_Nr'])
            for index, row in df.iterrows():
                pers_ID = str(row[0]).replace('.0', '')
                telnr_ID = str(row[1]).replace('.0', '')
                telnr = str(row[2]).replace(' ', '')
                if telnr_ID == 'nan' and telnr != 'nan':
                    insert_count += 1
                    vorwahl = telnr[0:3]
                    telnr   = telnr[3:]
                    if vorwahl == '055':
                        endgeraet = 'Festnetz'
                    else:
                        endgeraet = 'Mobile'
                    args = (pers_ID, '0041', vorwahl, telnr, 'Sonstige', endgeraet, 0, 'x')  # 'x' in the Tuple will be replaced by OUT-Value
                    if verbal:
                        print('    addTelNr', str(args), sep='')
                    result_args = myCursor.callproc('addTelNr', args)

    if attribut == 'IBAN':
        insert_count = 0
        print('Adding IBAN ..   ', end='')
        if verbal:
            print()
        if sheet_name == 'Unbereinigt_email_telnr_iban':
            df = pd.DataFrame(sheet_data, columns=['ID', 'IBAN_ID', 'IBAN'])
            for index, row in df.iterrows():
                pers_ID = str(row[0]).replace('.0', '')
                iban_ID = str(row[1]).replace('.0', '')
                iban = str(row[2])    # .replace(' ', '')
                if iban_ID == 'nan' and iban != 'nan':
                    insert_count += 1
                    args = (pers_ID, iban, 'x')
                    if verbal:
                        print('    addIBAN', str(args), sep='')
                    result_args = myCursor.callproc('addIBAN', args)

    print('    .. ', insert_count, 'reord(s) processed!')
    return insert_count


def updates_from_excel(filename, sheet_name, db_connection, verbal=False):
    update_count = 0
    sheet_data = pd.read_excel(filename, sheet_name=sheet_name)
    print('Updateing eMail, Tel_Nr, IBAN..   ', end='')
    if verbal:
        print()
    df = pd.DataFrame(sheet_data, columns=['ID', 'eMail_ID', 'eMail', 'Tel_Nr_ID', 'Tel_Nr', 'IBAN_ID', 'IBAN'])
    rec_updated = 0
    for index, row in df.iterrows():
        # print()
        # print(index, row)
        pers_ID = str(row[0]).replace('.0', '')
        email_ID = str(row[1]).replace('.0', '')
        email = str(row[2])
        if email_ID != 'nan' and email != 'nan':
            update_count += update_if_neccessary(db_connection, 'email_adressen', email_ID, 'eMail', email, verbal=True)

        tel_nr_ID = str(row[3]).replace('.0', '')
        tel_nr = str(row[4]).replace(' ', '')
        if tel_nr_ID != 'nan' and tel_nr != 'nan':
            update_count += update_if_neccessary(db_connection, 'telefonnummern', tel_nr_ID, 'nummer', tel_nr, verbal=True)

        iban_ID = str(row[5]).replace('.0', '')
        iban = str(row[6]).replace(' ', '')
        if iban_ID != 'nan' and iban != 'nan':
            # print(pers_ID, IBAN_ID, IBAN)
            update_count += update_if_neccessary(db_connection, 'iban', iban_ID, 'nummer', iban, verbal=True)

    return update_count

def execute_important_sql_queries(db_connection, verbal=False):
    myCursor = db_connection.cursor()
    print('Calling important updates...', end='')
    args = ()
    result_args = myCursor.callproc('important_updates', args)
    print('done')

def initial_load_pachtland(filename, db_connection, verbal=False):
    print('initial_load_pachtland...reading', pachlandzuteilung_fn)
    myCursor = db_connection.cursor()

    # sql = """TRUNCATE `landteile`;"""
    sql = """DELETE FROM `landteile`;"""
    myCursor.execute(sql)
    db_connection.commit()

    info_tabellen_landwirte = openpyxl.load_workbook(filename, data_only=True)
    paechter_sheets = [x for x in info_tabellen_landwirte.sheetnames if "_" in x]
    # print('paechter_sheets:', paechter_sheets)

    for aPaechter_sheet_name in paechter_sheets:
        aPaechter_sheet = info_tabellen_landwirte[aPaechter_sheet_name]
        paechter_name = aPaechter_sheet["L3"].value
        Paechter_id = aPaechter_sheet["M3"].value
        print(f'Processing {Paechter_id:5d} {aPaechter_sheet_name:30s} {paechter_name:30s}', end='')
        row_index = 11
        landteil_count = 0
        while True:
            flurname = aPaechter_sheet["B"+str(row_index)].value
            if flurname is None:
                break
            elif flurname == 'Total Aren:':
                row_index += 1
                continue
            else:
                landteil_count += 1
                gis_id = aPaechter_sheet["A" + str(row_index)].value
                geno_id = aPaechter_sheet["C" + str(row_index)].value
                flaeche_in_aren = 0
                flaeche_geno = aPaechter_sheet["D" + str(row_index)].value
                if flaeche_geno is None:
                    flaeche_bürger = aPaechter_sheet["E" + str(row_index)].value
                    if flaeche_bürger is None:
                        flaeche_in_aren = '0'
                    else:
                        flaeche_in_aren = flaeche_bürger
                else:
                    flaeche_in_aren = flaeche_geno
                bemerkungen = aPaechter_sheet["F" + str(row_index)].value
                zins_pro_are = aPaechter_sheet["G" + str(row_index)].value
                if zins_pro_are is None:
                    zins_pro_are = 0
                fix_pacht_zins = aPaechter_sheet["H" + str(row_index)].value
                if zins_pro_are > 0:
                    fix_pacht_zins = 0


                verpächter_id = aPaechter_sheet["I" + str(row_index)].value
                if verpächter_id is None:
                    verpächter_name = (aPaechter_sheet["J" + str(row_index)].value).replace(' ', '')
                    verpächter_vorname = (aPaechter_sheet["K" + str(row_index)].value)
                    verpächter_id = get_personen_id(db_connection, such_kriterien=[verpächter_name, verpächter_vorname], verbal=False)

                # if geno_id == '313.100.1':
                #     print(gis_id, flurname, geno_id, flaeche_in_aren, bemerkungen, zins_pro_are, zins_total_genossame, verpächter_id)
                sql = """INSERT INTO landteile (AV_Parzellen_Nr, GENO_Parzellen_Nr, Flur_Bezeichnung, Flaeche_In_Aren, Pachtzins_Pro_Are, Fix_Pachtzins, Paechter_ID, Verpaechter_ID) 
                         VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"""
                val = (gis_id, geno_id, flurname, flaeche_in_aren, zins_pro_are, fix_pacht_zins, Paechter_id, verpächter_id)
                myCursor.execute(sql, val)
                db_connection.commit()

            row_index += 1
        print(f'   -> {landteil_count:5d}', )
        # break

def update_if_neccessary(db_connection, tbl_name, id, field_name, field_value, verbal=False):
    verbal = False
    myCursor = db_connection.cursor()
    records_changed = 0
    if verbal:
        print(tbl_name, id, field_name, field_value)

    if tbl_name == 'email_adressen' or tbl_name == 'iban':
        sql_select = f'SELECT {field_name} FROM {tbl_name} WHERE id={id}'
        myCursor.execute(sql_select)
        myresult = myCursor.fetchall()
        old_value = myresult[0][0].replace(' ', '')
        if field_value != old_value:
            records_changed = 1
            print(tbl_name, '::  old: ', old_value, '   new:', field_value, '  id:', id)
            sql_update = f"UPDATE {tbl_name} SET {field_name} = '{field_value}' WHERE id={id}"
            print(sql_update, end='\n\n')
            myCursor.execute(sql_update)
            db_connection.commit()
    if tbl_name == 'telefonnummern':
        sql_select = f'SELECT vorwahl,{field_name} FROM {tbl_name} WHERE id={id}'
        myCursor.execute(sql_select)
        myresult = myCursor.fetchall()
        old_nummer = myresult[0][1]
        old_nummervorwahl = myresult[0][0]
        if old_nummervorwahl + old_nummer != field_value:
            records_changed = 1
            print(tbl_name, '::  old: ', old_nummervorwahl + old_nummer, '   new:', field_value)
            sql_update = f"UPDATE {tbl_name} SET Vorwahl = '{field_value[0:3]}', Nummer = '{field_value[3:]}' WHERE id={id}"
            print(sql_update, end='\n\n')
            myCursor.execute(sql_update)
            db_connection.commit()
    return records_changed

def get_personen_ids(db_connection, such_kriterien, verbal=False, db_table='personen_daten', search_attr='Such_Begriff', select_attributes=['ID', 'Vorname_Initial', 'Familien_Name', 'Private_Strassen_Adresse', 'Private_PLZ_Ort']):
    verbal = False
    myCursor = db_connection.cursor()

    prep_such_kriterien = []
    for a_such_kriterium in such_kriterien:
        a_such_kriterium.replace(' - ', '-')
        split_liste = a_such_kriterium.split('-')
        for an_item in split_liste:
            prep_such_kriterien.append(an_item)

    if verbal:
        print('get_personen_id', str(prep_such_kriterien))

    where_clauses = []
    for a_such_kriterium in prep_such_kriterien:
        where_clauses.append(f"{search_attr} LIKE Binary '%{a_such_kriterium}%'")

    where_clause_str = ' AND\n                  '.join(where_clauses)
    if False:
        print(where_clause_str)

    select_person = f"""
            SELECT 
               {','.join(select_attributes)}
            FROM {db_table} 
            WHERE {where_clause_str};
    """
    if verbal:
        print(select_person)

    myCursor.execute(select_person)
    result_set = myCursor.fetchall()
    if verbal:
        print(result_set)

    if len(result_set) == 0 and len(such_kriterien) > 2:
        # input('recorsive search?')
        result_set = get_personen_ids(db_connection, such_kriterien[:-1], verbal=verbal, db_table=db_table, search_attr=search_attr, select_attributes=select_attributes)


    if len(result_set) > 1:
        print('Multiple found')
        # input('weiter?')

    return result_set

def get_personen_id(db_connection, such_kriterien, verbal=False):
    myresult = get_personen_ids(db_connection, such_kriterien, verbal=verbal)

    verpächter_id = None
    if verbal:
        print("Records found:", len(myresult), myresult, '\n')
    if myresult == 1:
        verpächter_id = myresult[0][0]

    return verpächter_id

if __name__ == '__main__':

    connect_to_prod = True
    if connect_to_prod:
        stammdaten_schema = db_connect(host='192.168.253.24',
                                       port=3311,
                                       schema='genossame_wangen',
                                       user="root",
                                       password="Gen_88-mysql",
                                       trace=True)
    else:
        stammdaten_schema = db_connect(host='localhost',
                                       schema='genossame_wangen',
                                       user="App_User_Stammdaten",
                                       password="1234ABCD12abcd",
                                       trace=True)



    # data_import_fn = r'C:\Users\Landwirtschaft\Desktop\Geno_Daten_From_PTA.xlsx'
    # data_update_fn = r'C:\Users\Landwirtschaft\Desktop\Genossame_Alt\Genossame_Wangen_Daten_IBAN_EMAIL_TELNR_2023_05_27.xlsx'
    # data_update_fn = r'V:\Genossame_Wangen_Daten_Kopie.xlsx'
    initial_data_fn = r'V:\Geno_Wangen_Daten.xlsx'
    reco_data_fn = r'V:\Geno_Wangen_Daten.xlsx'
    pachlandzuteilung_fn = r'V:\Landwirtschaft\Pachtland\Infotabellen_Landwirte_2023_06_07.xlsx'

    # Initial Load
    # ============
    do_intial_load = False
    if do_intial_load:
        do_initial_load_buerger = True
        if do_initial_load_buerger:
            inital_load_fromExcel(initial_data_fn, stammdaten_schema, doit=True)

        do_initial_load_pachtland = True
        if do_initial_load_pachtland:
            initial_load_pachtland(pachlandzuteilung_fn, stammdaten_schema, verbal=True)


    # Updates and inserts from Reco
    # =============================
    do_reco = False
    if do_reco:
        do_inserts_from_reco = True
        if do_inserts_from_reco:
            rc = 0
            rc += inserts_from_excel(reco_data_fn, 'Unbereinigt_email_telnr_iban', 'EMAIL', stammdaten_schema, verbal=True)
            rc += inserts_from_excel(reco_data_fn, 'Unbereinigt_email_telnr_iban', 'TELNR', stammdaten_schema, verbal=True)
            rc += inserts_from_excel(reco_data_fn, 'Unbereinigt_email_telnr_iban', 'IBAN', stammdaten_schema, verbal=True)

            if rc > 0:
                print('\n\n---> All insert in', reco_data_fn)
            else:
                print('\n\n===> No inserts found in', reco_data_fn)

        do_updates_from_reco = True
        if do_updates_from_reco:
            rc = 0
            rc += updates_from_excel(reco_data_fn, 'Unbereinigt_email_telnr_iban', stammdaten_schema, verbal=True)

            if rc > 0:
                print('\n\n---> All updates in', reco_data_fn)
            else:
                print('\n\n===> No updates found in', reco_data_fn)

    # Newsletter Daten migrieren
    do_reco_Newsletter_set_id = True
    if do_reco_Newsletter_set_id:
        print('\n\n\n======> Newsletter Daten migrieren')
        excel_file = r"V:\EDV\Newsletter\MailadressenGenossenbürger.xlsx"
        excel_sheet = 'Mailadressen aktuell'
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        worksheet_sheet = workbook[excel_sheet]

        row = 2
        while worksheet_sheet["B"+str(row)].value is not None:
            if (worksheet_sheet["A"+str(row)].value is None) or (worksheet_sheet["A" + str(row)].value == '?') or (',' in str(worksheet_sheet["A" + str(row)].value)):
                such_kriterien = []
                such_kriterien.append(worksheet_sheet["B"+str(row)].value)
                such_kriterien.append(worksheet_sheet["D" + str(row)].value)
                if worksheet_sheet["C" + str(row)].value is not None:
                    such_kriterien.append(worksheet_sheet["C" + str(row)].value)
                elif worksheet_sheet["E" + str(row)].value is not None:
                    strasse = worksheet_sheet["E" + str(row)].value
                    strasse = strasse.replace('strasse', 'str.')
                    such_kriterien.append(strasse)
                # print(get_personen_id(stammdaten_schema, such_kriterien, verbal=False))
                people_found = get_personen_ids(stammdaten_schema, such_kriterien, verbal=False)

                ids = []
                for aRec in people_found:
                    ids.append(str(aRec[0]))
                ids_str = ','.join(ids)
                if len(ids) == 1:
                    worksheet_sheet["A" + str(row)] = str(ids[0])

                elif len(ids) == 0:
                    print('--> ', str(such_kriterien))
                    worksheet_sheet["A" + str(row)] = '?'

                else:
                    worksheet_sheet["A" + str(row)] = ids_str
                    print('==> ', str(such_kriterien))
                    worksheet_sheet["J" + str(row)] = str(people_found)
                    print('people_found (', len(people_found), '):', people_found,  end='\n\n\n')

            row += 1
        workbook.save(excel_file)
        workbook.close()

        # do compare and take over
        print('\n\n\n======> Compare and take over Data')
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        worksheet_sheet = workbook[excel_sheet]
        row = 2
        while worksheet_sheet["B"+str(row)].value is not None:
            pers_id = worksheet_sheet["A" + str(row)].value
            if (pers_id is not None) and (pers_id != '?') and (',' not in str(pers_id)):
                    geb_datum = worksheet_sheet["F" + str(row)].value
                    eMail_adr = worksheet_sheet["G" + str(row)].value
                    Natel_Nr = worksheet_sheet["H" + str(row)].value
                    bemerkungen = worksheet_sheet["I" + str(row)].value
                    print('+++++', pers_id, geb_datum, eMail_adr, Natel_Nr, bemerkungen)
                    if geb_datum is not None:

                    input('weiter_2')

            row += 1
        workbook.save(excel_file)
        workbook.close()




    # Cleanup Date
    # ============
    # execute_important_sql_queries(stammdaten_schema)

