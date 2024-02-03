#!/usr/bin/python3

# ------------------------------------------------------------------------------------------------
# Name  : Genossame_Common_Defs.py
# Source: https://raw.githubusercontent.com/walter-rothlin/Source-Code/master/Python_WaltisExamples/Genossame_Manager/Genossame_Common_Defs
#
# Description: Common Definitions for Genossame-Daten
#
# Autor: Walter Rothlin
#
# History:
# 21-Jun-2023   Walter Rothlin      Initial Version
# 03-Sep-2023   Walter Rothlin      Reorganized library
# 27-Nov-2023   Walter Rothlin      Mod get_person_details_from_DB_by_ID
# ------------------------------------------------------------------------------------------------
from waltisLibrary import *
import mysql.connector
import sqlparse
import csv
import json
import pandas as pd
import openpyxl
import hashlib

# Lambda function to check if a x is from WAHR / FALSCH.
ifTrue     = lambda x: True if (x == 'WAHR' or x == 'TRUE') else False
ifIntEmpty = lambda x: True if (x == ''     or x == 'TRUE') else False

session_attr_scriteria_addr_list = 'adress_s_criteria'
session_attr_scriteria_addr_orte_list = 'adress_orte_s_criteria'
session_attr_scriteria_orte_list = 'orte_s_criteria'

# =================
# Class Stammdaten
# =================
class Stammdaten:
    def __init__(self):
        self.__db_connection = db_connect(connect_to_prod=True, trace=True)

    def get_db_connection(self):
        return self.__db_connection

    def get_version(self):
        return("V1.0.0.1")

    def check_and_reconnect_db(self):
        sql = f"""SELECT count(*) from properties;"""
        # print(sql)
        try:
            mycursor = self.__db_connection.cursor(dictionary=True)
            mycursor.execute(sql)
            rs = mycursor.fetchall()[0]
            # print('rs:', rs, 'len(rs):', len(rs))
            connection_lost = False
        except:
            connection_lost = True
            print(f'{getTimestamp()}: check_and_reconnect_db()  --> Try to reconneced!!!')
            self.__db_connection = db_connect(connect_to_prod=True, trace=True)
        return not connection_lost

    def get_person_detail_from_DB_by_ID(self, id=None):
        if id is not None:
            sql = f"""SELECT * FROM Personen WHERE ID={id};"""
            # print(sql)
            mycursor = self.__db_connection.cursor(dictionary=True)
            mycursor.execute(sql)
            return mycursor.fetchall()[0]
        else:
            return {}

    def get_last_ID_for_table(self, table_name='Person', id_name='ID'):
        sql = f"""SELECT * from {table_name} WHERE ID=(SELECT max({id_name}) FROM {table_name});"""
        # print(sql)
        mycursor = self.__db_connection.cursor(dictionary=True)
        mycursor.execute(sql)
        rs = mycursor.fetchall()[0][id_name]
        # print('rs:', rs)
        return rs

    def get_ort_details_from_DB_by_ID(self, id=None, search_criterium=None, attr_list=['*'], case_sensitive=False):
        if id is None:
            sql = f"""
            SELECT
                *
            FROM Orte
            WHERE Name LIKE '%{search_criterium}%' OR Kanton = '{search_criterium}'
            ORDER BY ID;
            """
        else:
            sql = f"""
            SELECT
                *
            FROM Orte
            WHERE ID = {id};
            """
        # print(sql)
        mycursor = self.__db_connection.cursor(dictionary=True)
        mycursor.execute(sql)
        return mycursor.fetchall()

    def get_addr_ort_details_from_DB_by_ID(self, id=None, search_criterium=None, tabel_name='Adress_Daten', attr_list=['*'], case_sensitive=False):
        if id is None:
            sql = f"""
            SELECT
                *
            FROM {tabel_name}
            WHERE Strassen_Adresse_Ort LIKE '%{search_criterium}%'
            ORDER BY Strasse, Hausnummer;
            """
        else:
            sql = f"""
            SELECT
                *
            FROM {tabel_name}
            WHERE ID = {id};
            """
        # print(sql)
        mycursor = self.__db_connection.cursor(dictionary=True)
        mycursor.execute(sql)
        return mycursor.fetchall()

    def get_person_details_from_DB_by_ID(self, id=None, search_criterium=None, attr_list=['*'], case_sensitive=False):
        fieldStr = (',\n            ').join(attr_list)
        like_str = 'LIKE'
        if case_sensitive:
            like_str = 'LIKE Binary'

        if id is None:
            if search_criterium is None:
                sql = """
                    SELECT
                        """ + fieldStr + """
                    FROM Personen_Daten
                    Limit 0,20;
                """
            else:
                search_criterium = search_criterium.replace(' =', '=').replace('= ', '=')
                such_kriterien = search_criterium.split(' ')
                if False:
                    print('1:', such_kriterien)
                prep_such_kriterien = []
                add_where_clauses = []
                for a_such_kriterium in such_kriterien:
                    # print('1)', a_such_kriterium)
                    if '=' in a_such_kriterium:
                        add_where_clauses.append(a_such_kriterium)
                        continue
                    a_such_kriterium = a_such_kriterium.replace(' - ', '-')
                    a_such_kriterium = a_such_kriterium.replace('+', ' ')
                    # print('2)', a_such_kriterium)
                    split_liste = a_such_kriterium.split('-')
                    for an_item in split_liste:
                        prep_such_kriterien.append(an_item)

                where_clauses = []
                for a_such_kriterium in prep_such_kriterien:
                    where_clauses.append(f"Such_Begriff {like_str} '%{a_such_kriterium}%'")
                for a_such_kriterium in add_where_clauses:
                    where_clauses.append(f"{a_such_kriterium}")

                where_clause_str = ' AND\n                        '.join(where_clauses)
                # print(where_clause_str)

                sql = f"""
                    SELECT
                        {fieldStr}
                    FROM Personen_Daten
                    WHERE 
                        {where_clause_str};
                """
        else:
            sql = """
            SELECT
                """ + fieldStr + """
            FROM Personen_Daten 
            WHERE ID = """ + str(id) + """;
            """
        # print(sql)
        mycursor = self.__db_connection.cursor(dictionary=True)
        mycursor.execute(sql)
        return mycursor.fetchall()

    def get_Pers_Details_for_Pers_ID(self, id=None, table_name='IBAN', id_name='Personen_ID', attr_liste=['*']):
        ret_values = {}
        attr_liste_str = ', '.join(attr_liste)
        if id is not None:
            sql = f"""
                SELECT {attr_liste_str}
                FROM {table_name}
                WHERE {id_name} = {id};
            """
            # print(sql)
            mycursor = self.__db_connection.cursor(dictionary=True)
            mycursor.execute(sql)
            ret_values = mycursor.fetchall()
        return ret_values

    # def get_IBAN_Details_for_Pers_ID(self, pers_id=None):
    #    return

    def is_password_correct(self, username, password, password_is_hash=False):
        hashed_password = hash_password(password)

        # print("Original Password:", password)
        # print("Hashed Password:", hashed_password, len(hashed_password))

        sql = f"""
        SELECT Password FROM Personen_Daten WHERE ID IN (
            SELECT
                Pers_id
            FROM email_liste 
            WHERE eMail_Adresse = '{username}');
        """
        sql = f"""
        SELECT Personen_ID, Password 
        FROM login_table
        WHERE eMail = '{username}';
        """
        # print(sql)
        mycursor = self.__db_connection.cursor(dictionary=True)
        try:
            mycursor.execute(sql)
            rs = mycursor.fetchall()[0]
            print('rs:', rs)
            password_found = rs['Password']
            if password_found == password or password_found == hashed_password:
                return True, rs['Personen_ID']
            else:
                return False, rs['Personen_ID']
        except Exception:
            return False, None

    def get_priviliges_for_pers_ID(self, pers_id):
        sql = f"""
        SELECT Application, Privilige 
        FROM App_Priviliges
        WHERE Pers_ID = '{pers_id}';
        """
        # print(sql)
        mycursor = self.__db_connection.cursor(dictionary=True)
        try:
            mycursor.execute(sql)
            rs = mycursor.fetchall()
            ## print('rs:', rs)
            return rs
        except Exception:
            return None

    def insert_new_adresse(self, new_name_values, verbal=False):
        if verbal:
            print(f'''
            insert_new_adresse
            ------------------
            {new_name_values}
            ''')

        sql = f"""
        INSERT INTO `adressen` (`Strasse`, `Hausnummer`, `Orte_ID`, `Politisch_Wangen`) VALUES 
              ('{new_name_values["Strasse"]}','{new_name_values["Hausnummer"]}', '{new_name_values["Ort_ID"]}', '{new_name_values["Politisch_Wangen"]}');
        """
        # print(sql)
        mycursor = self.__db_connection.cursor(dictionary=True)
        try:
            mycursor.execute(sql)
            self.__db_connection.commit()
            rs = self.get_last_ID_for_table('Adressen')
            # print('rs:', rs)
            return rs
        except Exception:
            return None

    def insert_new_person(self, new_name_values, verbal=False):
        if verbal:
            print(f'''
            insert_new_person
            -----------------
            {new_name_values}
            ''')

        sql = f"""
        INSERT INTO `personen` (`Sex`, `Vorname`, `Ledig_Name`, `Partner_Name`, `Privat_Adressen_ID`) VALUES 
              ('{new_name_values["Sex"]}','{new_name_values["Vorname"]}', '{new_name_values["Ledig_Name"]}', '{new_name_values["Partner_Name"]}', 701);
        """
        # print(sql)
        mycursor = self.__db_connection.cursor(dictionary=True)
        try:
            mycursor.execute(sql)
            self.__db_connection.commit()
            rs = self.get_last_ID_for_table('Personen')
            # print('rs:', rs)
            return rs
        except Exception:
            return None

    def delete_iban_telnr_email(self, pid=None, change_type=None, id=None, verbal=False):
        if verbal:
            print(f'''
            delete_iban_telnr_email
            -------------------------
                pid:{pid}
                id:{id}
                change_type:{change_type}
            ''')
        myCursor = self.__db_connection.cursor(dictionary=False)
        if change_type == 'iban':

            args = (pid, id)
            if verbal:
                print(f"""...call proc .... deleteIBAN{args}""")
            result_args = myCursor.callproc('deleteIBAN', args)

        elif change_type == 'telnr':
            args = (pid, id)
            if verbal:
                print(f"""...call proc .... deleteTelnrFull{args}""")
            result_args = myCursor.callproc('deleteTelnrFull', args)

        elif change_type == 'email':
            args = (pid, id)
            if verbal:
                print(f"""...call proc .... deleteEmailAdrFull{args}""")
            result_args = myCursor.callproc('deleteEmailAdrFull', args)

        self.__db_connection.commit()
        return id

    def new_iban_telnr_email(self, pid=None, change_type=None, new_name_values={}, verbal=False):
        if verbal:
            print(f'''
            new_iban_telnr_email
            -------------------------
                pid:{pid}
                values:{new_name_values}
            ''')

        myCursor = self.__db_connection.cursor(dictionary=False)
        if change_type == 'iban':
            # args = (pid, new_name_values['Nummer'], new_name_values['Bezeichnung'], new_name_values['Bankname'], new_name_values['Bankort'], new_name_values['Prio'], id, new_name_values['Lautend_auf'], 'x')
            args = (pid, 'CH12 1234 1234 1234 1234 1', 'Privatkonto', 'SZKB', 'Wangen', 9, 'Lautend auf', 'x')
            if verbal:
                print(f"""...call proc .... addIBAN{args}""")
            result_args = myCursor.callproc('addIBAN', args)

        elif change_type == 'telnr':
            args = (pid, '0041', '055', '1234567', 'Privat', 'Mobile', 9, 'x')
            if verbal:
                print(f"""...call proc .... addTelNr{args}""")
            result_args = myCursor.callproc('addTelNr', args)

        elif change_type == 'email':
            args = (pid, 'abc.def@ertz.ch', 'Privat', 9, 'x')
            if verbal:
                print(f"""...call proc .... addEmailAdr{args}""")
            result_args = myCursor.callproc('addEmailAdr', args)

        self.__db_connection.commit()
        if verbal:
            print(f'callproc:new_iban_telnr_email{args}')
            print(f'        :return value{result_args}')
        return result_args[-1]

    def update_adress_ort(self, id=None, new_name_values={}, verbal=False):
        if verbal:
            print(f'''
            update_adress_ort
            -------------------------
                id:{id}
                values:{new_name_values}
            ''')

        sql_update = f"""
        UPDATE adressen SET 
              Strasse           = '{new_name_values['Strasse']}', 
              Hausnummer        = '{new_name_values['Hausnummer']}', 
              Postfachnummer    = '{new_name_values['Postfachnummer']}', 
              Adresszusatz      = '{new_name_values['Adresszusatz']}', 
              Wohnung           = '{new_name_values['Wohnung']}', 
              Kataster_Nr       = '{new_name_values['Kataster_Nr']}', 
              x_CH1903          = '{new_name_values['x_CH1903']}', 
              y_CH1903          = '{new_name_values['y_CH1903']}', 
              Politisch_Wangen  =  {new_name_values['Politisch_Wangen']}, 
              orte_id           =  {new_name_values['Ort_ID']}
        WHERE id={id};
        """
        # print(sql_update)

        myCursor = self.__db_connection.cursor(dictionary=True)
        myCursor.execute(sql_update)
        self.__db_connection.commit()


    def update_iban_telnr_email(self, pid=None, id=None, new_name_values={}, verbal=False):
        if verbal:
            print(f'''
            update_iban_telnr_email
            -------------------------
                pid:{pid}
                id:{id}
                values:{new_name_values}
            ''')

        myCursor = self.__db_connection.cursor(dictionary=True)
        if new_name_values['Change_Type'] == 'iban':
            args = (pid, new_name_values['Nummer'], new_name_values['Bezeichnung'], new_name_values['Bankname'], new_name_values['Bankort'], new_name_values['Prio'], id, new_name_values['Lautend_auf'])
            if verbal:
                print(f"""...call proc .... updateIBAN{args}""")
            result_args = myCursor.callproc('updateIBAN', args)

        elif new_name_values['Change_Type'] == 'telnr':
            args = (id, new_name_values['Laendercode'], new_name_values['Vorwahl'], new_name_values['Nummer'], new_name_values['Type'], new_name_values['Endgeraet'], new_name_values['Prio'])
            if verbal:
                print(f"""...call proc .... updateTelnr{args}""")
            result_args = myCursor.callproc('updateTelnr', args)

        elif new_name_values['Change_Type'] == 'email':
            args = (id, new_name_values['eMail'], new_name_values['Type'], new_name_values['Prio'])
            if verbal:
                print(f"""...call proc .... updateEmailAdr{args}""")
            result_args = myCursor.callproc('updateEmailAdr', args)

        self.__db_connection.commit()


    def update_pers_details_by_ID(self, new_name_values, verbal=False):
        if verbal:
            print(f'''
            update_pers_details_by_ID
            -------------------------
            ''')
        pers_id = new_name_values['ID']
        table_name = 'Personen'
        attr_types = self.get_Attribute_Types(table_name=table_name)
        old_name_values = self.get_person_detail_from_DB_by_ID(pers_id)
        if verbal:
            print('ATTR_TYPES:\n', attr_types)
            print('OLD_VALUES:\n', old_name_values)
            print('NEW_VALUES:\n', new_name_values)

        # geänderte Werte
        changed_values = {}
        for a_key in new_name_values:
            # old_value = old_name_values[a_key]
            # new_value = new_name_values[a_key]
            # print(f'{a_key}: {old_value}  --> {new_value}')
            if str(new_name_values[a_key]) != str(old_name_values[a_key]):
                # print(f'{a_key} ({attr_types[a_key]}): {old_name_values[a_key]} --> {new_name_values[a_key]}')
                if attr_types[a_key] == 'set':
                    if verbal:
                        print(f'Set: {a_key} ({attr_types[a_key]}): {old_name_values[a_key]} --> {new_name_values[a_key]}')
                    changed_values[a_key + ' (set)'] = new_name_values[a_key]
                    changed_values[a_key + ' (set)'] = new_name_values[a_key]
                elif attr_types[a_key] == 'int':
                    if verbal:
                        print(f'Int: {a_key} ({attr_types[a_key]}): {old_name_values[a_key]} --> {new_name_values[a_key]}')
                    changed_values[a_key + ' (int)'] = new_name_values[a_key]
                elif attr_types[a_key] == 'enum':
                    if verbal:
                        print(f'Enum: {a_key} ({attr_types[a_key]}): {old_name_values[a_key]} --> {new_name_values[a_key]}')
                    changed_values[a_key + ' (enum)'] = new_name_values[a_key]
                elif attr_types[a_key] == 'date':
                    if old_name_values[a_key] is not None:
                        o_year, o_month, o_day = str(old_name_values[a_key]).split('-')
                        if new_name_values[a_key] is not None and new_name_values[a_key] != '':
                            n_day, n_month, n_year = str(new_name_values[a_key]).split('.')
                            if o_year != n_year or o_month != n_month or o_day != n_day:
                                if verbal:
                                    print(f'Date: {a_key} ({attr_types[a_key]}): {old_name_values[a_key]} --> {new_name_values[a_key]}')
                                changed_values[a_key + ' (date)'] = new_name_values[a_key]
                        else:
                            changed_values[a_key + ' (date)'] = new_name_values[a_key]
                    else:
                        changed_values[a_key + ' (date)'] = new_name_values[a_key]
                else:
                    if verbal:
                        print(f'ELSE: {a_key} ({attr_types[a_key]}): {old_name_values[a_key]} --> {new_name_values[a_key]}')
                    changed_values[a_key] = new_name_values[a_key]
        # print(f'VALUES_To_Update for ID={pers_id}:\n{changed_values}\n\n')

        count_of_attr_changed = len(changed_values)
        if count_of_attr_changed > 0:
            sql_update = f"""
            UPDATE
                `{table_name}`
            SET
            """

            set_fields = []
            for a_key in changed_values:
                if '(set)' in a_key:
                    a_key_clean = a_key.replace('(set)', '').strip()
                    new_kategorien_value = changed_values[a_key].replace('{', '').replace('}', '').replace("', '", ',')
                    set_fields.append(f"`{a_key_clean}` = {new_kategorien_value}")
                elif '(enum)' in a_key:
                    a_key_clean = a_key.replace('(enum)', '').strip()
                    new_kategorien_value = changed_values[a_key].replace('{', '').replace('}', '').replace("', '", ',')
                    set_fields.append(f"`{a_key_clean}` = '{new_kategorien_value}'")
                elif '(int)' in a_key:
                    a_key_clean = a_key.replace('(int)', '').strip()
                    if changed_values[a_key] == '':
                        set_fields.append(f"`{a_key_clean}` = NULL")
                    else:
                        set_fields.append(f"`{a_key_clean}` = {changed_values[a_key]}")
                elif '(date)' in a_key:
                    a_key_clean = a_key.replace('(date)', '').strip()
                    if changed_values[a_key] == '':
                        set_fields.append(f"`{a_key_clean}` = NULL")
                    else:
                        set_fields.append(f"`{a_key_clean}` = STR_TO_DATE('{changed_values[a_key]}', '%d.%m.%Y')")
                else:
                    set_fields.append(f"`{a_key}` = '{changed_values[a_key]}'")

            sql_update += ',\n'.join(set_fields)
            sql_update += f"""
            WHERE (`ID` = {pers_id});
            """
            # print(sql_update)
            myCursor = self.__db_connection.cursor(dictionary=True)
            myCursor.execute(sql_update)
            self.__db_connection.commit()
            result_set = myCursor.fetchall()
            # print('result_set:', result_set)
        else:
            print('Nichts geändert')
        return {'Record_changed': 1, 'Attribute_changed': count_of_attr_changed}

    def get_Attribute_Types(self, table_name=None):
        ret_dict = {}
        if table_name is not None:
            sql = f"""
            SELECT
                Attribute,
                Attr_Type
            FROM table_meta_data 
            WHERE `Schema` = 'genossame_wangen' AND 
                  `Table`  = '{table_name}';
            """
            # print(sql)
            mycursor = self.__db_connection.cursor(dictionary=False)
            mycursor.execute(sql)
            for a_tuple in mycursor.fetchall():
                ret_dict[a_tuple[0]] = str(a_tuple[1]).replace("b'", "").replace("'", "")
            return ret_dict
        else:
            return {}



# =================
# Geno DB-Functions
# =================

def db_connect(connect_to_prod=True, trace=False):
    if connect_to_prod:
        stammdaten_schema = mysql_db_connect(db_host='192.168.253.24',
                                             port=3311,
                                             db_schema='genossame_wangen',
                                             db_user_name="Web_App_User",
                                             password="Geno_8855!",
                                             trace=trace)
    else:
        stammdaten_schema = mysql_db_connect(db_host='localhost',
                                             db_schema='genossame_wangen',
                                             db_user_name="App_User_Stammdaten",
                                             password="1234ABCD12abcd",
                                             trace=trace)
    return stammdaten_schema

def get_person_details_for_id(db_connection, pers_id, verbal=False, db_table='personen_daten', select_attributes=['ID', 'Vorname_Initial', 'Familien_Name', 'Private_Strassen_Adresse', 'Private_PLZ_Ort']):
    select_person = f"""
            SELECT 
               {','.join(select_attributes)}
            FROM {db_table} 
            WHERE ID = {pers_id};
    """
    myCursor = db_connection.cursor()
    if verbal:
        print(select_person)

    myCursor.execute(select_person)
    result_set = myCursor.fetchall()
    if verbal:
        print(result_set)
    return result_set


def get_personen_ids(db_connection, such_kriterien, verbal=False, db_table='personen_daten', search_attr='Such_Begriff', select_attributes=['ID', 'Vorname_Initial', 'Familien_Name', 'Private_Strassen_Adresse', 'Private_PLZ_Ort']):
    myCursor = db_connection.cursor()

    prep_such_kriterien = []
    for a_such_kriterium in such_kriterien:
        a_such_kriterium.replace(' - ', '-')
        split_liste = a_such_kriterium.split('-')
        for an_item in split_liste:
            prep_such_kriterien.append(an_item)

    if verbal:
        print('get_personen_id', str(prep_such_kriterien))

    where_clauses = []
    for a_such_kriterium in prep_such_kriterien:
        where_clauses.append(f"{search_attr} LIKE Binary '%{a_such_kriterium}%'")

    where_clause_str = ' AND\n                  '.join(where_clauses)
    if False:
        print(where_clause_str)

    select_person = f"""
            SELECT 
               {','.join(select_attributes)}
            FROM {db_table} 
            WHERE {where_clause_str};
    """
    if verbal:
        print(select_person)

    myCursor.execute(select_person)
    result_set = myCursor.fetchall()
    if verbal:
        print(result_set)

    if len(result_set) == 0 and len(such_kriterien) > 2:
        # input('recorsive search?')
        result_set = get_personen_ids(db_connection, such_kriterien[:-1], verbal=verbal, db_table=db_table, search_attr=search_attr, select_attributes=select_attributes)


    if len(result_set) > 1:
        print('Multiple found')
        # input('weiter?')
    return result_set


def get_personen_id(db_connection, such_kriterien, verbal=False):
    myresult = get_personen_ids(db_connection, such_kriterien, verbal=verbal)

    pers_id = None
    if verbal:
        print("Records found:", len(myresult), myresult, '\n')
    if myresult == 1:
        pers_id = myresult[0][0]
    return pers_id


def execute_important_sql_queries(db, verbal=True):
    myCursor = db.cursor(dictionary=True)
    if verbal:
        print('--> Calling stored-proc important updates...', end='')
    args = ()
    result_args = myCursor.callproc('important_updates', args)
    if verbal:
        print('done\n')

    # If Nach_Wangen_Gezogen und Von_Wangen_Weggezogen => Aelteres in Bemerkungen schreiben und NULL setzten
    # -------------------------------------------------------------------------------------------------------
    if verbal:
        print('''--> Checking if somebody has set ...
                 Nach_Wangen_Gezogen AND Von_Wangen_Weggezogen
                     Move older date to Bemerkungen (append)
                     Set older date to NULL
              ''', end='')
    select_person = f"""
        SELECT ID, 
               Bemerkungen, 
               DATE_FORMAT(Nach_Wangen_Gezogen,'%Y_%m_%d:Rückkehr nach Wangen') AS Rückkehr_am, 
               DATE_FORMAT(Von_Wangen_Weggezogen,'%Y_%m_%d:Wegzug von Wangen')  AS Wegzug_am
        FROM personen 
        WHERE Nach_Wangen_Gezogen IS NOT NULL AND
              Von_Wangen_Weggezogen IS NOT NULL;
    """
    if False:
        print(select_person)

    myCursor.execute(select_person)
    result_set = myCursor.fetchall()
    if len(result_set) > 0:
        print(result_set)
    for a_result in result_set:
        if a_result['Rückkehr_am'] < a_result['Wegzug_am']:
            print(f"{a_result['ID']} ist Wegzüger")
            res = update_db_attribute(db=db,
                                db_tbl_name='Personen', db_attr_name='Bemerkungen', db_attr_type='varchar', db_attr_set_enum_values='',
                                id_attr_name='ID', id=a_result['ID'],
                                new_value=f"|{a_result['Rückkehr_am']}", new_value_format=None,
                                take_action=True, verbal=False)
            res = update_db_attribute(db=db,
                                      db_tbl_name='Personen', db_attr_name='Nach_Wangen_Gezogen',
                                      id_attr_name='ID', id=a_result['ID'],
                                      new_value='NULL', new_value_format=None,
                                      take_action=True, verbal=False)
        else:
            print(f"{a_result['ID']} ist Rückkehrer")
            res = update_db_attribute(db=db,
                                db_tbl_name='Personen', db_attr_name='Bemerkungen', db_attr_type='varchar', db_attr_set_enum_values='',
                                id_attr_name='ID', id=a_result['ID'],
                                new_value=f"|{a_result['Wegzug_am']}", new_value_format=None,
                                take_action=True, verbal=False)
            res = update_db_attribute(db=db,
                                      db_tbl_name='Personen', db_attr_name='Von_Wangen_Weggezogen',
                                      id_attr_name='ID', id=a_result['ID'],
                                      new_value='NULL', new_value_format=None,
                                      take_action=True, verbal=False)
    if verbal:
        print('done!!\n')

    # If Nach_Wangen_Gezogen before 1.7. dieses Jahr => +Verwaltungsberechtigt,Nutzungsberechtigt
    # ------------------------------------------------------------------------------------------------------
    if verbal:
        print('''--> Checking if somebody has moved to Wangen ...
              ''', end='')
    select_person = f"""
        SELECT ID,
			   Vorname,
			   Ledig_Name,
               Kategorien,
               STR_TO_DATE(CONCAT(DATE_FORMAT(now(),'%Y'),'-07-01'),'%Y-%m-%d') AS Reference_Date,
               IF (Nach_Wangen_Gezogen <  STR_TO_DATE(CONCAT(DATE_FORMAT(now(),'%Y'),'-07-01'),'%Y-%m-%d'), "True", "False") AS is_Nutzungsberechtigt,
			   Nach_Wangen_Gezogen
        FROM personen 
        WHERE Nach_Wangen_Gezogen IS NOT NULL AND
              FIND_IN_SET('Nutzungsberechtigt', Kategorien) =  0;
    """
    if False:
        print(select_person)

    myCursor.execute(select_person)
    result_set = myCursor.fetchall()
    if len(result_set) > 0:
        print(result_set)

    for a_person in result_set:
        update_person = f"""
             UPDATE `personen` SET `Kategorien` = addSetValue(`Kategorien`,'Verwaltungsberechtigt')
             WHERE `ID` = {a_person['ID']};   
         """
        if False:
            print(update_person)
        myCursor.execute(update_person)
        db.commit()
        result_set = myCursor.fetchall()
        if len(result_set) > 0:
            print(result_set)

        if a_person['is_Nutzungsberechtigt'] == 'True':
            update_person = f"""
                 UPDATE `personen` SET `Kategorien` = addSetValue(`Kategorien`,'Nutzungsberechtigt')
                 WHERE `ID` = {a_person['ID']};   
             """
            if False:
                print(update_person)
            myCursor.execute(update_person)
            db.commit()
            result_set = myCursor.fetchall()
            if len(result_set) > 0:
                print(result_set)

    if verbal:
        print('''
                Kategorien = +Verwaltungsberechtigt
                if the move was before 1.7. then
                Kategorien = Nutzungsberechtigt
             done!!
             ''')

    # If Von_Wangen_Weggezogen => -Verwaltungsberechtigt und im darauf folgendem Jahr -Nutzungsberechtigt
    # ---------------------------------------------------------------------------------------------------
    if verbal:
        print('''--> Checking if somebody has moved away from Wangen ...
              ''', end='')
    select_person = f"""
        SELECT ID,
			   Vorname,
			   Ledig_Name,
               Kategorien,
               DATE_FORMAT(Von_Wangen_Weggezogen,'%Y') AS Wegzug_Jahr,
               DATE_FORMAT(now(),'%Y')                 AS Reference_Year,
               IF (DATE_FORMAT(Von_Wangen_Weggezogen,'%Y') < DATE_FORMAT(now(),'%Y'), "False", "True") AS is_Nutzungsberechtigt,
			   Von_Wangen_Weggezogen
        FROM personen 
        WHERE Von_Wangen_Weggezogen IS NOT NULL AND
              FIND_IN_SET('Nutzungsberechtigt', Kategorien) >  0;
    """
    if False:
        print(select_person)

    myCursor.execute(select_person)
    result_set = myCursor.fetchall()
    if len(result_set) > 0:
        print(result_set)

    for a_person in result_set:
        update_person = f"""
             UPDATE `personen` SET `Kategorien` = removeSetValue(`Kategorien`,'Verwaltungsberechtigt')
             WHERE `ID` = {a_person['ID']};   
         """
        if False:
            print(update_person)
        myCursor.execute(update_person)
        db.commit()
        result_set = myCursor.fetchall()
        if len(result_set) > 0:
            print(result_set)

        if a_person['is_Nutzungsberechtigt'] == 'False':
            update_person = f"""
                 UPDATE `personen` SET `Kategorien` = removeSetValue(`Kategorien`,'Nutzungsberechtigt')
                 WHERE `ID` = {a_person['ID']};   
             """
            if False:
                print(update_person)
            myCursor.execute(update_person)
            db.commit()
            result_set = myCursor.fetchall()
            if len(result_set) > 0:
                print(result_set)

    if verbal:
        print('''
                Kategorien = -Verwaltungsberechtigt
                Kategorien = -Nutzungsberechtigt (Im Jahr darauf)
             done!!
             ''')




    # Falls Todestag gesetzt, Zivilstand='Gestorben' und -Verwaltungsberechtigt und Newsletter_Abonniert_Am = NULL
    # ------------------------------------------------------------------------------------------------------------
    if verbal:
        print('''--> Checking if somebody has Todestag gesetzt, ...''', end='')
    select_person = f"""
        SELECT ID 
        FROM personen_daten 
        WHERE Todestag IS NOT NULL AND 
              FIND_IN_SET('Verwaltungsberechtigt', Kategorien) >  0;
    """
    if False:
        print(select_person)

    myCursor.execute(select_person)
    result_set = myCursor.fetchall()
    if len(result_set) > 0:
        print(result_set)
    for pers_id in result_set:

        update_person = f"""
            UPDATE `personen` SET `Kategorien` = removeSetValue(`Kategorien`,'Verwaltungsberechtigt'),
                                  `Zivilstand` = 'Gestorben',
                                  `Newsletter_Abonniert_Am` = NULL 
            WHERE `ID` = {pers_id['ID']};   
        """
        if False:
            print(update_person)
        myCursor.execute(update_person)
        db.commit()
        result_set = myCursor.fetchall()
        if len(result_set) > 0:
            print(result_set)

    if verbal:
        print('''
                Zivilstand = Gestorben
                Kategorien = -Verwaltungsberechtigt
                Newsletter_Abonniert_Am = NULL
             done!!
             ''')

    # Falls Todesjahr letztes Jahr oder älter ==> -Nutzungsberechtigt
    # ---------------------------------------------------------------
    if verbal:
        print('''--> Checking if somebody has passed away last year or earlier ...''', end='')
    select_person = f"""
        SELECT ID 
        FROM personen_daten 
        WHERE Todestag IS NOT NULL AND 
              FIND_IN_SET('Nutzungsberechtigt', Kategorien) >  0 AND 
              Todesjahr < DATE_FORMAT(now(),'%Y');
    """
    if False:
        print(select_person)

    myCursor.execute(select_person)
    result_set = myCursor.fetchall()
    if len(result_set) > 0:
        print(result_set)
    for pers_id in result_set:
        update_person = f"""
            UPDATE `personen` SET `Kategorien` = removeSetValue(`Kategorien`,'Nutzungsberechtigt')
            WHERE `ID` = {pers_id['ID']};   
        """
        if verbal:
            print(update_person)
        myCursor.execute(update_person)
        db.commit()
        result_set = myCursor.fetchall()
        if verbal:
            print(result_set)

    if verbal:
        print('''
                Kategorien = -Nutzungsberechtigt           
             done!!
             ''')

    # Re-Setzt Hat_16a und Hat_35a Teil neu
    # -------------------------------------
    if verbal:
        print('''--> Re-Setzt Hat_16a und Hat_35a Teil neu ...''', end='\n')
    res_count = remove_all_enum_value_in_set(db, 'personen', 'Kategorien', enum_val_to_remove='Hat_16a', take_action=True, verbal=False)
    if verbal and res_count > 0:
        print(f'Von {res_count:3d} Bürgern wurde der "Hat_16a" Teil entfernt ')

    res_count = remove_all_enum_value_in_set(db, 'personen', 'Kategorien', enum_val_to_remove='Hat_35a', take_action=True, verbal=False)
    if verbal and res_count > 0:
        print(f'Von {res_count:3d} Bürgern wurde der "Hat_35a" Teil entfernt ')


    count_of_16a_Teile = 0
    count_of_35a_Teile = 0
    select_person = f"""
        SELECT * 
        FROM Buergerteile;
    """
    if False:
        print(select_person)

    myCursor.execute(select_person)
    result_set = myCursor.fetchall()
    if len(result_set) > 0:
        # print(result_set)
        for a_buergerteil in result_set:
            if False:
                print(f"""{a_buergerteil['Verpaechter_Kategorien']} {a_buergerteil['Verpaechter_ID']}
                 {a_buergerteil['Verpaechter_Vorname']}
                 {a_buergerteil['Verpaechter_Name']} 
                 {a_buergerteil['Verpaechter_Strasse']}
                 {a_buergerteil['Verpaechter_PLZ_Ort']}
                 {a_buergerteil['Flaeche']}\n""")
            # print(f"""{a_buergerteil['Verpaechter_Kategorien']} {a_buergerteil['Verpaechter_ID']} {a_buergerteil['Flaeche']}\n""")
            flaeche_in_aren = a_buergerteil['Flaeche']
            if flaeche_in_aren > 15.0 and flaeche_in_aren < 17.0:
                update_person = f"""
                    UPDATE `personen` SET `Kategorien` = addSetValue(`Kategorien`,'HAT_16a')
                    WHERE `ID` = {a_buergerteil['Verpaechter_ID']};   
                """
                if False:
                    print(update_person)
                myCursor.execute(update_person)
                if myCursor.rowcount == 1:
                    count_of_16a_Teile += 1
                    db.commit()

            if flaeche_in_aren > 30.0 and flaeche_in_aren < 37.0:
                update_person = f"""
                    UPDATE `personen` SET `Kategorien` = addSetValue(`Kategorien`,'HAT_35a')
                    WHERE `ID` = {a_buergerteil['Verpaechter_ID']};   
                """
                if False:
                    print(update_person)
                myCursor.execute(update_person)
                if myCursor.rowcount == 1:
                    count_of_35a_Teile += 1
                    db.commit()


    if verbal:
        print(f'Bei {count_of_16a_Teile:3d} Bürgern wurde "Hat_16a" gesetzt ')
        print(f'Bei {count_of_35a_Teile:3d} Bürgern wurde "Hat_35a" gesetzt ')
        print('''             done!!
             ''')


def addPersonen_to_db_by_hash(db, arguments, take_action=False, verbal=False):
    verbal = True
    if verbal:
        print(f'''
           --> Calling addPersonen_to_db_by_hash(db,
                                    arguments               = {arguments}, 
                                    take_action             = {take_action},
                                    verbal                  = {verbal})''')
    pers_id = None
    if take_action:
        pers_id = addPersonen_to_db(db,
                                 arguments['Source'],
                                 arguments['Vorname'],
                                 arguments['Ledig_Name'],
                                 arguments['Partner_Name'],
                                 arguments['Partner_Name_Angenommen'],
                                 arguments['Private_Adressen_ID'],
                                 take_action=take_action,
                                 verbal=verbal)
    return pers_id


def addPersonen_to_db(db, source, vorname, ledig_name, partner_name, partner_name_angenommen, privat_adressen_id=701, take_action=False, verbal=False):
    try:
        if privat_adressen_id == '':
            privat_adressen_id = 701  # Unbekannte Strasse an unbekanntem Ort (4797)
        if verbal:
            print(f'''
               --> Calling addPersonen_to_db(db,
                                        source                  = {source}, 
                                        vorname                 = {vorname},
                                        ledig_name              = {ledig_name},
                                        partner_name            = {partner_name},
                                        partner_name_angenommen = {partner_name_angenommen},
                                        privat_adressen_id      = {privat_adressen_id},
                                        take_action             = {take_action},
                                        verbal                  = {verbal})''')

        pers_id = None

        if take_action:
            myCursor = db.cursor()
            args = (source, vorname, ledig_name, partner_name, partner_name_angenommen, privat_adressen_id, 'x')
            result_args = myCursor.callproc('addPersonen', args)
            db.commit()
            if verbal:
                print(f'callproc:addPersonen{args}')
                print(f'        :return value{result_args}')
            pers_id = result_args[-1]

        return pers_id

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None

# telnr functions
# ---------------
def CRUD_telnr_in_db(db, pers_id, telnr_detailed, take_action=False, verbal=False):
    this_fct_call = f'''
       --> Calling CRUD_telnr_in_db(db,
                        pers_id         = {pers_id}, 
                        email_detailed  = {telnr_detailed},
                        take_action     = {take_action},
                        verbal          = {verbal})'''

    try:
        myCursor = db.cursor()
        if verbal:
            print(this_fct_call)
        new_telnr_details = split_telnr_detailed_long(telnr_detailed)
        if verbal:
            print('splited:', new_telnr_details)
        current_telnrs_for_this_person = get_telnr_details_persid(db, pers_id, verbal=verbal)
        if current_telnrs_for_this_person is not None:
            if verbal:
                print('Old Telnrs for', pers_id, current_telnrs_for_this_person)

            # merge Telnrs

        if new_telnr_details['ID'] == '?':
            print(f"addTelnr({pers_id}, {new_telnr_details})")
            args = (pers_id, new_telnr_details['Laendercode'], new_telnr_details['Vorwahl'], new_telnr_details['Nummer'], new_telnr_details['Type'], new_telnr_details['Endgeraet'], new_telnr_details['Prio'], 'x')  # 'x' in the Tuple will be replaced by OUT-Value
            if take_action:
                if verbal:
                    print(f"""...call proc .... addTelnr{args}""")
                result_args = myCursor.callproc('addTelnr', args)
        elif new_telnr_details['Nummer'] == 'NULL':
            print(f"delete_telnr({pers_id}, {new_telnr_details['ID']})")
            args = (pers_id, new_telnr_details['ID'])

            if take_action:
                if verbal:
                    print(f"""...call proc .... deleteTelnr{args}""")
                result_args = myCursor.callproc('deleteTelnr', args)
        else:
            print(f"update_Telnr({pers_id}, {new_telnr_details})")
            args = (new_telnr_details['ID'], new_telnr_details['Laendercode'], new_telnr_details['Vorwahl'], new_telnr_details['Nummer'], new_telnr_details['Type'], new_telnr_details['Endgeraet'], new_telnr_details['Prio'])

            if take_action:
                if verbal:
                    print(f"""...call proc .... updateTelnr{args}""")
                result_args = myCursor.callproc('updateTelnr', args)


    except mysql.connector.Error as err:
        print(f"ERROR: {err}")
        print(this_fct_call)
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        print(this_fct_call)
        return None


def get_telnr_ids_for_persid(db, pers_id, verbal=False):
    ret_list = []
    sql_insert = f"""
            SELECT Telefonnummern_ID 
            FROM personen_has_telefonnummern
            WHERE Personen_ID = {pers_id}
             """
    if verbal:
        print(f"""... in get_telnr_ids_for_persid...
           {sql_insert}
           """)
    mycursor = db.cursor()
    mycursor.execute(sql_insert)
    for aId in mycursor.fetchall():
        ret_list.append(str(aId[0]))
    return ret_list


def get_telnr_details_for_ids(db, telnr_ids, verbal=False):
    telnr_ids_where_clause = ', '.join(telnr_ids)
    if verbal:
        print(f"""
            get_telnr_details_for_ids(db, 
                  {telnr_ids})

        telnr_ids_where_clause:{telnr_ids_where_clause}    
        """)
    if len(telnr_ids) > 0:
        sql_insert = f"""
                SELECT ID, Laendercode, Vorwahl, Nummer, Type, Endgeraet, Prio 
                FROM telefonnummern 
                WHERE ID in ({telnr_ids_where_clause})
                ORDER BY Prio
                """
        if verbal:
            print(f"""... in get_telnr_details_for_ids...
               {sql_insert}
               """)
        mycursor = db.cursor(dictionary=True)
        mycursor.execute(sql_insert)
        return mycursor.fetchall()
    else:
        return None


def get_telnr_details_persid(db, pers_id, verbal=False):
    telnr_ids = get_telnr_ids_for_persid(db, pers_id, verbal=verbal)
    return get_telnr_details_for_ids(db, telnr_ids, verbal=verbal)


def TBC_fix_email_details_persid(db, pers_id, verbal=False):
    email_details = get_email_details_persid(db, pers_id, verbal=verbal)
    last_prio = 0
    for a_eMail_details in email_details:
        a_eMail_details['eMail'] = a_eMail_details['eMail'].lower().replace(' ', '')
        a_eMail_details['Prio'] = last_prio
        last_prio += 1
    print(email_details)


def split_telnr_detailed_long(detailed_long_str, prio_default=0, type_default='Sonstige', type_endgeraet='Festnetz', verbal=False, telnr_only=False):
    if verbal:
        print(f" called .... split_telnr_detailed_long('{detailed_long_str}', '{verbal}')")
    detailed_long_str = detailed_long_str.replace(' ', '')
    detailed_list = detailed_long_str.split(':')
    if len(detailed_list) > 2:
        if verbal:
            print(detailed_list)
        ret_value = {'Laendercode': '0041',
                     'Vorwahl': detailed_list[0][0:3],
                     'Nummer': detailed_list[0][3:10],
                     'ID': detailed_list[1],
                     'Prio': detailed_list[2],
                     'Type': detailed_list[3],
                     'Endgeraet': detailed_list[4],
                     }
    elif len(detailed_list) > 1:
        ret_value = {'Laendercode': '',
                     'Vorwahl': '',
                     'Nummer': detailed_list[0],
                     'ID': detailed_list[1],
                     'Prio': prio_default,
                     'Type': type_default,
                     'Endgeraet': type_endgeraet,
                     }
    else:
        ret_value = {'Laendercode': '',
                     'Vorwahl': '',
                     'Nummer': detailed_list[0],
                     'ID': '?',
                     'Prio': prio_default,
                     'Type': type_default,
                     'Endgeraet': type_endgeraet,
                     }

    if telnr_only:
        ret_value = ret_value['Vorwahl'] + ' ' + ret_value['Nummer']
    return ret_value

# email functions
# ---------------
def CRUD_email_in_db(db, pers_id, email_detailed, take_action=False, verbal=False):
    this_fct_call = f'''
       --> Calling CRUD_email_in_db(db,
                        pers_id         = {pers_id}, 
                        email_detailed  = {email_detailed},
                        take_action     = {take_action},
                        verbal          = {verbal})'''
    try:
        myCursor = db.cursor()
        if verbal:
            print(this_fct_call)
        new_email_details = split_email_detailed_long(email_detailed)
        if verbal:
            print('splited:', new_email_details)
        current_emails_for_this_person = get_email_details_persid(db, pers_id, verbal=verbal)
        if current_emails_for_this_person is not None:
            if verbal:
                print('Old emails for', pers_id, current_emails_for_this_person)

            # merge eMails

        if new_email_details['ID'] == '?':
            print(f"add_email({pers_id}, {new_email_details})")
            args = (pers_id, new_email_details['eMail'], new_email_details['Type'], new_email_details['Prio'], 'x')  # 'x' in the Tuple will be replaced by OUT-Value
            if take_action:
                if verbal:
                    print(f"""...call proc .... addEmailAdr{args}""")
                result_args = myCursor.callproc('addEmailAdr', args)
        elif new_email_details['eMail'] == 'NULL':
            print(f"delete_email({pers_id}, {new_email_details['ID']})")
            args = (pers_id, new_email_details['ID'])

            if take_action:
                if verbal:
                    print(f"""...call proc .... deleteEmailAdr{args}""")
                result_args = myCursor.callproc('deleteEmailAdr', args)
        else:
            print(f"update_email({pers_id}, {new_email_details})")
            args = (new_email_details['ID'], new_email_details['eMail'], new_email_details['Type'], new_email_details['Prio'])

            if take_action:
                if verbal:
                    print(f"""...call proc .... updateEmailAdr{args}""")
                result_args = myCursor.callproc('updateEmailAdr', args)


    except mysql.connector.Error as err:
        print(f"ERROR: {err}")
        print(this_fct_call)
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        print(this_fct_call)
        return None

def get_email_ids_for_persid(db, pers_id, verbal=False):
    ret_list = []
    sql_insert = f"""
            SELECT EMail_adressen_ID 
            FROM personen_has_email_adressen 
            WHERE Personen_ID = {pers_id}
             """
    if verbal:
        print(f"""... in get_email_ids_for_persid...
           {sql_insert}
           """)
    mycursor = db.cursor()
    mycursor.execute(sql_insert)
    for aId in mycursor.fetchall():
        ret_list.append(str(aId[0]))
    return ret_list

def get_email_details_for_ids(db, email_ids, verbal=False):
    email_ids_where_clause = ', '.join(email_ids)
    if verbal:
        print(f"""
            get_email_details_for_ids(db, 
                  {email_ids})
                  
        email_ids_where_clause:{email_ids_where_clause}    
        """)
    if len(email_ids) > 0:
        sql_insert = f"""
                SELECT ID, eMail, Type, Prio 
                FROM email_adressen 
                WHERE ID in ({email_ids_where_clause})
                ORDER BY Prio
                """
        if verbal:
            print(f"""... in get_email_details_for_ids...
               {sql_insert}
               """)
        mycursor = db.cursor(dictionary=True)
        mycursor.execute(sql_insert)
        return mycursor.fetchall()
    else:
        return None

def get_email_details_persid(db, pers_id, verbal=False):
    email_ids = get_email_ids_for_persid(db, pers_id, verbal=verbal)
    return get_email_details_for_ids(db, email_ids, verbal=verbal)

def fix_email_details_persid(db, pers_id, verbal=False):
    email_details = get_email_details_persid(db, pers_id, verbal=verbal)
    last_prio = 0
    for a_eMail_details in email_details:
        a_eMail_details['eMail'] = a_eMail_details['eMail'].lower().replace(' ', '')
        a_eMail_details['Prio'] = last_prio
        last_prio += 1
    print(email_details)

def split_email_detailed_long(detailed_long_str, type_default='Sonstige', prio_default=0, verbal=False, email_only=False):
    if verbal:
        print(f" called .... split_email_detailed_long('{detailed_long_str}', '{verbal}')")
    detailed_long_str = detailed_long_str.replace(' ', '')
    detailed_list = detailed_long_str.split(':')
    if len(detailed_list) > 2:
        if verbal:
            print(detailed_list)
        ret_value = {'ID': detailed_list[1],
                     'eMail': detailed_list[0],
                     'Type': detailed_list[3],
                     'Prio': detailed_list[2],
                     }
    elif len(detailed_list) > 1:
        ret_value = {'ID': detailed_list[1],
                     'eMail': detailed_list[0],
                     'Type': type_default,
                     'Prio': prio_default,
                     }
    else:
        ret_value = {'ID': '?',
                     'eMail': detailed_list[0],
                     'Type': type_default,
                     'Prio': prio_default,
                     }

    if email_only:
        ret_value = ret_value['eMail']
    return ret_value

def AUTO_TEST__split_email_detailed_long(verbal=True):
    print('\n\n==> AUTO_TEST__split_email_detailed_long.....\n')
    print('1) ', split_email_detailed_long('walter@rothlin.com  :321:  0:Sonstige:', verbal=verbal), '\n' + generateStringRepeats(10, '----'))
    print('2) ', split_email_detailed_long('walter.rothlin@fh-hwz.ch  :320:  1:Geschaeft:', verbal=verbal), '\n' + generateStringRepeats(10, '----'))
    print('3) ', split_email_detailed_long('walterrr@rothlin.com', verbal=verbal), '\n' + generateStringRepeats(10, '----'))
    print('\n\n... AUTO_TEST__split_email_detailed_long Done!!\n')

# telnr functions
# ---------------
def CRUD_IBAN_in_db(db, pers_id, IBAN_detailed, take_action=False, verbal=False):
    this_fct_call = f'''
       --> Calling CRUD_IBAN_in_db(db,
                        pers_id         = {pers_id}, 
                        IBAN_detailed   = {IBAN_detailed},
                        take_action     = {take_action},
                        verbal          = {verbal})'''

    try:
        myCursor = db.cursor()
        if verbal:
            print(this_fct_call)
        new_IBAN_details = split_IBAN_detailed_long(IBAN_detailed)
        if verbal:
            print('splited:', new_IBAN_details)

        current_IBANs_for_this_person = get_IBAN_details_persid(db, pers_id, verbal=verbal)
        if current_IBANs_for_this_person is not None:
            if verbal:
                print('Old IBANs for', pers_id, current_IBANs_for_this_person)

            # merge Telnrs

        if new_IBAN_details['ID'] == '?':
            print(f"addIBAN({pers_id}, {new_IBAN_details})")
            args = (pers_id, new_IBAN_details['Nummer'], new_IBAN_details['Bezeichnung'], new_IBAN_details['Bankname'], new_IBAN_details['Bankort'], new_IBAN_details['Prio'], 'x')  # 'x' in the Tuple will be replaced by OUT-Value
            if take_action:
                if verbal:
                    print(f"""...call proc .... addIBAN{args}""")
                result_args = myCursor.callproc('addIBAN', args)
        elif new_IBAN_details['Nummer'] == 'NULL':
            print(f"delete_telnr({pers_id}, {new_IBAN_details['ID']})")
            args = (pers_id, new_IBAN_details['ID'])

            if take_action:
                if verbal:
                    print(f"""...call proc .... deleteIBAN{args}""")
                result_args = myCursor.callproc('deleteIBAN', args)
        else:
            print(f"update_IBAN({pers_id}, {new_IBAN_details})")
            args = (pers_id, new_IBAN_details['Nummer'], new_IBAN_details['Bezeichnung'], new_IBAN_details['Bankname'], new_IBAN_details['Bankort'], new_IBAN_details['Prio'], new_IBAN_details['ID'])

            if take_action:
                if verbal:
                    print(f"""...call proc .... updateIBAN{args}""")
                result_args = myCursor.callproc('updateIBAN', args)


    except mysql.connector.Error as err:
        print(f"ERROR: {err}")
        print(this_fct_call)
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        print(this_fct_call)
        return None


def get_IBAN_ids_for_persid(db, pers_id, verbal=False):
    ret_list = []
    sql_insert = f"""
            SELECT ID 
            FROM iban
            WHERE Personen_ID = {pers_id}
             """
    if verbal:
        print(f"""... in get_IBAN_ids_for_persid...
           {sql_insert}
           """)
    mycursor = db.cursor()
    mycursor.execute(sql_insert)
    for aId in mycursor.fetchall():
        ret_list.append(str(aId[0]))
    return ret_list


def get_IBAN_details_for_ids(db, IBAN_ids, verbal=False):
    IBAN_ids_where_clause = ', '.join(IBAN_ids)
    if verbal:
        print(f"""
            get_IBAN_details_for_ids(db, 
                  {IBAN_ids})

        IBAN_ids_where_clause:{IBAN_ids_where_clause}    
        """)
    if len(IBAN_ids) > 0:
        sql_insert = f"""
                SELECT ID, Nummer, Bezeichnung, Bankname, Bankort, Prio 
                FROM iban 
                WHERE ID in ({IBAN_ids_where_clause})
                ORDER BY Prio
                """
        if verbal:
            print(f"""... in get_IBAN_details_for_ids...
               {sql_insert}
               """)
        mycursor = db.cursor(dictionary=True)
        mycursor.execute(sql_insert)
        return mycursor.fetchall()
    else:
        return None


def get_IBAN_details_persid(db, pers_id, verbal=False):
    IBAN_ids = get_IBAN_ids_for_persid(db, pers_id, verbal=verbal)
    return get_IBAN_details_for_ids(db, IBAN_ids, verbal=verbal)


def TBC_fix_email_details_persid(db, pers_id, verbal=False):
    email_details = get_email_details_persid(db, pers_id, verbal=verbal)
    last_prio = 0
    for a_eMail_details in email_details:
        a_eMail_details['eMail'] = a_eMail_details['eMail'].lower().replace(' ', '')
        a_eMail_details['Prio'] = last_prio
        last_prio += 1
    print(email_details)


def split_IBAN_detailed_long(detailed_long_str, prio_default=0, bezeichnung_default='', bankname_default='', bankort_default='', verbal=False, IBAN_only=False):
    if verbal:
        print(f" called .... split_IBAN_detailed_long('{detailed_long_str}', '{verbal}')")
    detailed_long_str = detailed_long_str.replace(' ', '')
    detailed_list = detailed_long_str.split(':')

    # print('CH97 0077 7001 5721 7077 8:', format_IBAN('CH97 0077 7001 5721 7077 8'))
    if len(detailed_list) > 2:
        if verbal:
            print(detailed_list)
        ret_value = {'Nummer': format_IBAN(detailed_list[0]),
                     'ID': detailed_list[1],
                     'Prio': detailed_list[2],
                     'Bankname': detailed_list[3],
                     'Bankort': detailed_list[4],
                     'Bezeichnung': detailed_list[5],
                     }
    elif len(detailed_list) > 1:
        ret_value = {'Nummer': detailed_list[0],
                     'ID': detailed_list[1],
                     'Prio': prio_default,
                     'Bankname': bankname_default,
                     'Bankort': bankort_default,
                     'Bezeichnung': bezeichnung_default,
                     }
    else:
        ret_value = {'Nummer': detailed_list[0],
                     'ID': '?',
                     'Prio': prio_default,
                     'Bankname': bankname_default,
                     'Bankort': bankort_default,
                     'Bezeichnung': bezeichnung_default,
                     }

    if IBAN_only:
        ret_value = ret_value['Vorwahl'] + ' ' + ret_value['Nummer']
    return ret_value


# div  functions
# ---------------
def update_or_insert_value(db_connection, pers_id, attribute_name, value, verbal=False):
    count_of_inserts = 0
    count_of_updates = 0
    count_of_matches = 0
    myCursor = db_connection.cursor()
    if verbal:
        print(f"update_or_insert_value({pers_id}, '{attribute_name}', '{value}', {verbal})")
    if attribute_name == 'eMai_adressen':
        sql_select = f'SELECT eMail_adresse, Type, Prio FROM email_liste WHERE Pers_ID={pers_id}'
        myCursor.execute(sql_select)
        myresult = myCursor.fetchall()
        if len(myresult) == 0:
            print(f"add_email({pers_id}, {value['eMail'].lower()})")
            args = (pers_id, value['eMail'].lower(), 'Sonstige', 0, 'x')  # 'x' in the Tuple will be replaced by OUT-Value
            if verbal:
                print('    addEmailAdr', str(args), sep='')
            result_args = myCursor.callproc('addEmailAdr', args)
            count_of_inserts += 1
        else:
            for a_email_adr in myresult:
                if a_email_adr[0].lower() == value['eMail'].lower():
                    if verbal:
                        print('Matches: ', a_email_adr[0], value['eMail'])
                    count_of_matches += 1
                    break
                else:
                    print(f'{pers_id:4s}   DB:', a_email_adr[0])
                    print('New_Value:', value['eMail'], '\n')
                    current_emails = get_email_ids_for_persid(db_connection, pers_id)
                    if len(current_emails) == 1:
                        old_email_details = get_email_details_for_ids(db_connection, current_emails)[0]
                        sql_del = f'DELETE FROM personen_has_email_adressen WHERE Personen_ID={pers_id} AND EMail_adressen_ID={old_email_details[0]}'
                        print(old_email_details)
                        print('sql_del:', sql_del)
                        myCursor = db_connection.cursor()
                        myCursor.execute(sql_del)
                        db_connection.commit()

    if attribute_name == 'telefon_nummer':
        sql_select = f'SELECT Tel_Nr, Type, Endgeraet, Prio FROM Telnr_liste WHERE Pers_ID={pers_id}'
        myCursor.execute(sql_select)
        myresult = myCursor.fetchall()
        value['tel_nr'] = value['tel_nr'].replace(' ', '')
        if len(myresult) == 0:
            print(f"addTelNr({pers_id}, {value['tel_nr']})")
            vorwahl = value['tel_nr'][0:3]
            telnr = value['tel_nr'][3:]
            if vorwahl == '055':
                endgeraet = 'Festnetz'
            else:
                endgeraet = 'Mobile'
            args = (pers_id, '0041', vorwahl, telnr, 'Sonstige', endgeraet, 0, 'x')
            if verbal:
                print('    addTelNr', str(args), sep='')
            result_args = myCursor.callproc('addTelNr', args)
            count_of_inserts += 1
        else:
            for a_telnr in myresult:
                if verbal:
                    print('FromDB  :', a_telnr)
                    print('FromParm:', value)
                if a_telnr[0].replace(' ', '') == value['tel_nr'].replace(' ', ''):
                    if verbal:
                        print('Matches: ', a_telnr[0], value['tel_nr'])
                    count_of_matches += 1
                    break
                else:
                    print(f'{pers_id:4s}   DB:', a_telnr[0])
                    print('New_Value:', value['tel_nr'], '\n')
                    if verbal:
                        input('weiter_3')

    if attribute_name == 'Geburtstag' or attribute_name == 'Newsletter_Abonniert_Am':
        a_date = str(value['Date'])
        # print('a_date         :', a_date)
        dd = a_date[8:10]
        mm = a_date[5:7]
        yyyy = a_date[:4]
        a_date = dd + '.' + mm + '.' + yyyy
        # print('a_date         :', a_date, '\n\n')

        sql_select = f"SELECT DATE_FORMAT({attribute_name},'%d.%m.%Y') FROM Personen WHERE ID={pers_id}"
        # print(sql_select)
        myCursor.execute(sql_select)
        myresult = myCursor.fetchall()
        # print(myresult)
        a_date_from_db = myresult[0][0]
        if verbal:
            print('FromDB  :', a_date_from_db)
            print('FromParm:', a_date)
        if myresult[0][0] == a_date:
            if verbal:
                print('Match Date: ', a_date_from_db, a_date)
            count_of_matches += 1
        else:
            sql_update = f"Update Personen SET {attribute_name} = STR_TO_DATE('{a_date}','%d.%m.%Y') WHERE ID={pers_id}"
            myCursor.execute(sql_update)
            count_of_updates += myCursor.rowcount
            db_connection.commit()
            if verbal:
                print(f'{pers_id:4s}   DB:', a_date_from_db)
                print('New_Value:', a_date, '\n')
                print('sql_select:', sql_update)
                input('weiter_4')

    return {'count_of_inserts': count_of_inserts,
            'count_of_updates': count_of_updates,
            'count_of_matches': count_of_matches}





# ---------------------------------------------
# Functions for Common use
# ---------------------------------------------

def get_ort_id_from_db_by_pers_id(db=None, pers_id=None, take_action=False, verbal=False):
    if verbal:
        print(f'''
           --> Calling get_ort_id_from_db_by_pers_id(db,
                                    pers_id          = {pers_id}, 
                                    take_action      = {take_action},
                                    verbal           = {verbal})''')

    myCursor = db.cursor()
    ret_id = -1
    args = (pers_id, 'x')
    if take_action:
        if verbal:
            print(f"""...call proc .... getOrtIdFromPersId{args}""")
        result_args = myCursor.callproc('getOrtIdFromPersId', args)
        ret_id = result_args[1]
    return ret_id

def get_ort_id_from_db_by_plz_ort(db=None, plz_ort='', kanton='SZ', land_id=1, take_action=False, verbal=False):
    if verbal:
        print(f'''
           --> Calling get_ort_id_from_db_by_plz_ort(db,
                                    plz_ort          = {plz_ort}, 
                                    take_action      = {take_action},
                                    verbal           = {verbal})''')

    myCursor = db.cursor()
    ret_id = -1
    result = split_plz_ort(plz_ort)
    args = (result['PLZ'], result['Ort'], kanton, land_id, 'x')
    if take_action:
        if verbal:
            print(f"""...call proc .... getOrtId{args}""")
        result_args = myCursor.callproc('getOrtId', args)
        ret_id = result_args[4]
    return ret_id

def get_ort_id_from_ws_or_db(ws=None, db=None, pers_id=None, title_row=None, row=None, do_reset_cell=False, reset_cell_value=None, take_action=False, verbal=False):
    if verbal:
        print(f'''
           --> Calling get_ort_id_from_ws_or_db(ws, db,
                                    pers_id          = {pers_id}, 
                                    title_row        = {title_row},
                                    row              = {row},
                                    do_reset_cell    = {do_reset_cell},
                                    reset_cell_value = {reset_cell_value},
                                    take_action      = {take_action},
                                    verbal           = {verbal})''')

    values = get_cell_values_by_column_titles(
        ws,
        title_row=title_row,
        row=row,
        column_names=['Private_Ort_ID', 'Private_PLZ_Ort'],
        do_reset_cell=take_action,
        reset_cell_value='',
        take_action=take_action,
        verbal=False)

    if values['Private_Ort_ID'] == '':
        if values['Private_PLZ_Ort'] == '':
            values['Private_Ort_ID'] = get_ort_id_from_db_by_pers_id(db=db, pers_id=pers_id, take_action=take_action, verbal=verbal)
        else:
            values['Private_Ort_ID'] = get_ort_id_from_db_by_plz_ort(db=db, plz_ort=values['Private_PLZ_Ort'], take_action=take_action, verbal=verbal)
    ret_values = values
    return ret_values

def get_adress_id_from_db_by_strassen_adresse(db=None, strassen_adresse='', ort_id='', take_action=False, verbal=False):
    if verbal:
        print(f'''
           --> Calling get_adress_id_from_db_by_strassen_adresse(db,
                                    strassen_adresse = {strassen_adresse}, 
                                    ort_id           = {ort_id}, 
                                    take_action      = {take_action},
                                    verbal           = {verbal})''')

    myCursor = db.cursor()
    ret_id = -1
    result = split_adress_street_nr(strassen_adresse)
    args = (result['Street'], result['No'], ort_id, 'x')
    if take_action:
        if verbal:
            print(f"""...call proc .... getAdressId_by_ortID{args}""")
        result_args = myCursor.callproc('getAdressId_by_ortID', args)
        ret_id = result_args[3]

    return ret_id

def pre_process_CUD(db, reco_data_fn, reco_sheetname, db_attr_excel_column_mapping=None, verbal=False, take_action=False):
    if verbal:
        print(f'''
           --> Calling pre_process_CUD(stammdaten_schema,
                    File    = {reco_data_fn}, 
                    Sheet   = {reco_sheetname},
                    Mapping = {db_attr_excel_column_mapping}  
                    verbal  = {verbal}, 
                    take_action = {take_action})''')

        print(f'Open Excel:{reco_data_fn} [{reco_sheetname}]')

    # Open and read the Excel
    # -----------------------
    workbook = openpyxl.load_workbook(reco_data_fn, data_only=True)
    worksheet_sheet = workbook[reco_sheetname]

    statistics = {
            'Familien_Name': 0,
            'Private_PLZ_Ort': 0,
            'Private_Strassen_Adresse': 0}

    title_row = 1
    first_value_row = 2
    row = first_value_row
    count_of_total_lines = 0
    while worksheet_sheet["A" + str(row)].value is not None:
        id_value_from_excel = worksheet_sheet["A" + str(row)].value
        count_of_total_lines += 1
        for a_mapping in db_attr_excel_column_mapping:
            # print('--> Preprocessor field:',  a_mapping['excel'])
            excel_column_name = a_mapping['excel']

            values = get_cell_values_by_column_titles(
                worksheet_sheet,
                title_row=title_row,
                row=row,
                column_names=[excel_column_name],
                do_reset_cell=take_action,
                reset_cell_value='',
                take_action=take_action,
                verbal=False)

            new_value_from_excel = values[excel_column_name]
            if new_value_from_excel is not None and new_value_from_excel != '':
                print(f"{row:5d}: {a_mapping['excel']:20s}:{new_value_from_excel:20s}")

                if excel_column_name == 'Familien_Name':
                    statistics['Familien_Name'] += 1
                    sex_value = get_cell_values_by_column_titles(
                        worksheet_sheet,
                        title_row=title_row,
                        row=row,
                        column_names=['Geschlecht'],
                        do_reset_cell=False,
                        reset_cell_value='',
                        take_action=take_action,
                        verbal=False)
                    sex_value = sex_value['Geschlecht']
                    ret_val = split_familien_name(new_value_from_excel, sex_value, verbal=False)
                    # print('1) ', ret_val)
                    set_cell_value_by_column_title(ret_val['Ledig_Name']             , worksheet_sheet, title_row=title_row, row=row, column_name='Ledig_Name'             , take_action=True, verbal=False)
                    set_cell_value_by_column_title(ret_val['Partner_Name']           , worksheet_sheet, title_row=title_row, row=row, column_name='Partner_Name'           , take_action=True, verbal=False)
                    set_cell_value_by_column_title(ret_val['Partner_Name_Angenommen'], worksheet_sheet, title_row=title_row, row=row, column_name='Partner_Name_Angenommen', take_action=True, verbal=False)
                    print(f"       -->  {ret_val['Ledig_Name']:20s}:{ret_val['Partner_Name']:20s}:{ret_val['Partner_Name_Angenommen']:1d}")

                elif excel_column_name == 'Private_PLZ_Ort':
                    statistics['Private_PLZ_Ort'] += 1
                    ort_id = get_ort_id_from_db_by_plz_ort(db=db, plz_ort=new_value_from_excel, take_action=take_action, verbal=False)
                    # print('1) ', ort_id)
                    ret_val = set_cell_value_by_column_title(ort_id, worksheet_sheet, title_row=title_row, row=row, column_name='Private_Ort_ID', take_action=True, verbal=False)
                    # print('2) ', ret_val)
                    print(f"       -->  {ort_id:5d}")

                elif excel_column_name == 'Private_Strassen_Adresse':
                    statistics['Private_Strassen_Adresse'] += 1
                    new_value_from_excel = new_value_from_excel.replace('strasse', 'str.')
                    # print(f'    Private_Strasse: "{excel_column_name}": "{new_value_from_excel}"')
                    ort_id = get_cell_values_by_column_titles(
                        worksheet_sheet,
                        title_row=title_row,
                        row=row,
                        column_names=['Private_Ort_ID'],
                        do_reset_cell=True,
                        reset_cell_value='',
                        take_action=take_action,
                        verbal=False)['Private_Ort_ID']
                    # print('1) ', ort_id)
                    addr_id = get_adress_id_from_db_by_strassen_adresse(db=db, strassen_adresse=new_value_from_excel, ort_id=ort_id,          take_action=True, verbal=False)
                    set_cell_value_by_column_title(addr_id, worksheet_sheet, title_row=title_row, row=row, column_name='Private_Adressen_ID', take_action=True, verbal=False)
                    print(f"       -->  {addr_id:5d}")

                else:
                    print(end='')

        # get details for id from DB
        if id_value_from_excel == 'P':
            pers_values = get_cell_values_by_column_titles(worksheet_sheet, title_row=title_row, row=row, column_names=['Vorname', 'Ledig_Name', 'Partner_Name'], do_reset_cell=False, reset_cell_value=None, take_action=False, verbal=False)
            key_values = {key: value for key, value in pers_values.items() if is_not_empty(value)}

            non_empty_values = []
            for a_value in key_values.values():
                non_empty_values.append(a_value)

            ## print('non_empty_values:', non_empty_values)
            ret_val = get_personen_ids(db, such_kriterien=non_empty_values, verbal=True)
            ## print('PPPPP', ret_val)
            if len(ret_val) == 0:
                set_cell_value_by_column_title('?', worksheet_sheet, title_row=title_row, row=row, column_name='ID', take_action=True, verbal=False)
            elif len(ret_val) == 1:
                set_cell_value_by_column_title(ret_val[0][0], worksheet_sheet, title_row=title_row, row=row, column_name='ID', take_action=True, verbal=False)
                get_cell_values_by_column_titles(worksheet_sheet, title_row=title_row, row=row, column_names=['Vorname', 'Ledig_Name', 'Partner_Name'], do_reset_cell=True, reset_cell_value='', take_action=True, verbal=False)
            else:
                set_cell_value_by_column_title(str(ret_val), worksheet_sheet, title_row=title_row, row=row, column_name='DetailsFromDB', take_action=True, verbal=False)

        elif id_value_from_excel == '?':
            pass

        else:
            ret_val = get_person_details_for_id(db, id_value_from_excel, verbal=False)
            set_cell_value_by_column_title(str(ret_val), worksheet_sheet, title_row=title_row, row=row, column_name='DetailsFromDB', take_action=True, verbal=False)

        # check if all attributes are empty
        if are_all_values_empty(worksheet_sheet, title_row=1, row=row, exclude_attr_names=['ID', 'DetailsFromDB', 'ToDoes']):
            set_cell_value_by_column_title('All Done', worksheet_sheet, title_row=title_row, row=row, column_name='ToDoes', take_action=True, verbal=False)
        else:
            set_cell_value_by_column_title('', worksheet_sheet, title_row=title_row, row=row, column_name='ToDoes', take_action=True, verbal=False)

        row += 1

    statistics['Line Processed in Excel'] = count_of_total_lines
    print('\n\nPre-Processing Statistics:\n  ', statistics)

    # Save the Excel
    # --------------
    if take_action:
        if verbal:
            print(f'Close Excel:{reco_data_fn} [{reco_sheetname}]')
        workbook.save(reco_data_fn)
        workbook.close()


def process_CUD(stammdaten_schema, reco_data_fn, reco_sheetname, excel_db_field_mapping, verbal=False, take_action=False):
    if verbal:
        print(f'''
           --> Calling process_CUD(stammdaten_schema,
                                    {reco_data_fn}, 
                                    {reco_sheetname},
                                    excel_db_field_mapping, 
                                    verbal={verbal}, 
                                    take_action={take_action})''')

        print(f'Open Excel:{reco_data_fn} [{reco_sheetname}]')

    # Open and read the Excel
    # -----------------------
    workbook = openpyxl.load_workbook(reco_data_fn, data_only=True)
    worksheet_sheet = workbook[reco_sheetname]

    verbal_while_insert = False
    verbal_while_update = verbal
    count_of_new_records = 0
    count_of_updated_records = 0
    count_of_updated_attributs = 0
    count_of_total_lines = 0



    title_row = 1
    first_value_row = 2
    row = first_value_row
    while worksheet_sheet["A" + str(row)].value is not None:
        count_of_total_lines += 1

        pers_id = get_cell_value_by_column_title(worksheet_sheet, title_row=title_row, row=row, column_name='ID', verbal=False)['ID']
        if pers_id == '?':
            count_of_new_records += 1
            values = get_cell_values_by_column_titles(
                             worksheet_sheet,
                             title_row        = title_row,
                             row              = row,
                             column_names     = ['Vorname', 'Ledig_Name', 'Partner_Name', 'Partner_Name_Angenommen', 'Private_Adressen_ID'],
                             do_reset_cell    = take_action,
                             reset_cell_value = '',
                             take_action      = take_action,
                             verbal           = verbal_while_insert)
            values['Source'] = 'Loader_1'
            ret_val   = addPersonen_to_db_by_hash(stammdaten_schema, values, take_action=take_action, verbal=verbal_while_insert)
            ret_val_1 = set_cell_value_by_column_title(ret_val, worksheet_sheet, title_row=1, row=row, column_name='ID', take_action=take_action, verbal=verbal_while_insert)
            if verbal_while_insert:
                print(ret_val, ret_val_1)

        elif pers_id != 'P':
            db_table_name = 'Personen'
            for a_mapping in excel_db_field_mapping:
                excel_column_name = a_mapping['excel']
                # print('excel_column_name: ', excel_column_name)
                '''
                if a_mapping.get('db') == 'PreProcessor':
                    print('--> Preprocessor field:', a_mapping)
                    

                    values = get_cell_values_by_column_titles(
                        worksheet_sheet,
                        title_row=title_row,
                        row=row,
                        column_names=[excel_column_name],
                        do_reset_cell=take_action,
                        reset_cell_value='',
                        take_action=take_action,
                        verbal=True)

                    new_value_from_excel = values[excel_column_name]
                    if new_value_from_excel is not None and new_value_from_excel != '':
                        if a_mapping.get('db') is not None:
                            db_attr_name = a_mapping['db']
                        else:
                            db_attr_name = a_mapping['excel']

                        if a_mapping.get('excel') is not None:
                            excel_column_name = a_mapping['excel']
                        else:
                            excel_column_name = a_mapping['db']
                        # print('db_attr_name:', db_attr_name, '    excel_column_name:', excel_column_name)
                        # halt()
                '''
                values = get_cell_values_by_column_titles(
                    worksheet_sheet,
                    title_row=title_row,
                    row=row,
                    column_names=[excel_column_name],
                    do_reset_cell=take_action,
                    reset_cell_value='',
                    take_action=take_action,
                    verbal=False)
                new_value_from_excel = values[excel_column_name]
                db_attr_name = ''
                if new_value_from_excel is not None and new_value_from_excel != '':
                    if a_mapping.get('db') is not None:
                        db_attr_name = a_mapping['db']
                    else:
                        db_attr_name = a_mapping['excel']

                    if a_mapping.get('excel') is not None:
                        excel_column_name = a_mapping['excel']
                    else:
                        excel_column_name = a_mapping['db']

                if new_value_from_excel is None or new_value_from_excel == '' or db_attr_name == 'Pre':
                    if False:
                        print(f"{pers_id:5d} No update for {excel_column_name} {db_attr_name}")
                else:
                    print(f"{pers_id:5d} {excel_column_name} --> {new_value_from_excel}")
                    if (db_attr_name == 'Join'):
                        # print("JOIN    .... ", excel_column_name, new_value_from_excel)
                        if excel_column_name.startswith('eMail'):
                            print(f'    eMail: {excel_column_name}: {new_value_from_excel}')
                            ret_val = CRUD_email_in_db(stammdaten_schema, pers_id, new_value_from_excel.replace(' ', ''), take_action=True, verbal=verbal_while_update)
                            # print(ret_val)
                        elif excel_column_name.startswith('Tel_Nr'):
                            print(f'    Tel_Nr: "{excel_column_name}": "{new_value_from_excel}"')
                            ret_val = CRUD_telnr_in_db(stammdaten_schema, pers_id, new_value_from_excel.replace(' ', ''), take_action=True, verbal=verbal_while_update)
                            # print(ret_val)
                        elif excel_column_name.startswith('IBAN'):
                            print(f'    Tel_Nr: "{excel_column_name}": "{new_value_from_excel}"')
                            ret_val = CRUD_IBAN_in_db(stammdaten_schema, pers_id, new_value_from_excel.replace(' ', ''), take_action=True, verbal=verbal_while_update)
                            # print(ret_val)
                        elif excel_column_name == 'Private_Strassen_Adresse':
                            new_value_from_excel = new_value_from_excel.replace('strasse', 'str.')
                            print(f'    Private_Strasse: "{excel_column_name}": "{new_value_from_excel}"')
                            ret_val = get_ort_id_from_ws_or_db(ws=worksheet_sheet, db=stammdaten_schema, pers_id=pers_id, title_row=title_row, row=row, take_action=True, verbal=verbal_while_update)
                            ort_id = ret_val['Private_Ort_ID']
                            set_cell_value_by_column_title(ret_val['Private_Ort_ID'], worksheet_sheet, title_row=title_row, row=row, column_name='Private_Ort_ID', take_action=True, verbal=verbal_while_update)
                            set_cell_value_by_column_title(''                       , worksheet_sheet, title_row=title_row, row=row, column_name='Private_Ort_ID', take_action=True, verbal=verbal_while_update)

                            addr_id = get_adress_id_from_db_by_strassen_adresse(db=stammdaten_schema, strassen_adresse=new_value_from_excel, ort_id=ort_id, take_action=True, verbal=verbal_while_update)
                            set_cell_value_by_column_title(addr_id, worksheet_sheet, title_row=title_row, row=row, column_name='Private_Adressen_ID', take_action=True, verbal=verbal_while_update)
                            print(addr_id)
                            print('\n')

                    else:
                        attr_type = get_db_attr_type(stammdaten_schema, table=db_table_name, attribute=db_attr_name, take_action=take_action, verbal=verbal_while_update)
                        count_of_updated_attributs += update_db_attribute(stammdaten_schema,
                                db_tbl_name=db_table_name, db_attr_name=db_attr_name, db_attr_type=attr_type['type'], db_attr_set_enum_values=attr_type['enums'],
                                id_attr_name='ID', id=pers_id,
                                new_value=new_value_from_excel, new_value_format=None,
                                take_action=take_action, verbal=verbal_while_update)
                    count_of_updated_records += 1
        row += 1

    # Save the Excel
    # --------------
    if take_action:
        if verbal:
            print(f'Close Excel:{reco_data_fn} [{reco_sheetname}]')
        workbook.save(reco_data_fn)
        workbook.close()


    print(f'''
    # Statistics
    # ==========
    Changes Found in File: {count_of_total_lines:4d}
          Inserted Reords: {count_of_new_records:4d}
          Updated Records: {count_of_updated_records:4d}
      
        Updated Attributs: {count_of_updated_attributs:4d}
    ''')
# ---------------------------------------------
# END: Functions for Common use
# ---------------------------------------------

def get_session_attibute(session, attribute_name='user_name'):
    if session and session.get(attribute_name):
        return session.get(attribute_name)
    else:
        return 'None'

def update_if_neccessary(db_connection, tbl_name, id, field_name, field_value, verbal=True, take_action=False):

    if verbal:
        print(f'''
           --> Calling update_if_neccessary(db_connection,
                                                  {tbl_name}, 
                                                  {id},
                                                  {field_name},
                                                  verbal={verbal}, 
                                                  take_action={take_action})''')
    verbal = False
    myCursor = db_connection.cursor()
    records_changed = 0
    if verbal:
        print(tbl_name, id, field_name, field_value)

    if tbl_name == 'email_adressen' or tbl_name == 'iban':
        sql_select = f'SELECT {field_name} FROM {tbl_name} WHERE id={id}'
        myCursor.execute(sql_select)
        myresult = myCursor.fetchall()
        old_value = myresult[0][0].replace(' ', '')
        if field_value != old_value:
            records_changed = 1
            print(tbl_name, '::  old: ', old_value, '   new:', field_value, '  id:', id)
            sql_update = f"UPDATE {tbl_name} SET {field_name} = '{field_value}' WHERE id={id}"
            print(sql_update, end='\n\n')
            if take_action:
                myCursor.execute(sql_update)
                db_connection.commit()
    if tbl_name == 'telefonnummern':
        sql_select = f'SELECT vorwahl,{field_name} FROM {tbl_name} WHERE id={id}'
        myCursor.execute(sql_select)
        myresult = myCursor.fetchall()
        old_nummer = myresult[0][1]
        old_nummervorwahl = myresult[0][0]
        if old_nummervorwahl + old_nummer != field_value:
            records_changed = 1
            print(tbl_name, '::  old: ', old_nummervorwahl + old_nummer, '   new:', field_value)
            if field_value == 'keine':
                sql_update = f"UPDATE {tbl_name} SET Vorwahl = '', Nummer = '{field_value}' WHERE id={id}"
            else:
                sql_update = f"UPDATE {tbl_name} SET Vorwahl = '{field_value[0:3]}', Nummer = '{field_value[3:]}' WHERE id={id}"
            print(sql_update, end='\n\n')
            if take_action:
                myCursor.execute(sql_update)
                db_connection.commit()
    return records_changed

def hash_password(password):
    # Create a new SHA-256 hash object
    sha256 = hashlib.sha256()

    # Update the hash object with the password bytes
    sha256.update(password.encode('utf-8'))

    # Get the hexadecimal representation of the hash
    hashed_password = sha256.hexdigest()

    return hashed_password

def initial_load_pachtland(db_connection, filename,  verbal=False):
    print('initial_load_pachtland...reading', filename)
    gt_landteil_count = 0
    gt_streu_count = 0
    gt_geno_landteil_count = 0
    gt_buerger_landteil_count = 0
    gt_buerger_16a_count = 0
    gt_buerger_35a_count = 0

    # Deleting left overs in DB from Landteile
    # ----------------------------------------
    myCursor = db_connection.cursor()


    sql = """DELETE FROM `landteile`;"""    # sql = """TRUNCATE `landteile`;"""  # needs user rights
    myCursor.execute(sql)
    db_connection.commit()
    print(f'{myCursor.rowcount} Landteile wurden gelöscht ')

    if verbal:
        print('--> Calling stored-proc reset_table_autoincrement_landteile...', end='')
    args = ('landteile', 'x')
    result_args = myCursor.callproc('reset_table_autoincrement_landteile', args)
    print(result_args, end='')
    if verbal:
        print('....done\n')


    # Prcessing Info-Tables in Excel and load the Landteile
    # -----------------------------------------------------
    info_tabellen_landwirte = openpyxl.load_workbook(filename, data_only=True)
    paechter_sheets = [x for x in info_tabellen_landwirte.sheetnames if "_" in x]
    # print('paechter_sheets:', paechter_sheets)

    for aPaechter_sheet_name in paechter_sheets:
        if aPaechter_sheet_name in ['Bürgerteile_Speziell']:
            print(f'Not processing {aPaechter_sheet_name:20s}', end='\n')
            continue
        else:
            # print(f'Processing       {aPaechter_sheet_name:20s}', end='\n')
            pass

        if True:    # aPaechter_sheet_name in ['Müller_Urs']:
            aPaechter_sheet = info_tabellen_landwirte[aPaechter_sheet_name]
            paechter_name = aPaechter_sheet["L3"].value
            Paechter_id = aPaechter_sheet["M3"].value
            print(f'Processing {Paechter_id:5d} {aPaechter_sheet_name:30s} {paechter_name:30s}', end='')


            row_index = 11
            landteil_count = 0
            streu_count = 0
            geno_landteil_count = 0
            buerger_landteil_count = 0
            while True:
                flurname = aPaechter_sheet["B"+str(row_index)].value
                if flurname is None:
                    break
                elif flurname == 'Total Aren:':
                    row_index += 1
                    continue
                else:
                    landteil_count += 1
                    gt_landteil_count += 1
                    gis_id = aPaechter_sheet["A" + str(row_index)].value
                    geno_id = aPaechter_sheet["C" + str(row_index)].value
                    flaeche_in_aren = 0
                    flaeche_geno = aPaechter_sheet["D" + str(row_index)].value
                    if flaeche_geno is None:
                        flaeche_bürger = aPaechter_sheet["E" + str(row_index)].value
                        if flaeche_bürger is None:
                            flaeche_in_aren = '0'
                        else:
                            flaeche_in_aren = flaeche_bürger
                    else:
                        flaeche_in_aren = flaeche_geno
                    bemerkungen = aPaechter_sheet["F" + str(row_index)].value
                    zins_pro_are = aPaechter_sheet["G" + str(row_index)].value
                    if zins_pro_are is None:
                        zins_pro_are = 0
                    fix_pacht_zins = aPaechter_sheet["H" + str(row_index)].value
                    if zins_pro_are > 0:
                        fix_pacht_zins = 0


                    verpächter_id = aPaechter_sheet["I" + str(row_index)].value
                    if verpächter_id is None:
                        verpächter_name = (aPaechter_sheet["J" + str(row_index)].value).replace(' ', '')
                        verpächter_vorname = (aPaechter_sheet["K" + str(row_index)].value)
                        verpächter_id = get_personen_id(db_connection, such_kriterien=[verpächter_name, verpächter_vorname], verbal=False)

                    vorheriger_pächter_id = aPaechter_sheet["Q" + str(row_index)].value
                    vorheriger_verpächter_id = aPaechter_sheet["R" + str(row_index)].value

                    if verpächter_id == 625:
                        if flurname == 'Streue':
                            streu_count += 1
                            gt_streu_count += 1
                        else:
                            geno_landteil_count += 1
                            gt_geno_landteil_count += 1

                    else:
                        buerger_landteil_count += 1
                        gt_buerger_landteil_count += 1
                        if flaeche_bürger == 16:
                            gt_buerger_16a_count += 1
                        elif flaeche_bürger == 35:
                            gt_buerger_35a_count += 1
                        else:
                            print(f'WARNING: Bürgerlandteil nicht 16a oder 35a!!')

                    sql = """INSERT INTO landteile (AV_Parzellen_Nr, GENO_Parzellen_Nr, Flur_Bezeichnung, Flaeche_In_Aren, Bemerkungen, Pachtzins_Pro_Are, Fix_Pachtzins, Paechter_ID, Verpaechter_ID, Vorheriger_Paechter_ID, Vorheriger_Verpaechter_ID) 
                             VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                    val = (gis_id, geno_id, flurname, flaeche_in_aren, bemerkungen, zins_pro_are, fix_pacht_zins, Paechter_id, verpächter_id, vorheriger_pächter_id, vorheriger_verpächter_id)
                    if False and geno_id in ['126.200.1', '126.200.2']:
                        print(val)
                    myCursor.execute(sql, val)
                    db_connection.commit()

                row_index += 1
            print(f'   -> {landteil_count:2d} (Geno:{geno_landteil_count:2d} + Bürger:{buerger_landteil_count:2d} + Streue:{streu_count:2d})')
    print(f'  Total:{gt_landteil_count:2d} (Geno:{gt_geno_landteil_count:2d} + Bürger:{gt_buerger_landteil_count:2d} + Streue:{gt_streu_count:2d})')
    print(f'  Detail Bürgerteile:{gt_buerger_landteil_count:2d} = 16a:{gt_buerger_16a_count:2d} + 35a:{gt_buerger_35a_count:2d})')

# -------------------------------------------
# ++++++++++++ Main Main Main +++++++++++++++
# -------------------------------------------
if __name__ == '__main__':
    connect_to_prod = True
    AUTO_TEST__split_email_detailed_long(verbal=True)



